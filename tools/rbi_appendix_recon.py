"""One-off recon: dump RBI Appendix Table 2 (Centre→States transfers, all-India)."""
from __future__ import annotations

import io
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

ROOT = Path(__file__).resolve().parents[1]
PATH = ROOT / ".runtime" / "raw" / "rbi" / "state_finances" / "02_APP_devolution_transfers.xlsx"

wb = load_workbook(PATH, data_only=True)
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"=== {sn} ===")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row)):
        vals = [c.value for c in row]
        if any(v is not None for v in vals):
            print(i, vals)
    print()
