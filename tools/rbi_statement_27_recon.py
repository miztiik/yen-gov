"""Recon for RBI State Finances Statement 27 (health expenditure share).

Fetches the Statement 27 workbook with a Chrome-style UA (RBI CDN
rejects bare clients with HTML error pages → BadZipFile downstream),
saves a byte-faithful copy to .runtime/raw/rbi/state_finances/ AND to
backend/tests/fixtures/rbi_statement_27/ (the integration-test
fixture), and dumps sheets + header rows + a sample of data rows so
we can decide whether the existing rbi_xlsx IndicatorSpec shape
covers it without further parser changes.

Usage:
    python tools/rbi_statement_27_recon.py [--no-fetch]
"""
from __future__ import annotations

import argparse
import io
import sys
from datetime import datetime, timezone
from pathlib import Path

# Force UTF-8 stdout (Windows defaults to cp1252 which chokes on ₹/em-dash).
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import httpx
from openpyxl import load_workbook


URL = (
    "https://rbidocs.rbi.org.in/rdocs/Publications/DOCs/"
    "27_ST23012026CC86B1004D0246F9A46EE80264885103.XLSX"
)
UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 yen-gov/0.1"
)


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--no-fetch", action="store_true",
                   help="Reuse the cached copy under .runtime/raw/ instead of re-fetching.")
    args = p.parse_args()

    repo_root = Path(__file__).resolve().parents[1]
    cache_dir = repo_root / ".runtime" / "raw" / "rbi" / "state_finances"
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_path = cache_dir / "27_ST23012026_HealthExpenditureShare.xlsx"

    fixture_dir = repo_root / "backend" / "tests" / "fixtures" / "rbi_statement_27"
    fixture_dir.mkdir(parents=True, exist_ok=True)
    fixture_path = fixture_dir / "27_ST23012026_HealthExpenditureShare.xlsx"

    if not args.no_fetch or not cache_path.exists():
        print(f"== fetching {URL}")
        with httpx.Client(headers={"User-Agent": UA}, timeout=60.0,
                          follow_redirects=True) as client:
            r = client.get(URL)
            r.raise_for_status()
            content = r.content
            fetched_at = datetime.now(timezone.utc).replace(microsecond=0)
        print(f"   {len(content):,} bytes, status={r.status_code}, "
              f"fetched_at={fetched_at.isoformat()}Z")
        cache_path.write_bytes(content)
        fixture_path.write_bytes(content)
        print(f"   cached:  {cache_path.relative_to(repo_root).as_posix()}")
        print(f"   fixture: {fixture_path.relative_to(repo_root).as_posix()}")
    else:
        print(f"== reusing cache {cache_path.relative_to(repo_root).as_posix()}")
        # Also ensure fixture copy exists.
        if not fixture_path.exists():
            fixture_path.write_bytes(cache_path.read_bytes())

    wb = load_workbook(cache_path, data_only=True)
    print(f"\n== sheets ({len(wb.sheetnames)}):")
    for s in wb.sheetnames:
        ws = wb[s]
        print(f"   - {s!r}  ({ws.max_row} rows x {ws.max_column} cols)")

    for s in wb.sheetnames:
        ws = wb[s]
        print(f"\n== full dump of {s!r} (up to 60 rows):")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 60:
                print(f"   … ({ws.max_row - i} more rows)")
                break
            cells = [("" if v is None else str(v)).strip()[:60] for v in row]
            while cells and not cells[-1]:
                cells.pop()
            print(f"   r{i:02d}: {cells}")

    wb.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
