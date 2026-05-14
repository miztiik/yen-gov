"""Ingest RBI Handbook of Statistics on Indian Economy (2024-25 edition) — state SDP tables.

Reads four cached XLSX files under ``.runtime/raw/rbi/handbook_economy_2024_25/``::

  T05_NSDP_Statewise_Current.xlsx   -> economy/state_nsdp_current_inr_crore
  T06_NSDP_Statewise_Constant.xlsx  -> economy/state_nsdp_constant_inr_crore
  T09_PCNSDP_Statewise_Current.xlsx -> economy/state_per_capita_nsdp_current_inr_long
  T10_PCNSDP_Statewise_Constant.xlsx-> economy/state_per_capita_nsdp_constant_inr_long

Each sheet stacks four base-year sections in the same column-1 ordering:
  (Base Year : 1993-94)   1994-95..1999-00     (current/constant; pre-2011-12 = factor cost)
  (Base Year : 1999-2000) 1999-00..2004-05
  (Base Year : 2004-05)   2004-05..2011-12
  (Base Year : 2011-12)   2011-12..2024-25

For overlapping years we keep the LATEST base (closest to today's methodology),
giving a continuous spliced series 1994-95 -> 2024-25 (31y) for NSDP, and
2000-01 -> 2024-25 (25y) for per-capita NSDP.

`series_breaks` are declared at each base transition and at 2011-12 (factor cost
-> basic prices definition change for NSDP). Per-row `vintage` records the base
year so the renderer can show it.

Run from repo root with the backend venv active::

    python tools/rbi_hbs_ingest_state_gdp.py
"""
from __future__ import annotations

import io
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

CACHE = Path(".runtime/raw/rbi/handbook_economy_2024_25")
OUT = Path("datasets/indicators/in/economy")

# Snapshot URLs for sources[]. These are the pinned 2024-25 edition URLs we
# downloaded into CACHE on 2026-05-14 (lexicographic edition stamp 29082025
# = Aug 29, 2025 = Handbook of Statistics on Indian Economy 2024-25 edition).
HBS_LANDING = "https://www.rbi.org.in/Scripts/AnnualPublications.aspx?head=Handbook+of+Statistics+on+Indian+Economy"
SNAPSHOT_URLS = {
    "T05": "https://rbidocs.rbi.org.in/rdocs/Publications/DOCs/05T_2908202556D0D1A9FA0C4615A7889EC1F025BACE.XLSX",
    "T06": "https://rbidocs.rbi.org.in/rdocs/Publications/DOCs/06T_290820258621D22235014AC19EE27C859382FEAF.XLSX",
    "T09": "https://rbidocs.rbi.org.in/rdocs/Publications/DOCs/09T_2908202582956A1D380840F5870B18841EEEF815.XLSX",
    "T10": "https://rbidocs.rbi.org.in/rdocs/Publications/DOCs/10T_290820257B1763CF72624007ABF4F4E48CC941B3.XLSX",
}
# UTC timestamp recorded on these artifacts. Reflects when our pipeline read
# the bytes (a single download session). Not the RBI publication date.
FETCHED_AT = datetime(2026, 5, 14, 18, 0, 0, tzinfo=timezone.utc).isoformat().replace("+00:00", "Z")

# State name (RBI column header) -> ECI code. RBI uses "&" not "and"; our
# states.json uses "and"; the mapping is hand-pinned.
NAME_TO_ECI = {
    "Andhra Pradesh": "S01",
    "Arunachal Pradesh": "S02",
    "Assam": "S03",
    "Bihar": "S04",
    "Chhattisgarh": "S26",
    "Goa": "S05",
    "Gujarat": "S06",
    "Haryana": "S07",
    "Himachal Pradesh": "S08",
    "Jammu & Kashmir": "U08",
    "Jharkhand": "S27",
    "Karnataka": "S10",
    "Kerala": "S11",
    "Madhya Pradesh": "S12",
    "Maharashtra": "S13",
    "Manipur": "S14",
    "Meghalaya": "S15",
    "Mizoram": "S16",
    "Nagaland": "S17",
    "Odisha": "S18",
    "Punjab": "S19",
    "Rajasthan": "S20",
    "Sikkim": "S21",
    "Tamil Nadu": "S22",
    "Telangana": "S29",
    "Tripura": "S23",
    "Uttar Pradesh": "S24",
    "Uttarakhand": "S28",
    "West Bengal": "S25",
    "Andaman & Nicobar Islands": "U01",
    "Chandigarh": "U02",
    "Delhi": "U05",
    "Puducherry": "U07",
}

# All-India aggregate column (only present in Tables 9/10). Becomes entity_id "IN".
ALL_INDIA_NAMES = {"All- India per capita NNI", "All-India per capita NNI"}


def _fy_to_time(fy_label: str) -> str:
    """Convert RBI fiscal-year label '1994-95' -> '1994-04' (April start)."""
    start = fy_label.split("-")[0]
    return f"{start}-04"


def _is_year(s: object) -> str | None:
    """Return canonical 'YYYY-YY' year label if s looks like one, else None."""
    if not isinstance(s, str):
        return None
    s = s.strip()
    if len(s) == 7 and s[4] == "-" and s[:4].isdigit() and s[5:7].isdigit():
        return s
    return None


def _is_base_marker(s: object) -> str | None:
    """Return base-year label if s is a '(Base Year : XXXX-YY)' marker."""
    if not isinstance(s, str):
        return None
    if "Base Year" not in s and "Base :" not in s:
        return None
    # Extract last token
    inside = s.replace("(", "").replace(")", "").strip()
    parts = inside.split(":")
    if len(parts) < 2:
        return None
    return parts[-1].strip()


def _coerce_value(v: object) -> float | None:
    """Convert RBI cell to float or None (treat '-', '.', '' as missing)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        v = v.strip()
        if v in ("-", ".", "", "*", "NA", "n.a.", "—"):
            return None
        try:
            return float(v.replace(",", ""))
        except ValueError:
            return None
    return None


def parse_workbook(xlsx: Path) -> dict[str, dict[str, dict[str, float]]]:
    """Return {entity_id: {time: {base: value}}} for every state column.

    Reads every sheet in the workbook (T_5, T_5(Contd.), T_5(Concld.) etc.).
    """
    out: dict[str, dict[str, dict[str, float]]] = {}
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        # Header row carries the state names. Tables 5/6 use row index 4
        # (col 1 = "Year"); Tables 9/10 use row index 3 (col 1 = "State /
        # Union Territory"). Detect by which one starts with "Year"/"State".
        header_idx = None
        for i, row in enumerate(rows[:8]):
            cell = row[1] if len(row) > 1 else None
            if isinstance(cell, str) and cell.strip() in ("Year", "State / Union Territory"):
                header_idx = i
                break
        if header_idx is None:
            continue
        header = rows[header_idx]
        # Map column index -> entity_id (skip cols not in NAME_TO_ECI / ALL_INDIA)
        col_to_entity: dict[int, str] = {}
        for ci, cell in enumerate(header):
            if not isinstance(cell, str):
                continue
            name = cell.strip()
            if name in NAME_TO_ECI:
                col_to_entity[ci] = NAME_TO_ECI[name]
            elif name in ALL_INDIA_NAMES:
                col_to_entity[ci] = "IN"

        # Walk data rows, tracking current base
        current_base: str | None = None
        for row in rows[header_idx + 1:]:
            label = row[1] if len(row) > 1 else None
            base = _is_base_marker(label)
            if base:
                current_base = base
                continue
            year = _is_year(label)
            if not year or current_base is None:
                continue
            time = _fy_to_time(year)
            for ci, eid in col_to_entity.items():
                v = _coerce_value(row[ci]) if ci < len(row) else None
                if v is None:
                    continue
                out.setdefault(eid, {}).setdefault(time, {})[current_base] = v
    wb.close()
    return out


# Preference order: most recent base wins for any given (entity, time).
# 2011-12 base values dominate 2004-05 dominates 1999-2000 dominates 1993-94.
BASE_PRIORITY = ["2011-12", "2004-05", "1999-2000", "1993-94"]


def collapse_to_long(parsed: dict[str, dict[str, dict[str, float]]]) -> list[dict]:
    """Pick the latest-base value per (entity, time) and emit long-form rows
    sorted by entity then time, each carrying its `vintage` (base year)."""
    rows: list[dict] = []
    for eid in sorted(parsed):
        for time in sorted(parsed[eid]):
            by_base = parsed[eid][time]
            chosen_base = next(
                (b for b in BASE_PRIORITY if b in by_base), None
            )
            if chosen_base is None:
                continue
            rows.append({
                "entity_id": eid,
                "time": time,
                "value": by_base[chosen_base],
                "vintage": f"Base {chosen_base}",
            })
    return rows


# --- Indicator metadata --------------------------------------------------

SERIES_BREAKS_NSDP = [
    {"at_time": "1999-04", "kind": "rebase",
     "note": "MoSPI rebase: 1993-94 -> 1999-2000 base."},
    {"at_time": "2004-04", "kind": "rebase",
     "note": "MoSPI rebase: 1999-2000 -> 2004-05 base."},
    {"at_time": "2011-04", "kind": "definition_change",
     "note": "MoSPI rebase to 2011-12 base AND switched headline aggregate from 'NSDP at factor cost' (pre-2011-12) to 'NSDP at basic prices' (2011-12 onwards). Growth rates spanning 2010-11 -> 2011-12 are not strictly comparable."},
    {"at_time": "2014-04", "kind": "coverage_change",
     "note": "Telangana carved out of Andhra Pradesh on 2 June 2014. RBI back-projects S29 (Telangana) to 2011-12 by carving from undivided AP; pre-2014-15 S29 values are MoSPI back-estimates, not original state DES figures. S01 (residual Andhra Pradesh) is reported on the post-bifurcation footprint from 2011-12 too."},
    {"at_time": "2019-04", "kind": "coverage_change",
     "note": "J&K reorganisation (Aug 2019): U08 series from 2019-20 onwards covers UT of Jammu & Kashmir only (Ladakh not separately reported in this RBI table)."},
]

SPECS = [
    {
        "id": "economy/state_nsdp_current_inr_crore",
        "title": "Net State Domestic Product (current prices)",
        "table_key": "T05",
        "xlsx": "T05_NSDP_Statewise_Current.xlsx",
        "table_label": "Table 5: Net State Domestic Product - State-wise (At Current Prices)",
        "value_kind": "currency",
        "unit": "INR (crore)",
        "icon": "trending-up",
        "description": (
            "Net State Domestic Product (NSDP) of each state and union "
            "territory, in Lakh Rupees Crore at current (nominal) prices. "
            "Long spliced series 1994-95 -> 2024-25 (31 fiscal years), "
            "stitched across MoSPI's four base-year revisions (1993-94, "
            "1999-2000, 2004-05, 2011-12). For overlapping years the most "
            "recent base is kept. Pre-2011-12 figures are NSDP 'at factor "
            "cost'; 2011-12 onwards are NSDP 'at basic prices'."
        ),
        "notes": (
            "Source: RBI Handbook of Statistics on Indian Economy 2024-25 "
            "edition, Table 5. The same workbook stacks four base-year "
            "sections; we keep the latest-base value per (state, year) and "
            "stamp the chosen base on each row's `vintage` field. Telangana "
            "(S29) starts in 2014-15 (carve-out from Andhra Pradesh). U08 "
            "(J&K) post-2019-20 is UT-only (Ladakh excluded). 2024-25 is "
            "Provisional Estimates per RBI; 2023-24 is First Revised."
        ),
    },
    {
        "id": "economy/state_nsdp_constant_inr_crore",
        "title": "Net State Domestic Product (constant prices, spliced)",
        "table_key": "T06",
        "xlsx": "T06_NSDP_Statewise_Constant.xlsx",
        "table_label": "Table 6: Net State Domestic Product - State-wise (At Constant Prices)",
        "value_kind": "currency",
        "unit": "INR (crore)",
        "icon": "trending-up",
        "description": (
            "Real Net State Domestic Product (constant prices), 1994-95 -> "
            "2024-25, spliced across MoSPI's four base years (1993-94, "
            "1999-2000, 2004-05, 2011-12). The most recent base is kept "
            "for any overlap. Use the 1993-94/1999-2000/2004-05 segments "
            "for within-base growth; growth rates that span a base "
            "transition (e.g. 2010-11 -> 2011-12) are not directly "
            "comparable. Pre-2011-12 = NSDP at factor cost; from 2011-12 "
            "= NSDP at basic prices."
        ),
        "notes": (
            "Source: RBI Handbook of Statistics on Indian Economy 2024-25 "
            "edition, Table 6. Each row's `vintage` records the base year "
            "the value was published under. See `series_breaks` for the "
            "rebase points; the renderer should refuse to compute a "
            "growth rate that crosses a break."
        ),
    },
    {
        "id": "economy/state_per_capita_nsdp_current_inr_long",
        "title": "State per-capita NSDP (current prices, long series)",
        "table_key": "T09",
        "xlsx": "T09_PCNSDP_Statewise_Current.xlsx",
        "table_label": "Table 9: Per Capita Net State Domestic Product - State-wise (At Current Prices)",
        "value_kind": "currency",
        "unit": "INR",
        "icon": "trending-up",
        "description": (
            "Per-capita Net State Domestic Product (current prices), "
            "2000-01 -> 2024-25 (25 fiscal years), spliced across MoSPI's "
            "1999-2000, 2004-05, and 2011-12 base years. Long-history "
            "sibling of `economy/state_per_capita_nsdp_current_inr` (which "
            "covers 2004-05 -> 2023-24 from ICED at the 2011-12 base "
            "only). Use this artifact when you need depth; use the ICED "
            "sibling when you need single-base purity."
        ),
        "notes": (
            "Source: RBI Handbook of Statistics on Indian Economy 2024-25 "
            "edition, Table 9. RBI's per-capita series begins in 2000-01 "
            "(no 1993-94 base section here). All-India per-capita NNI "
            "(IN) included as the national reference line."
        ),
    },
    {
        "id": "economy/state_per_capita_nsdp_constant_inr_long",
        "title": "State per-capita NSDP (constant prices, long series, spliced)",
        "table_key": "T10",
        "xlsx": "T10_PCNSDP_Statewise_Constant.xlsx",
        "table_label": "Table 10: Per Capita Net State Domestic Product - State-wise (At Constant Prices)",
        "value_kind": "currency",
        "unit": "INR",
        "icon": "trending-up",
        "description": (
            "Real per-capita NSDP, 2000-01 -> 2024-25, spliced across "
            "MoSPI's 1999-2000, 2004-05, and 2011-12 base years. The "
            "most recent base is kept for overlapping years; each row's "
            "`vintage` records which base produced the value. Long-history "
            "sibling of `economy/state_per_capita_nsdp_constant_2011_12_inr` "
            "(which is single-base 2011-12 only)."
        ),
        "notes": (
            "Source: RBI Handbook of Statistics on Indian Economy 2024-25 "
            "edition, Table 10. Pre-2011-12 figures are real per-capita "
            "NSDP at factor cost (older bases); 2011-12 onwards are at "
            "basic prices. Cross-base growth rates not strictly comparable."
        ),
    },
]


def emit(spec: dict, parsed: dict[str, dict[str, dict[str, float]]]) -> None:
    rows = collapse_to_long(parsed)
    if not rows:
        raise SystemExit(f"empty rows for {spec['id']}")
    times = sorted({r["time"] for r in rows})
    entities = sorted({r["entity_id"] for r in rows})
    art = {
        "$schema": "https://yen-gov.github.io/schemas/indicator.schema.json",
        "$schema_version": "1.3",
        "sources": [
            {
                "url": SNAPSHOT_URLS[spec["table_key"]],
                "fetched_at": FETCHED_AT,
                "name": f"RBI Handbook of Statistics on Indian Economy 2024-25 - {spec['table_label']}",
                "authority": "Reserve Bank of India (compiled from National Statistics Office, MoSPI)",
            },
            {
                "url": HBS_LANDING,
                "fetched_at": FETCHED_AT,
                "name": "RBI Handbook of Statistics on Indian Economy - landing page",
                "authority": "Reserve Bank of India",
            },
        ],
        "license": {
            "id": "RBI-publication",
            "name": "Reserve Bank of India publication (open for non-commercial use with attribution)",
            "url": "https://www.rbi.org.in/Scripts/Disclaimer.aspx",
            "redistributable": True,
        },
        "coverage": {
            "spatial": f"India (states + UTs); {len(entities)} entities",
            "temporal": f"{times[0]}..{times[-1]}",
            "admin_level": "state",
        },
        "indicator": {
            "id": spec["id"],
            "title": spec["title"],
            "description": spec["description"],
            "entity_kind": "state",
            "time_grain": "fiscal_year",
            "value_kind": spec["value_kind"],
            "direction": "higher_is_better",
            "scale_hint": "linear",
            "unit": spec["unit"],
            "icon": spec["icon"],
            "attribution_geography": "where_resident",
            "comparability": "comparable_with_normalisation",
            "implementing_authority": "state",
            "methodology_vintage": "MoSPI multi-base spliced (1993-94 / 1999-2000 / 2004-05 / 2011-12); RBI Handbook 2024-25 edition",
            "notes": spec["notes"],
            "series_breaks": SERIES_BREAKS_NSDP,
        },
        "rows": rows,
    }
    out_path = OUT / f"{spec['id'].split('/')[-1]}.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(art, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"  wrote {out_path}  rows={len(rows)} entities={len(entities)} span={times[0]}..{times[-1]}")


def main() -> None:
    for spec in SPECS:
        xlsx = CACHE / spec["xlsx"]
        print(f"\n=== {spec['id']} <- {spec['xlsx']} ===")
        parsed = parse_workbook(xlsx)
        print(f"  parsed entities: {len(parsed)}")
        emit(spec, parsed)


if __name__ == "__main__":
    main()
