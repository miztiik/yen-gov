"""Inspect RBI HBS-IE Table 89 layout."""
from __future__ import annotations

import io
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

path = Path(".runtime/raw/rbi/hbs_ie/T89_KeyDeficitIndicators_Centre_2025.xlsx")
wb = load_workbook(path, data_only=True)
print(f"sheets: {wb.sheetnames}")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n--- sheet {name!r}  dims: rows={ws.max_row} cols={ws.max_column} ---")
    for r in range(1, min(ws.max_row, 70) + 1):
        cells = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                cells.append("·")
            else:
                cells.append(str(v).replace("\n", " ")[:30])
        print(f"  row{r:>3}: " + " | ".join(cells))
