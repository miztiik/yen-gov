"""Quick recon of RBI Handbook of Statistics on Indian States — SDP tables 19–24.

Run from repo root:
    python tools/rbi_handbook_states_inspect.py

Dumps sheet names, header rows, and a sample of state rows so we can
write the parser against the real workbook layout.
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import sys as _sys
CACHE = Path(_sys.argv[1] if len(_sys.argv) > 1 else ".runtime/raw/rbi/handbook_states_2024_25")

for xlsx in sorted(CACHE.glob("*.xlsx")):
    print(f"\n=== {xlsx.name} ===")
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"  sheet: {sheet_name}  dims={ws.calculate_dimension()}")
        rows = list(ws.iter_rows(values_only=True))
        for i, row in enumerate(rows[:8]):
            print(f"   row {i}: {[c for c in row[:14]]}")
        print(f"   ... total rows: {len(rows)}")
        # show last 5 rows to see footnotes
        for i, row in enumerate(rows[-5:], start=len(rows) - 5):
            print(f"   row {i}: {[c for c in row[:14]]}")
    wb.close()
