"""Inspect ECI Statistical Report Section 3 (List of Political Parties Participated).

Question: does it carry the numeric eci_code that we currently get by
cross-fetching the live-results partywise page?

Run:
    python tools/eci_recon/inspect_section3.py <path-to-xlsx>
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")


def main(path: Path) -> None:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    print(f"Workbook: {path}")
    print(f"Sheets:   {wb.sheetnames}")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\n=== Sheet: {name} ===")
        rows = list(ws.iter_rows(min_row=1, max_row=25, values_only=True))
        for i, row in enumerate(rows, start=1):
            cells = ["" if v is None else str(v) for v in row]
            # Strip trailing empties for readability
            while cells and not cells[-1].strip():
                cells.pop()
            if cells:
                print(f"  R{i:02d}: {cells}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("usage: inspect_section3.py <xlsx>", file=sys.stderr)
        sys.exit(2)
    main(Path(sys.argv[1]))
