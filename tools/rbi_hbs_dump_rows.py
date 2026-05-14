"""Dump full header row + a sample data row to confirm column layout."""
from __future__ import annotations
import io, sys
from pathlib import Path
import openpyxl
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
xlsx = Path(sys.argv[1])
wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"\n=== {xlsx.name} :: {sname} ===")
    rows = list(ws.iter_rows(values_only=True))
    for i in (3, 4, 5, 6, 7, 30, 31):
        if i < len(rows):
            print(f" row{i}: {rows[i]}")
wb.close()
