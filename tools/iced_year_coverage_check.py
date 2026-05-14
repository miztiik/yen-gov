"""Check year-coverage end-dates on the dup-looking files so we know if they
extend (MERGE) or are subsumed (SKIP-DUP) by existing artifacts."""
from __future__ import annotations
import io
import json
import sys
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

ROOT = Path(__file__).resolve().parents[1]
EPH = ROOT / "datasets" / "ephemeral_datasets"
IND = ROOT / "datasets" / "indicators" / "in"


def years_in_xlsx_row0(p: Path) -> list[str]:
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    r0 = next(ws.iter_rows(values_only=True))
    return [str(c) for c in r0 if c is not None]


def years_in_xlsx_col1(p: Path) -> set[str]:
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    next(rows)  # header
    out: set[str] = set()
    for r in rows:
        if len(r) > 1 and r[1] is not None:
            out.add(str(r[1]))
    return out


def indicator_years(p: Path) -> set[str]:
    data = json.loads(p.read_text(encoding="utf-8"))
    return {str(row.get("time", row.get("year"))) for row in data.get("data", []) if row.get("time") or row.get("year")}


def section(name: str) -> None:
    print(f"\n--- {name} ---")


# Wide files (year columns)
section("Electricity_Installed_Capacity_*.xlsx — wide (years across columns)")
for p in sorted(EPH.glob("Electricity_Installed_Capacity_*.xlsx")):
    yrs = years_in_xlsx_row0(p)
    print(f"  {p.name}  first/last year: {yrs[1]!r} ... {yrs[-1]!r}  ({len(yrs)-1} year cols)")

section("Net_Capacity_Addition_*.xlsx — wide")
for p in sorted(EPH.glob("Net_Capacity_Addition_*.xlsx")):
    yrs = years_in_xlsx_row0(p)
    print(f"  {p.name}  first/last year: {yrs[1]!r} ... {yrs[-1]!r}  ({len(yrs)-1} year cols)")

# Long files (year column index 1)
section("Per_Capita_Consumption / Per_Capita_Income — long")
for name in ["Per_Capita_Consumption_1778791816369.xlsx", "Per_Capita_Income_1778791727578.xlsx"]:
    p = EPH / name
    yrs = sorted(years_in_xlsx_col1(p))
    print(f"  {p.name}  years: {yrs[0]} ... {yrs[-1]}  ({len(yrs)} unique)")

section("Retired_Thermal_Plants — long, year col 0")
for p in sorted(EPH.glob("Retired_Thermal_Plants_*.xlsx")):
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    yrs = sorted({str(r[0]) for r in rows[1:] if r and r[0] is not None})
    has_state = "state" in str(rows[0]).lower() if rows else False
    nonblank_states = sum(1 for r in rows[1:] if has_state and len(r) > 1 and r[1])
    print(f"  {p.name}  years {yrs[0]}..{yrs[-1]} ({len(yrs)})  rows={len(rows)-1}  has_state_col={has_state}  nonblank_state_rows={nonblank_states}")

# Existing indicators we suspect overlap
section("EXISTING: state_per_capita_electricity_consumption_kwh.json")
p = IND / "energy" / "state_per_capita_electricity_consumption_kwh.json"
if p.exists():
    yrs = sorted(indicator_years(p))
    print(f"  years: {yrs[0]} ... {yrs[-1]}  ({len(yrs)})")

section("EXISTING: state_per_capita_nsdp_current_inr.json")
p = IND / "economy" / "state_per_capita_nsdp_current_inr.json"
if p.exists():
    yrs = sorted(indicator_years(p))
    print(f"  years: {yrs[0]} ... {yrs[-1]}  ({len(yrs)})")

section("EXISTING: installed_capacity_by_source_mw.json")
p = IND / "energy" / "installed_capacity_by_source_mw.json"
if p.exists():
    yrs = sorted(indicator_years(p))
    print(f"  years: {yrs[0]} ... {yrs[-1]}  ({len(yrs)})")
