"""Recon ECI 2023 Statistical Report XLSX shape vs the 2024+ adapter.

For each of the four 2023 states with new-portal landing pages, downloads
Section 10 (Detailed_Results) and Section 3 (List_Of_Political_Parties_
Participated) XLSX files, then dumps the first ~20 rows of each so we can
diff column layout against the 2024+ assumption baked into:

  - sources/eci/statistical_report_detailed.py  (Section 10 parser)
  - sources/eci/section3.py                     (Section 3 parser)

Read-only. Lives in tools/ per CLAUDE.md §3 — adapter belongs under
backend/sources/eci/ once we know the shape matches.
"""

from __future__ import annotations

import io
import sys
import urllib.request

from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"
)

BASE = "https://www.eci.gov.in/eci-backend/public/all_files/full-statistical-reports"

# (state_label, slug-on-eci) — the 4 cohorts of interest.
STATES = [
    ("Madhya Pradesh", "mp"),
    ("Chhattisgarh", "chhattisgarh"),
    ("Mizoram", "mizoram"),
    ("Telangana", "telangana"),
]

STATEMENTS = [
    ("Section 10", "Detailed_Results"),
    ("Section 3", "List_Of_Political_Parties_Participated"),
]


BROWSER_HEADERS = {
    "User-Agent": UA,
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "en-GB,en-US;q=0.9,en;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
    "Referer": "https://www.eci.gov.in/mp-legislative-election-2023-statistical-report",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "same-origin",
    "Sec-Fetch-User": "?1",
    "Upgrade-Insecure-Requests": "1",
}


def fetch(url: str) -> bytes:
    # urllib doesn't auto-decompress br/gzip; drop Accept-Encoding so server
    # returns identity. Keep the rest of the browser fingerprint.
    headers = {k: v for k, v in BROWSER_HEADERS.items() if k != "Accept-Encoding"}
    req = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(req, timeout=60) as resp:
        return resp.read()


def dump_xlsx(label: str, content: bytes, max_rows: int = 25) -> None:
    print(f"--- {label} ({len(content):,} bytes) ---")
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    print(f"sheets: {wb.sheetnames}")
    ws = wb[wb.sheetnames[0]]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        cells = [
            str(c)[:35] if c is not None else "" for c in row[:16]
        ]
        print(f"  r{i:>2}: {cells}")
    wb.close()


def main() -> None:
    for state_label, slug in STATES:
        print(f"\n===== {state_label} ({slug}) =====")
        for label, fname in STATEMENTS:
            url = f"{BASE}/{slug}/2023/{fname}.xlsx"
            print(f"\n# {label}: {url}")
            try:
                content = fetch(url)
            except Exception as exc:
                print(f"  FETCH FAILED: {exc!r}")
                continue
            dump_xlsx(label, content)


if __name__ == "__main__":
    main()
