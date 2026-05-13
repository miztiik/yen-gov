"""Inspect AppT1 Major Deficit Indicators workbook layout.

The pre-cached file ``AppT1_MajorDeficitIndicators_2026.xlsx`` is
documented in the long-coverage ledger as having a "two-row-per-year
layout" (₹ Crore + % GDP interleaved) which differs from the App T2
single-row layout already handled by ``rbi_appendix_national``.

This script dumps every sheet's first ~40 rows × first ~12 columns so
we can pick the right ``AppendixSpec`` shape and column-pinning rule.
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

PATH = Path(".runtime/raw/rbi/state_finances/AppT1_MajorDeficitIndicators_2026.xlsx")


def _fmt(v: object) -> str:
    if v is None:
        return ""
    s = str(v).replace("\n", " | ").strip()
    return s[:50]


def main() -> None:
    wb = load_workbook(PATH, read_only=True, data_only=True)
    print(f"# {PATH.name} — {wb.sheetnames}")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n## Sheet: {sheet_name}  ({ws.max_row} rows × {ws.max_column} cols)\n")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 40:
                print("    ... (truncated)")
                break
            cells = [_fmt(c) for c in row[:12]]
            print(f"  r{i:02d}: " + " | ".join(cells))


if __name__ == "__main__":
    main()
