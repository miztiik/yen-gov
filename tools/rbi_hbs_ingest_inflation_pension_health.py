"""Ingest RBI Handbook tables (HBS-IS 2024-25) for state pension indicators.

Emits artifacts under datasets/indicators/in/{prices,fiscal}/. The 5 health
shards previously emitted under datasets/indicators/in/health/ were retired
in PR-D6 (grain-rip plan §D6); the 3 national inflation shards (WPI / CPI-IW /
CPI-Combined) previously emitted under datasets/indicators/in/prices/ from
HBS-IE Tables 36 + 37 were retired in PR-D1 (grain-rip plan §D1) — no
canonical national-inflation successor is planned; state-level CPI lives in
the post-B8 facetted shard ``prices/cpi_inflation_pct``. See
docs/reference/topics/prices.md for the retirement note.

Shared building blocks (state-name map, value coercion, year-label parsing,
landing-page URLs, license block, write helper) live in
``backend/yen_gov/sources/rbi_hbs/``.
"""
from __future__ import annotations

import json  # noqa: F401  -- kept for spec-table introspection during dev runs
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

from yen_gov.sources.rbi_hbs import (
    HBS_IS_LANDING,
    LICENSE_RBI,
    NAME_TO_ECI,
    coerce_value as _coerce,
    setup_utf8_stdout,
    write_artifact,
    year_label_to_time as _year_label_to_time,
)

setup_utf8_stdout()

STATES_CACHE = Path(".runtime/raw/rbi/handbook_states_2024_25")
OUT = Path("datasets/indicators/in")

FETCHED_AT = datetime(2026, 5, 14, 19, 0, 0, tzinfo=timezone.utc).isoformat().replace("+00:00", "Z")


# ---------------------------------------------------------------------------
# Pattern B — State × year (rows = state, cols = years), one or two sheets.
# ---------------------------------------------------------------------------


def parse_state_year_table(
    xlsx: Path, header_row: int = 3, calendar: bool = False
) -> list[dict]:
    """Walk every sheet. Header row N is at index ``header_row`` (0-based);
    cells in cols 2+ are year labels. Subsequent rows have state name in col 1
    and value cells in cols 2+. Returns long-form ``[{entity_id, time, value}]``.

    For tables with two range-split sheets (e.g. T_171(i) + T_171(ii)), the
    same state appears in each sheet under disjoint year ranges; we union them
    in one output stream.
    """
    out: list[dict] = []
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    seen: set[tuple[str, str]] = set()  # de-dup if a year repeats across sheets
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        if header_row >= len(rows):
            continue
        header = rows[header_row]
        col_to_time: dict[int, str] = {}
        for ci, cell in enumerate(header):
            if ci < 2:
                continue
            t = _year_label_to_time(cell, calendar=calendar)
            if t:
                col_to_time[ci] = t
        if not col_to_time:
            continue
        for row in rows[header_row + 1:]:
            label = row[1] if len(row) > 1 else None
            if not isinstance(label, str):
                continue
            name = label.strip()
            eid = NAME_TO_ECI.get(name)
            if eid is None:
                continue
            for ci, time in col_to_time.items():
                v = _coerce(row[ci]) if ci < len(row) else None
                if v is None:
                    continue
                key = (eid, time)
                if key in seen:
                    continue
                seen.add(key)
                out.append({"entity_id": eid, "time": time, "value": v})
    wb.close()
    out.sort(key=lambda r: (r["entity_id"], r["time"]))
    return out


# ---------------------------------------------------------------------------
# Spec table — one entry per artifact
# ---------------------------------------------------------------------------

SOURCE_RBI = "Reserve Bank of India (compiled from Office of the Economic Adviser, MoCI; or Labour Bureau, MoLE; per table)"

# --- State indicator specs (driven by parse_state_year_table) ---

STATE_SPECS = [
    # ---- State CPI inflation (T108-T111) ----
    {
        "out_path": "prices/state_cpi_general_inflation_pct.json",
        "id": "prices/state_cpi_general_inflation_pct",
        "title": "State-wise CPI inflation (General) — annual average",
        "xlsx": STATES_CACHE / "T108_StateCpiGeneral.xlsx",
        "table_label": "Table 108: State-wise Average Inflation (CPI) - General",
        "snapshot_url": "https://rbidocs.rbi.org.in/rdocs/Publications/DOCs/108T_111220251942D16B3BED4D73BE851D15D2329168.XLSX",
        "landing": HBS_IS_LANDING,
        "header_row": 3,
        "calendar": False,
        "value_kind": "rate",
        "unit": "% YoY",
        "time_grain": "fiscal_year",
        "direction": "neutral",
        "description": (
            "State-wise headline CPI inflation (General sub-index), year-on-year %, "
            "annual average per fiscal year. Sub-national sibling of the national "
            "`prices/national_cpi_combined_index_annual`. Tamil Nadu inflation can "
            "diverge meaningfully from Bihar inflation in any given year because of "
            "local food, fuel, and housing dynamics — citizens experience their own "
            "state's number, not the national average."
        ),
        "notes": "Source: RBI Handbook of Statistics on Indian States 2024-25 edition, Table 108. RBI publishes already as YoY % inflation (not index level), so no further computation needed.",
    },
    {
        "out_path": "prices/state_cpi_food_inflation_pct.json",
        "id": "prices/state_cpi_food_inflation_pct",
        "title": "State-wise CPI inflation (Food and Beverages)",
        "xlsx": STATES_CACHE / "T109_StateCpiFood.xlsx",
        "table_label": "Table 109: State-wise Average Inflation (CPI) - Food and Beverages",
        "snapshot_url": "https://rbidocs.rbi.org.in/rdocs/Publications/DOCs/109T_111220250E067E49350B48659E5A50C3D357BB46.XLSX",
        "landing": HBS_IS_LANDING,
        "header_row": 3,
        "calendar": False,
        "value_kind": "rate",
        "unit": "% YoY",
        "time_grain": "fiscal_year",
        "direction": "neutral",
        "description": (
            "State-wise CPI inflation in the Food and Beverages sub-basket — the "
            "single biggest household expenditure category (~46% of CPI-Combined). "
            "Food inflation is what citizens actually feel; it is also the most "
            "monsoon- and global-commodity-shock-driven sub-index, so swings here "
            "are largely supply-side, not policy-attributable."
        ),
        "notes": "Source: RBI Handbook of Statistics on Indian States 2024-25 edition, Table 109.",
    },
    {
        "out_path": "prices/state_cpi_fuel_inflation_pct.json",
        "id": "prices/state_cpi_fuel_inflation_pct",
        "title": "State-wise CPI inflation (Fuel and Light)",
        "xlsx": STATES_CACHE / "T110_StateCpiFuel.xlsx",
        "table_label": "Table 110: State-wise Average Inflation (CPI) - Fuel and Light",
        "snapshot_url": "https://rbidocs.rbi.org.in/rdocs/Publications/DOCs/110T_11122025784C1BDC2482496B9E25DA1BA4B38A3D.XLSX",
        "landing": HBS_IS_LANDING,
        "header_row": 3,
        "calendar": False,
        "value_kind": "rate",
        "unit": "% YoY",
        "time_grain": "fiscal_year",
        "direction": "neutral",
        "description": (
            "State-wise CPI inflation in the Fuel and Light sub-basket (LPG, kerosene, "
            "electricity tariffs, firewood). Driven by a mix of central petroleum "
            "policy, state electricity-tariff orders, and global crude. State variation "
            "reflects state-specific subsidy regimes (e.g. free electricity up to N "
            "units) and ESC-tariff slabs as much as supply shocks."
        ),
        "notes": "Source: RBI Handbook of Statistics on Indian States 2024-25 edition, Table 110.",
    },
    {
        "out_path": "prices/state_cpi_housing_urban_inflation_pct.json",
        "id": "prices/state_cpi_housing_urban_inflation_pct",
        "title": "State-wise CPI inflation (Housing — Urban only)",
        "xlsx": STATES_CACHE / "T111_StateCpiHousingUrban.xlsx",
        "table_label": "Table 111: State-wise Average Inflation (CPI) - Housing (Urban)",
        "snapshot_url": "https://rbidocs.rbi.org.in/rdocs/Publications/DOCs/111T_111220252F7D9704AB664C35A58770BAC73518DE.XLSX",
        "landing": HBS_IS_LANDING,
        "header_row": 3,
        "calendar": False,
        "value_kind": "rate",
        "unit": "% YoY",
        "time_grain": "fiscal_year",
        "direction": "neutral",
        "description": (
            "State-wise CPI Housing inflation, URBAN ONLY — NSO does not publish a "
            "rural housing CPI because rural housing in the CPI methodology is "
            "imputed from owner-occupied dwellings differently. Use this as the "
            "rent / urban housing-cost signal; it is structurally smoother than "
            "Food or Fuel and reflects mostly base-rent revisions in the surveyed "
            "centres."
        ),
        "notes": "Source: RBI Handbook of Statistics on Indian States 2024-25 edition, Table 111. Coverage is urban centres only by methodology — not a yen-gov omission.",
    },
    # ---- State pension expenditure (T171) ----
    {
        "out_path": "fiscal/state_pension_expenditure_inr_crore.json",
        "id": "fiscal/state_pension_expenditure_inr_crore",
        "title": "State pension expenditure (revenue account)",
        "xlsx": STATES_CACHE / "T171_StatePension.xlsx",
        "table_label": "Table 171: State-wise Pension",
        "snapshot_url": "https://rbidocs.rbi.org.in/rdocs/Publications/DOCs/171T_1112202558897643693142228AC6C769081FB16B.XLSX",
        "landing": HBS_IS_LANDING,
        "header_row": 3,
        "calendar": False,
        "value_kind": "currency",
        "unit": "INR (crore)",
        "time_grain": "fiscal_year",
        "direction": "neutral",
        "description": (
            "Annual state-government pension expenditure (revenue account, ₹ Crore) "
            "from FY 2004-05 onwards. Covers retirement and family pensions paid to "
            "state-government employees and pre-NPS hires; does NOT include the "
            "centrally-sponsored social-pension schemes (IGNOAPS, IGNWPS) or the "
            "National Pension System (NPS) contribution flows. Rapidly-growing line "
            "item in most state budgets — relevant to fiscal-sustainability and the "
            "Old Pension Scheme (OPS) restoration debate."
        ),
        "notes": (
            "Source: RBI Handbook of Statistics on Indian States 2024-25 edition, "
            "Table 171. Suffix codes in column headers — '(A)' = Actuals, '(RE)' = "
            "Revised Estimates, '(BE)' = Budget Estimates — are stripped at parse "
            "time but the underlying revision tier should be considered when "
            "comparing FY24/FY25 with earlier years."
        ),
    },
]


def _build_artifact(spec: dict, rows: list[dict], series_breaks: list[dict] | None = None) -> dict:
    times = sorted({r["time"] for r in rows})
    entities = sorted({r["entity_id"] for r in rows})
    indicator = {
        "id": spec["id"],
        "title": spec["title"],
        "description": spec["description"],
        "entity_kind": "country" if entities == ["IN"] else "state",
        "time_grain": spec["time_grain"],
        "value_kind": spec["value_kind"],
        "direction": spec.get("direction", "neutral"),
        "scale_hint": "linear",
        "unit": spec["unit"],
        "icon": spec.get("icon", "trending-up"),
        "attribution_geography": spec.get("attribution_geography", "where_resident"),
        "comparability": spec.get("comparability", "comparable_with_normalisation"),
        "implementing_authority": spec.get("implementing_authority", "centre"),
        "methodology_vintage": spec.get("methodology_vintage", "RBI Handbook 2024-25 edition"),
        "notes": spec["notes"],
    }
    if series_breaks:
        indicator["series_breaks"] = series_breaks
    return {
        "$schema": "https://yen-gov.github.io/schemas/indicator.schema.json",
        "$schema_version": "1.3",
        "sources": [
            {
                "url": spec["snapshot_url"],
                "fetched_at": FETCHED_AT,
                "name": f"RBI Handbook 2024-25 — {spec['table_label']}",
                "authority": SOURCE_RBI,
            },
            {
                "url": spec["landing"],
                "fetched_at": FETCHED_AT,
                "name": "RBI Handbook landing page",
                "authority": "Reserve Bank of India",
            },
        ],
        "license": LICENSE_RBI,
        "coverage": {
            "spatial": "India" if entities == ["IN"] else f"India (states + UTs); {len(entities)} entities",
            "temporal": f"{times[0]}..{times[-1]}",
            "admin_level": "country" if entities == ["IN"] else "state",
        },
        "indicator": indicator,
        "rows": rows,
    }


def _write(spec: dict, art: dict) -> None:
    write_artifact(OUT / spec["out_path"], art)


def main() -> None:
    print("\n=== State indicators (HBS-IS) ===")
    for spec in STATE_SPECS:
        rows = parse_state_year_table(spec["xlsx"], spec["header_row"], spec["calendar"])
        if not rows:
            print(f"  WARN no rows for {spec['id']}")
            continue
        art = _build_artifact(spec, rows)
        _write(spec, art)


if __name__ == "__main__":
    main()
