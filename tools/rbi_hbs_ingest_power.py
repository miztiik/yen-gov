"""Ingest RBI Handbook of Statistics on Indian States (HBS-IS) 2024-25 Power
section — Tables 138, 139, 140, 141, 142, 143.

Emits 7 state-level artifacts under datasets/indicators/in/energy/:

  Pattern A (states-as-rows × FY-as-cols, two-sheet split, FY05-FY25, 21y):
    - state_per_capita_availability_kwh.json  (T138)
    - state_power_availability_mu.json        (T139)
    - state_installed_capacity_total_mw.json  (T140)
    - state_power_requirement_mu.json         (T141)

  Pattern B (multi-period grouped sheets, 2 FY per sheet × 6 sheets, FY14-FY25, 12y):
    - state_peak_demand_mw.json  (T142, Peak Demand sub-column)
    - state_peak_met_mw.json     (T142, Peak Met sub-column)
    (Deficit MW + % are derivable: peak_met - peak_demand.)

  Pattern C (states-as-rows × calendar-year-as-cols, two-sheet split, 2007-2024, 18y):
    - state_renewable_grid_capacity_mw.json   (T143)

T143 covers ALL renewables (wind + solar + small hydro + bio + waste-to-energy)
combined — RBI does not publish a per-source split in HBS-IS. The closest RBI
gets to a fuel-mix view of state-level capacity. Originating data: MoSPI
Energy Statistics (per the workbook footer).

Shared building blocks (state-name map, value coercion, year-label parsing,
landing-page URLs, license block, write helper) live in
``backend/yen_gov/sources/rbi_hbs/``; this script imports them rather than
duplicating them inline.
"""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

import openpyxl

from yen_gov.sources.rbi_hbs import (
    HBS_IS_LANDING,
    LICENSE_RBI,
    NAME_TO_ECI,
    coerce_value as _coerce,
    cy_label_to_time as _cy_label_to_time,
    fy_label_to_time as _fy_label_to_time,
    setup_utf8_stdout,
    write_artifact,
)

setup_utf8_stdout()

STATES_CACHE = Path(".runtime/raw/rbi/handbook_states_2024_25")
OUT = Path("datasets/indicators/in")

FETCHED_AT = datetime(2026, 5, 14, 19, 30, 0, tzinfo=timezone.utc).isoformat().replace("+00:00", "Z")


def parse_state_year_table(xlsx: Path, header_row: int = 3, calendar: bool = False) -> list[dict]:
    """Walk every sheet. Header row is at index ``header_row`` (0-based);
    cells in cols 2+ are FY labels like '2004-05' (or calendar-year ints when
    ``calendar=True``). Subsequent rows have the state name in col 1. Two-sheet
    splits are unioned automatically. 'Total' / 'Others' / footer rows whose
    label is not in NAME_TO_ECI are silently skipped.
    """
    out: list[dict] = []
    seen: set[tuple[str, str]] = set()
    label_to_time = _cy_label_to_time if calendar else _fy_label_to_time
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        if header_row >= len(rows):
            continue
        header = rows[header_row]
        col_to_time: dict[int, str] = {}
        for ci, cell in enumerate(header):
            if ci < 2:
                continue
            t = label_to_time(cell)
            if t:
                col_to_time[ci] = t
        if not col_to_time:
            continue
        for row in rows[header_row + 1:]:
            label = row[1] if len(row) > 1 else None
            if not isinstance(label, str):
                continue
            name = label.strip()
            eid = NAME_TO_ECI.get(name)
            if eid is None:
                continue
            for ci, time in col_to_time.items():
                v = _coerce(row[ci]) if ci < len(row) else None
                if v is None:
                    continue
                key = (eid, time)
                if key in seen:
                    continue
                seen.add(key)
                out.append({"entity_id": eid, "time": time, "value": v})
    wb.close()
    out.sort(key=lambda r: (r["entity_id"], r["time"]))
    return out


def parse_t142_peak(xlsx: Path, sub_col_offset: int) -> list[dict]:
    """T142 has a multi-period grouped layout — each sheet covers 2 fiscal
    years, with each FY occupying 4 columns: Peak Demand MW, Peak Met MW,
    Surplus/Deficit MW, Surplus/Deficit %. Six sheets cover FY13-14..FY24-25.

    Layout per sheet (0-indexed rows):
      row 2: '2013-14' in col 2, '2014-15' in col 6   (FY-period header)
      row 3: 'Peak Demand', 'Peak Met', 'Surplus/Deficit', None, repeat
      row 4: units '(Megawatt)' x4, repeat
      row 5+: state name in col 1, then 4 cols per FY-period

    ``sub_col_offset`` selects which sub-column to extract within each FY block:
      0 = Peak Demand, 1 = Peak Met, 2 = Surplus/Deficit (MW).
    """
    out: list[dict] = []
    seen: set[tuple[str, str]] = set()
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 6:
            continue
        # Period header is row index 2; FY labels at cols 2 and 6 (the start
        # of each 4-col block). Walk col-by-col looking for FY labels.
        period_row = rows[2]
        period_starts: list[tuple[int, str]] = []
        for ci, cell in enumerate(period_row):
            t = _fy_label_to_time(cell)
            if t:
                period_starts.append((ci, t))
        if not period_starts:
            continue
        for row in rows[5:]:
            label = row[1] if len(row) > 1 else None
            if not isinstance(label, str):
                continue
            name = label.strip()
            eid = NAME_TO_ECI.get(name)
            if eid is None:
                continue
            for start_col, time in period_starts:
                target = start_col + sub_col_offset
                v = _coerce(row[target]) if target < len(row) else None
                if v is None:
                    continue
                key = (eid, time)
                if key in seen:
                    continue
                seen.add(key)
                out.append({"entity_id": eid, "time": time, "value": v})
    wb.close()
    out.sort(key=lambda r: (r["entity_id"], r["time"]))
    return out


SOURCE_RBI = "Reserve Bank of India (compiled from Central Electricity Authority, Ministry of Power)"

SPECS = [
    {
        "out_path": "energy/state_per_capita_availability_kwh.json",
        "id": "energy/state_per_capita_availability_kwh",
        "title": "State-wise per-capita availability of power",
        "xlsx": STATES_CACHE / "T138_PerCapitaAvailabilityOfPower.xlsx",
        "table_label": "Table 138: State-wise Per Capita Availability of Power",
        "snapshot_url": "https://rbidocs.rbi.org.in/rdocs/Publications/DOCs/138T_111220252D21BC51E5A842E4B25E954391C10A41.XLSX",
        "value_kind": "raw",
        "unit": "kWh per person per year",
        "icon": "zap",
        "direction": "higher_is_better",
        "description": (
            "Annual per-capita electricity availability (kWh / person / year), by "
            "state and Union Territory, fiscal year. Computed by CEA as state "
            "energy availability (MU) ÷ state mid-year population. The most "
            "citizen-relevant single number for 'how much power do people in this "
            "state actually get to use' — captures both supply expansion and "
            "transmission losses. The all-India figure roughly tripled from "
            "~600 kWh in FY05 to ~1300 kWh in FY24, but the inter-state spread "
            "is wide (Goa / Punjab / Gujarat above 2,000; Bihar / Manipur / "
            "Nagaland still below 400)."
        ),
        "notes": (
            "Source: RBI Handbook of Statistics on Indian States 2024-25 edition, "
            "Table 138. Originating data: Central Electricity Authority, Ministry "
            "of Power. Per-capita is on geographical-area population (resident), "
            "not consumption-weighted — large industrial-export states (e.g. "
            "Chhattisgarh, Sikkim) post inflated numbers because the power they "
            "generate is consumed elsewhere."
        ),
    },
    {
        "out_path": "energy/state_power_availability_mu.json",
        "id": "energy/state_power_availability_mu",
        "title": "State-wise availability of power (MU)",
        "xlsx": STATES_CACHE / "T139_AvailabilityOfPower.xlsx",
        "table_label": "Table 139: State-wise Availability of Power",
        "snapshot_url": "https://rbidocs.rbi.org.in/rdocs/Publications/DOCs/139T_111220256569FF1C11E248B7A8ED64D31C999776.XLSX",
        "value_kind": "raw",
        "unit": "MU (million units = GWh)",
        "icon": "zap",
        "direction": "higher_is_better",
        "description": (
            "Annual availability of energy in the state (MU = million units = GWh), "
            "fiscal year. 'Availability' is the energy actually supplied to the "
            "state grid after accounting for inter-state imports / exports and "
            "transmission losses — i.e. what was put on the bus for distribution "
            "companies to sell to consumers. Read alongside Power Requirement "
            "(T141) — the gap between requirement and availability is the "
            "'energy not supplied' (deficit), which CEA tracks separately."
        ),
        "notes": (
            "Source: RBI Handbook of Statistics on Indian States 2024-25 edition, "
            "Table 139. Originating data: Central Electricity Authority. Note that "
            "'availability' is a transmission-side metric, not consumption — "
            "for actual electricity sold to end users see ICED-derived "
            "`energy/state_electricity_sales_mu.json`."
        ),
    },
    {
        "out_path": "energy/state_installed_capacity_total_mw.json",
        "id": "energy/state_installed_capacity_total_mw",
        "title": "State-wise installed capacity of power (MW)",
        "xlsx": STATES_CACHE / "T140_InstalledCapacityOfPower.xlsx",
        "table_label": "Table 140: State-wise Installed Capacity of Power",
        "snapshot_url": "https://rbidocs.rbi.org.in/rdocs/Publications/DOCs/140T_111220254D8DA0B69B444492B6E9BAF30F3395C8.XLSX",
        "value_kind": "raw",
        "unit": "MW",
        "icon": "zap",
        "direction": "higher_is_better",
        "description": (
            "Total installed electricity-generation capacity located in the state "
            "(MW), end-of-fiscal-year. Aggregate of all fuel sources (coal, gas, "
            "diesel, nuclear, hydro, renewable). For the state × source breakdown "
            "RBI's Handbook does not publish — see CEA monthly Executive Summary "
            "for that. This RBI series gives the deepest continuous all-source "
            "total: 21 fiscal years (FY05 → FY25) versus 10 years from the "
            "ICED-derived sibling `energy/state_installed_capacity_geographical_mw`. "
            "Prefer this RBI series for long-history charts; ICED for the most "
            "recent year's sub-aggregates."
        ),
        "notes": (
            "Source: RBI Handbook of Statistics on Indian States 2024-25 edition, "
            "Table 140. Originating data: Central Electricity Authority. "
            "Capacity is geographically located — power-trading states "
            "(e.g. Chhattisgarh exports much of its thermal output) appear "
            "with capacity disproportionate to their consumption. Compare "
            "with `state_per_capita_availability_kwh` for the consumption-side "
            "view."
        ),
    },
    {
        "out_path": "energy/state_power_requirement_mu.json",
        "id": "energy/state_power_requirement_mu",
        "title": "State-wise power requirement (MU)",
        "xlsx": STATES_CACHE / "T141_PowerRequirement.xlsx",
        "table_label": "Table 141: State-wise Power Requirement",
        "snapshot_url": "https://rbidocs.rbi.org.in/rdocs/Publications/DOCs/141T_11122025615E98A72EA6475CAACFFD12FFC83057.XLSX",
        "value_kind": "raw",
        "unit": "MU (million units = GWh)",
        "icon": "zap",
        "direction": "neutral",
        "description": (
            "Annual energy requirement assessed by the state (MU = million units "
            "= GWh), fiscal year — the demand-side counterpart to Availability "
            "(T139). Requirement minus availability gives the 'energy not "
            "supplied' deficit; the percentage gap (deficit / requirement) is "
            "the headline 'power deficit %' India tracked closely through the "
            "2000s. The all-India deficit shrank from ~10% in FY05 to under "
            "0.5% from FY18 onwards as supply caught up; persistent state-level "
            "deficits today flag distribution / scheduling issues, not "
            "generation shortage."
        ),
        "notes": (
            "Source: RBI Handbook of Statistics on Indian States 2024-25 edition, "
            "Table 141. Originating data: Central Electricity Authority. "
            "Direction is neutral because higher requirement reflects "
            "industrial activity (good) and / or inefficient consumption "
            "(not good) — interpret in conjunction with per-capita GSDP and "
            "T-and-D loss series."
        ),
    },
]


# T142 — Peak Demand vs Peak Met. Two artifacts; deficit derivable.
T142_SPECS = [
    {
        "out_path": "energy/state_peak_demand_mw.json",
        "id": "energy/state_peak_demand_mw",
        "title": "State-wise peak power demand (MW)",
        "xlsx": STATES_CACHE / "T142_PeakDemandVsPeakMet.xlsx",
        "table_label": "Table 142: State-wise Actual Power Supply Position — Peak Demand",
        "snapshot_url": "https://rbidocs.rbi.org.in/rdocs/Publications/DOCs/142T_111220252ABDAB650F53423F990E3EA91A0B5B5A.XLSX",
        "sub_col_offset": 0,
        "value_kind": "raw",
        "unit": "MW",
        "icon": "zap",
        "direction": "neutral",
        "description": (
            "Highest single-instant electricity demand observed in the state during "
            "the fiscal year (MW). 'Peak' is the system-wide simultaneous demand "
            "as recorded by the State Load Despatch Centre — typically a hot "
            "summer afternoon (north / west India) or a winter evening (Punjab, "
            "Delhi). Read alongside Peak Met (`state_peak_met_mw`) — the gap is "
            "the unmet peak demand, which is more operationally critical than "
            "the energy-deficit % because shortages here force load-shedding."
        ),
        "notes": (
            "Source: RBI Handbook of Statistics on Indian States 2024-25 edition, "
            "Table 142. Originating data: Central Electricity Authority. "
            "Note RBI relabelled 'Surplus / Deficit' to 'Demand Not Met' from "
            "FY 2019-20 onwards; the underlying definition is unchanged. "
            "Direction is neutral — higher peak demand reflects economic "
            "activity (good) and / or supply-side inefficiency at off-peak (not "
            "good); compare with `state_power_requirement_mu` for the energy-"
            "volume view."
        ),
    },
    {
        "out_path": "energy/state_peak_met_mw.json",
        "id": "energy/state_peak_met_mw",
        "title": "State-wise peak power supplied (MW)",
        "xlsx": STATES_CACHE / "T142_PeakDemandVsPeakMet.xlsx",
        "table_label": "Table 142: State-wise Actual Power Supply Position — Peak Met",
        "snapshot_url": "https://rbidocs.rbi.org.in/rdocs/Publications/DOCs/142T_111220252ABDAB650F53423F990E3EA91A0B5B5A.XLSX",
        "sub_col_offset": 1,
        "value_kind": "raw",
        "unit": "MW",
        "icon": "zap",
        "direction": "higher_is_better",
        "description": (
            "Maximum instantaneous power actually supplied in the state during "
            "the fiscal year (MW). The pair (peak_demand, peak_met) tells the "
            "load-shedding story: peak_met < peak_demand in any year means the "
            "grid had to drop load to keep frequency stable. India's all-India "
            "peak deficit fell from ~12% in FY05 to under 1% from FY18 onwards, "
            "but state-level shortfalls persist — Bihar, UP, Punjab, J&K, and "
            "Andhra Pradesh routinely under-met their own peak in the FY13-FY25 "
            "window covered here."
        ),
        "notes": (
            "Source: RBI Handbook of Statistics on Indian States 2024-25 edition, "
            "Table 142. Originating data: Central Electricity Authority. "
            "Coverage starts FY 2013-14 in this RBI edition (12 fiscal years). "
            "Pair with `state_peak_demand_mw` to compute deficit MW and "
            "deficit % at render time."
        ),
    },
]


# T143 — State-wise Grid Interactive Renewable Power capacity (calendar year, end-March).
T143_SPECS = [
    {
        "out_path": "energy/state_renewable_grid_capacity_mw.json",
        "id": "energy/state_renewable_grid_capacity_mw",
        "title": "State-wise installed grid-interactive renewable capacity (MW)",
        "xlsx": STATES_CACHE / "T143_GridInteractiveRenewableCapacity.xlsx",
        "table_label": "Table 143: State-wise Total Installed Capacity of Grid Interactive Renewable Power",
        "snapshot_url": "https://rbidocs.rbi.org.in/rdocs/Publications/DOCs/143T_11122025415C21E8727B4EA0B009173F7DC19E84.XLSX",
        "value_kind": "raw",
        "unit": "MW",
        "icon": "sun",
        "direction": "higher_is_better",
        "description": (
            "Cumulative grid-connected renewable-power generation capacity "
            "installed in the state, MW, as at end-March of the calendar year. "
            "Includes wind + solar + small hydro + biomass + waste-to-energy "
            "combined (RBI's table does not publish a per-source split). The "
            "closest proxy in the RBI Handbook for a state-level fuel-mix view "
            "of capacity. Coverage 2007–2024 = 18 years; demonstrates the scale "
            "of India's RE expansion (national total 10,256 MW in 2007 → "
            "143,645 MW in 2024, a 14× increase). Rajasthan (517→26,693 MW), "
            "Gujarat (644→25,472), Tamil Nadu (3,802→19,983), Karnataka, and "
            "Maharashtra dominate; Bihar, Odisha, north-eastern states remain "
            "in single-digit GW territory."
        ),
        "notes": (
            "Source: RBI Handbook of Statistics on Indian States 2024-25 edition, "
            "Table 143. Originating data per the workbook footer: 'Energy "
            "Statistics, Ministry of Statistics and Programme Implementation, "
            "Government of India.' Time grain is calendar-year (as-at-end-March "
            "snapshot) — labelled with the year of the March 31 reading. "
            "Telangana has data from 2015 (state created June 2014); Ladakh "
            "from 2023 (UT created October 2019). 'Total' / 'Others' rows in "
            "the source workbook are skipped at parse time."
        ),
    },
]


def _build_t142_extra_metadata(spec: dict) -> dict:
    return {}


def _build_artifact(spec: dict, rows: list[dict]) -> dict:
    times = sorted({r["time"] for r in rows})
    entities = sorted({r["entity_id"] for r in rows})
    time_grain = spec.get("time_grain", "fiscal_year")
    return {
        "$schema": "https://yen-gov.github.io/schemas/indicator.schema.json",
        "$schema_version": "1.3",
        "sources": [
            {
                "url": spec["snapshot_url"],
                "fetched_at": FETCHED_AT,
                "name": f"RBI Handbook 2024-25 — {spec['table_label']}",
                "authority": SOURCE_RBI,
            },
            {
                "url": HBS_IS_LANDING,
                "fetched_at": FETCHED_AT,
                "name": "RBI Handbook of Statistics on Indian States — landing page",
                "authority": "Reserve Bank of India",
            },
        ],
        "license": LICENSE_RBI,
        "coverage": {
            "spatial": f"India (states + UTs); {len(entities)} entities",
            "temporal": f"{times[0]}..{times[-1]}",
            "admin_level": "state",
        },
        "indicator": {
            "id": spec["id"],
            "title": spec["title"],
            "description": spec["description"],
            "entity_kind": "state",
            "time_grain": time_grain,
            "value_kind": spec["value_kind"],
            "direction": spec.get("direction", "neutral"),
            "scale_hint": "linear",
            "unit": spec["unit"],
            "icon": spec.get("icon", "zap"),
            "attribution_geography": "where_administered",
            "comparability": "comparable_with_normalisation",
            "implementing_authority": "centre",
            "methodology_vintage": spec.get("methodology_vintage", (
                "RBI Handbook of Statistics on Indian States 2024-25 edition; "
                "originating data Central Electricity Authority, Ministry of Power."
            )),
            "notes": spec["notes"],
        },
        "rows": rows,
    }


def _write(spec: dict, art: dict) -> None:
    write_artifact(OUT / spec["out_path"], art)


def main() -> None:
    print("=== HBS-IS Power section (T138-T141) — FY × state ===")
    for spec in SPECS:
        rows = parse_state_year_table(spec["xlsx"], header_row=3)
        if not rows:
            print(f"  WARN no rows for {spec['id']}")
            continue
        art = _build_artifact(spec, rows)
        _write(spec, art)

    print("\n=== HBS-IS Power section (T142) — Peak demand vs peak met ===")
    for spec in T142_SPECS:
        rows = parse_t142_peak(spec["xlsx"], spec["sub_col_offset"])
        if not rows:
            print(f"  WARN no rows for {spec['id']}")
            continue
        art = _build_artifact(spec, rows)
        _write(spec, art)

    print("\n=== HBS-IS Power section (T143) — Renewable grid capacity (CY) ===")
    for spec in T143_SPECS:
        rows = parse_state_year_table(spec["xlsx"], header_row=4, calendar=True)
        if not rows:
            print(f"  WARN no rows for {spec['id']}")
            continue
        spec["time_grain"] = "year"
        spec["methodology_vintage"] = (
            "RBI Handbook of Statistics on Indian States 2024-25 edition; "
            "originating data Energy Statistics, MoSPI, Government of India. "
            "End-March cumulative installed capacity snapshots."
        )
        art = _build_artifact(spec, rows)
        _write(spec, art)


if __name__ == "__main__":
    main()
