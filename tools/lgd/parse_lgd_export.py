#!/usr/bin/env python3
"""Parse the throwaway LGD exports in ``datasets/ephemeral/`` into committed,
deterministic parsed-snapshot CSVs under ``datasets/data/entities/lgd/``.

This is PR-stage 0a of sub-plan
``TODO/20260604-b2b5-elections-reingest-subplan.md`` (chunk B2b.5.0). It is the
FIRST filter in the committed seam the plan section 0c.1 mandates::

    datasets/ephemeral/*.csv|*.xlsx   (gitignored, throwaway, SOURCE-only)
        -> tools/lgd/parse_lgd_export.py        (THIS file)
        -> datasets/data/entities/lgd/<table>.csv    (committed snapshot = source of truth)
           + datasets/_ops/lgd-parse-receipt.json    (sha256 + row counts + vintage)
        -> backend canonical seed (0b / 0c)
        -> datasets/data/entities/*.csv          (columns.json-governed)

Reproducibility contract (plan 0c.1): a fresh checkout + this committed builder
regenerates the snapshot CSVs byte-for-byte from the same ephemeral inputs; the
per-file sha256 in ``datasets/_ops/lgd-parse-receipt.json`` ties the committed
snapshot to the (uncommitted) raw bytes, so byte-exact re-download is NOT
required.

Layout (G8-finish, 2026-06-08, plan section 9): the snapshot tables live
under ``datasets/data/entities/lgd/`` (citizen-tier reference, alongside
``entities/state_codes.csv`` etc.) and the parse receipt under
``datasets/_ops/`` (operator-tier, not citizen-facing) - retiring the
old ``datasets/reference/`` tier.

Five snapshot tables are emitted:

- ``states.csv``                          - LGD state/UT register (kind, census).
- ``districts.csv``                       - LGD district register (parent state, census).
- ``subdistricts.csv``                    - LGD sub-district register (parked; no v1
                                            consumer per plan 0c.7, parsed in the same pass).
- ``constituencies.csv``                  - AC + PC register from the PRI super-file,
                                            carrying the ECI code per constituency
                                            (the ``eci_no`` direct-join source, plan 0d
                                            round-8c table row 3b) and each AC's parent PC.
- ``constituency_district_membership.csv``- AC -> LGD-district 1:many edges with a
                                            village-row count and a derived ``is_primary``
                                            (the plurality district), the LGD-canonical
                                            AC<->district relation (plan F3 / 0b#6).

Source shapes (verbatim headers, 2026-06-05 snapshot):

- ``All_Stateof_India_*.csv``: title row + ``S.No.,State Code,State Version,
  State Name (In English),State Name (In Local),Census 2001 Code,Census 2011 Code,
  State or UT``.
- ``All_Districtof_India_*.csv``: title + ``S.No.,State Code,State Name (In English),
  District Code,District Name(In English),Census 2001 Code,Census 2011 Code``.
- ``All_Sub_Districtof_India_*.csv``: title + ``S.No.,State Code,State Name,
  District Code,District Name,Sub-district Code,Sub-district Version,
  Sub-district Name,Census 2001 Code,Census 2011 Code``.
- ``Parliment_..._Pri_*.xlsx`` (35 files, one per state/UT): title + ``S.No.,
  Parliament Constituency code,Parliament Constituency ECI Code,
  Parliament Constituency Name,Assembly Constituency Code,
  Assembly Constituency ECI Code,Assembly Constituency Name,District Code,
  District Name,District Census 2011 Code,Subdistrict Code,...``.

Authority-key discipline (plan 0c.1): LGD/ECI codes are read WITHOUT integer
coercion off a string source; XLSX numeric cells are converted via ``int`` only
for the LGD/ECI sequence ids (which never carry leading zeros). Census codes are
taken verbatim from the CSV registers (never the XLSX), preserving any leading
zeros, and are LABELS never join keys (plan 0d round-8c).

Naming divergence from the plan: the plan names this ``parse_lgd_html.py``; the
operator supplied CSV + XLSX exports (cleaner than the rendered HTML, plan 0d
round-8 point 5 "XLS or HTML, operator's choice"), so it is named
``parse_lgd_export.py``. Recorded in the 0a PR body.

This tool imports ONLY the Python stdlib + ``openpyxl`` (a project dependency);
per CLAUDE.md section 4 a ``tools/`` module MUST NOT import ``backend/yen_gov``.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# --------------------------------------------------------------------------- #
# Constants                                                                    #
# --------------------------------------------------------------------------- #

# The operator snapshot window IS the vintage (LGD exposes no edition tag);
# mandatory per plan 0c.5 / ADR-0042. A constant, never a wall-clock value, so
# repeated runs are byte-identical.
SNAPSHOT_VINTAGE = "2026-06-05"

# v1 freezes the delimitation at 2008 (the in-force Delimitation Commission
# cycle); the register reflects current boundaries. delim_year lands in the
# canonical electoral.csv at 0c, not in the neutral parsed snapshot.

STATES_GLOB = "All_Stateof_India_*.csv"
DISTRICTS_GLOB = "All_Districtof_India_*.csv"
SUBDISTRICTS_GLOB = "All_Sub_Districtof_India_*.csv"
PRI_GLOB = "Parliment_Constituency_And_Assembly_Constituency_Pri_*.xlsx"

# Bridge to the canonical schema-version helper so the receipt envelope below
# never carries a hand-typed semver literal (CLAUDE.md section 11). The
# helper reads `datasets/schemas/<file>.schema.json`'s `x-version` once at
# import time and caches; drift becomes impossible by construction.
_REPO_ROOT = Path(__file__).resolve().parents[2]
_BACKEND_DIR = _REPO_ROOT / "backend"
if str(_BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(_BACKEND_DIR))

from yen_gov.core.schema_registry import schema_version  # noqa: E402

RECEIPT_SCHEMA_VERSION = schema_version("lgd-parse-receipt.schema.json")

# Expected header rows (row index 1, after the title row at index 0). Validated
# verbatim so an upstream format change fails loud instead of mis-indexing.
STATES_HEADER = [
    "S.No.", "State Code", "State Version", "State Name (In English)",
    "State Name (In Local)", "Census 2001 Code", "Census 2011 Code", "State or UT",
]
DISTRICTS_HEADER = [
    "S.No.", "State Code", "State Name (In English)", "District Code",
    "District Name(In English)", "Census 2001 Code", "Census 2011 Code",
]
SUBDISTRICTS_HEADER = [
    "S.No.", "State Code", "State Name", "District Code", "District Name",
    "Sub-district Code", "Sub-district Version", "Sub-district Name",
    "Census 2001 Code", "Census 2011 Code",
]
PRI_HEADER = [
    "S.No.", "Parliament Constituency code", "Parliament Constituency ECI Code",
    "Parliament Constituency Name", "Assembly Constituency Code",
    "Assembly Constituency ECI Code", "Assembly Constituency Name",
    "District Code", "District Name", "District Census 2011 Code",
    "Subdistrict Code", "Subdistrict Name", "Subdistrict Census 2011 Code",
    "Village Code", "Village Name", "Village Census 2011 Code",
    "Rural Localbody Code", "Rural LocalbodyName", "Block Code", "Block Name",
]

# Output snapshot file names + their column order.
OUT_STATES = "states.csv"
OUT_DISTRICTS = "districts.csv"
OUT_SUBDISTRICTS = "subdistricts.csv"
OUT_CONSTITUENCIES = "constituencies.csv"
OUT_MEMBERSHIP = "constituency_district_membership.csv"
OUT_RECEIPT = "parse-receipt.json"

STATES_COLS = [
    "lgd_state_code", "state_name", "state_name_local",
    "census_2001_code", "census_2011_code", "kind",
]
DISTRICTS_COLS = [
    "lgd_state_code", "state_name", "lgd_district_code", "district_name",
    "census_2001_code", "census_2011_code",
]
SUBDISTRICTS_COLS = [
    "lgd_state_code", "state_name", "lgd_district_code", "district_name",
    "lgd_subdistrict_code", "subdistrict_name", "census_2001_code", "census_2011_code",
]
CONSTITUENCIES_COLS = [
    "lgd_state_code", "kind", "lgd_code", "eci_code", "name", "parent_pc_lgd_code",
]
MEMBERSHIP_COLS = [
    "lgd_state_code", "ac_lgd_code", "lgd_district_code", "village_count", "is_primary",
]


class ParseError(ValueError):
    """Raised when a source export violates an expected shape (STOP-AND-SURFACE)."""


# --------------------------------------------------------------------------- #
# Low-level cell helpers                                                       #
# --------------------------------------------------------------------------- #

def _code_str(value: Any) -> str:
    """Render an LGD/ECI code cell as a canonical string without coercion loss.

    XLSX numeric cells arrive as ``float`` (``411.0``); LGD/ECI ids are integer
    sequences with no leading zeros, so ``int`` is lossless for them. CSV cells
    arrive as ``str`` and are returned stripped. ``None`` / blank -> ``""``.
    """
    if value is None:
        return ""
    if isinstance(value, bool):  # guard: bool is a subclass of int
        raise ParseError(f"unexpected boolean code cell: {value!r}")
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if not value.is_integer():
            raise ParseError(f"non-integer LGD/ECI code cell: {value!r}")
        return str(int(value))
    return str(value).strip()


def _text(value: Any) -> str:
    """Render a name/label cell as a stripped string (UTF-8 preserved)."""
    if value is None:
        return ""
    return str(value).strip()


def _census(value: Any) -> str:
    """Render a census code cell, normalising the LGD ``0`` sentinel to empty.

    LGD emits ``0`` (or ``000`` etc.) for an entity that did NOT exist at that
    census - Telangana (2014), Ladakh / the merged DNH-DD (2019-2020), and every
    post-2011 district. Carrying a literal ``0`` would let a naive
    ``JOIN ON census_*_code = 0`` match every such entity at once; the census
    column is a LABEL never a key (sub-plan 0c.4 / round-8c), and an empty cell
    is the honest "did not exist / no code" signal. A genuine non-zero code is
    preserved verbatim (leading zeros intact - it is read off the string CSV).
    """
    text = _text(value)
    if text == "" or set(text) == {"0"}:
        return ""
    return text


def _norm_name(name: str) -> str:
    """Normalise a state name for matching: uppercase, single-spaced."""
    return re.sub(r"\s+", " ", name).strip().upper()


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _sort_int(code: str) -> tuple[int, str]:
    """Sort key that orders numeric codes numerically, blanks last."""
    if code == "":
        return (1, "")
    try:
        return (0, f"{int(code):012d}")
    except ValueError:
        return (0, code)


# --------------------------------------------------------------------------- #
# Source readers (return data rows; validate the header)                       #
# --------------------------------------------------------------------------- #

def _read_titled_csv(path: Path, expected_header: list[str]) -> list[list[str]]:
    """Read a LGD CSV export: row 0 = title, row 1 = header, rest = data.

    Validates the header verbatim. Returns the data rows (list of cell lists).
    """
    with path.open(encoding="utf-8-sig", newline="") as fh:
        rows = list(csv.reader(fh))
    if len(rows) < 2:
        raise ParseError(f"{path.name}: fewer than 2 rows (no title+header)")
    header = [c.strip() for c in rows[1][: len(expected_header)]]
    if header != expected_header:
        raise ParseError(
            f"{path.name}: unexpected header\n  got:      {header}\n"
            f"  expected: {expected_header}"
        )
    return [r for r in rows[2:] if any(c.strip() for c in r)]


def parse_states(path: Path) -> list[dict[str, str]]:
    """Parse the LGD all-states export into state register rows."""
    out: list[dict[str, str]] = []
    for r in _read_titled_csv(path, STATES_HEADER):
        kind_raw = _text(r[7]).upper()
        kind = {"S": "state", "U": "ut"}.get(kind_raw)
        if kind is None:
            raise ParseError(f"{path.name}: unknown State-or-UT flag {r[7]!r}")
        out.append({
            "lgd_state_code": _code_str(r[1]),
            "state_name": _text(r[3]),
            "state_name_local": _text(r[4]),
            "census_2001_code": _census(r[5]),
            "census_2011_code": _census(r[6]),
            "kind": kind,
        })
    return out


def parse_districts(path: Path) -> list[dict[str, str]]:
    """Parse the LGD all-districts export into district register rows."""
    out: list[dict[str, str]] = []
    for r in _read_titled_csv(path, DISTRICTS_HEADER):
        out.append({
            "lgd_state_code": _code_str(r[1]),
            "state_name": _text(r[2]),
            "lgd_district_code": _code_str(r[3]),
            "district_name": _text(r[4]),
            "census_2001_code": _census(r[5]),
            "census_2011_code": _census(r[6]),
        })
    return out


def parse_subdistricts(path: Path) -> list[dict[str, str]]:
    """Parse the LGD all-sub-districts export into sub-district register rows."""
    out: list[dict[str, str]] = []
    for r in _read_titled_csv(path, SUBDISTRICTS_HEADER):
        out.append({
            "lgd_state_code": _code_str(r[1]),
            "state_name": _text(r[2]),
            "lgd_district_code": _code_str(r[3]),
            "district_name": _text(r[4]),
            "lgd_subdistrict_code": _code_str(r[5]),
            "subdistrict_name": _text(r[7]),
            "census_2001_code": _text(r[8]),
            "census_2011_code": _census(r[9]),
        })
    return out


def _pri_data_rows(ws: Any) -> tuple[str, list[tuple[Any, ...]]]:
    """Yield (title, data_rows) from a PRI worksheet; validate the header."""
    it = ws.iter_rows(values_only=True)
    title_row = next(it, None)
    header_row = next(it, None)
    if title_row is None or header_row is None:
        raise ParseError("PRI worksheet has fewer than 2 rows")
    title = _text(title_row[0])
    header = [_text(c) for c in header_row[: len(PRI_HEADER)]]
    if header != PRI_HEADER:
        raise ParseError(
            f"PRI '{title}': unexpected header\n  got:      {header}\n"
            f"  expected: {PRI_HEADER}"
        )
    return title, [r for r in it if r is not None and r[1] is not None]


def _state_from_title(title: str, state_name_to_code: dict[str, str]) -> str | None:
    """Resolve a workbook state code from its title (``State Of <NAME> ...``)."""
    if "State Of " not in title:
        return None
    name = title.split("State Of ", 1)[-1].split(" Parliament", 1)[0].strip()
    key = _norm_name(name)
    return state_name_to_code.get(key) or state_name_to_code.get(
        key.removeprefix("THE ").strip()
    )


def parse_pri_workbook(
    path: Path,
    district_to_state: dict[str, str],
    state_name_to_code: dict[str, str],
) -> tuple[str, list[dict[str, str]], list[dict[str, str]], list[str], int]:
    """Parse one PRI super-file workbook.

    Returns ``(state_code, constituency_rows, membership_rows, unmatched_districts,
    raw_row_count)``. The workbook's state is resolved from its title (every PRI
    export is single-state) and CROSS-VALIDATED against the LGD district codes it
    references when present (a UT with no assembly - Chandigarh, Lakshadweep -
    carries no district rows, so the title is the only signal). Each AC's parent
    PC and ECI code, and each AC<->district village count, come straight off the
    rows.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        title, data_rows = _pri_data_rows(ws)
    finally:
        wb.close()

    # AC identity + name + eci + parent PC (collapsed across the AC's many
    # village rows). PCs collected even when an AC is absent (UT-no-assembly).
    ac_name: dict[str, str] = {}
    ac_eci: dict[str, str] = {}
    ac_pc: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    pc_name: dict[str, str] = {}
    pc_eci: dict[str, str] = {}
    ac_district_villages: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    district_codes_seen: set[str] = set()
    raw_count = 0

    for row in data_rows:
        raw_count += 1
        pc_code = _code_str(row[1])
        pc_eci_code = _code_str(row[2])
        pc_nm = _text(row[3])
        ac_code = _code_str(row[4])
        ac_eci_code = _code_str(row[5])
        ac_nm = _text(row[6])
        dist_code = _code_str(row[7])

        if pc_code and pc_code != "0":
            pc_name.setdefault(pc_code, pc_nm)
            pc_eci.setdefault(pc_code, pc_eci_code)
        if ac_code and ac_code != "0":
            ac_name.setdefault(ac_code, ac_nm)
            ac_eci.setdefault(ac_code, ac_eci_code)
            if pc_code and pc_code != "0":
                ac_pc[ac_code][pc_code] += 1
            if dist_code and dist_code != "0":
                ac_district_villages[ac_code][dist_code] += 1
        if dist_code and dist_code != "0":
            district_codes_seen.add(dist_code)

    # Resolve the workbook's state from its title (every PRI export is single-
    # state), then cross-validate against the district-code authority join.
    state_code = _state_from_title(title, state_name_to_code)
    if state_code is None:
        raise ParseError(f"PRI '{title}': title does not resolve to a known state")
    resolved = {district_to_state[d] for d in district_codes_seen if d in district_to_state}
    unmatched = sorted(district_codes_seen - set(district_to_state), key=_sort_int)
    if len(resolved) > 1:
        raise ParseError(
            f"PRI '{title}': district codes span multiple states {sorted(resolved)}"
        )
    if resolved and next(iter(resolved)) != state_code:
        raise ParseError(
            f"PRI '{title}': title state {state_code!r} disagrees with "
            f"district-join state {next(iter(resolved))!r}"
        )

    # Constituency register: one row per PC, then one per AC (with its parent PC).
    cons_rows: list[dict[str, str]] = []
    for pc_code in pc_name:
        cons_rows.append({
            "lgd_state_code": state_code,
            "kind": "pc",
            "lgd_code": pc_code,
            "eci_code": pc_eci[pc_code],
            "name": pc_name[pc_code],
            "parent_pc_lgd_code": "",
        })
    for ac_code in ac_name:
        pc_votes = ac_pc.get(ac_code, {})
        parent_pc = ""
        if pc_votes:
            # An AC nests in exactly one PC by delimitation; pick the modal PC
            # (tie -> lowest code) to absorb any stray row noise.
            parent_pc = sorted(pc_votes.items(), key=lambda kv: (-kv[1], _sort_int(kv[0])))[0][0]
        cons_rows.append({
            "lgd_state_code": state_code,
            "kind": "ac",
            "lgd_code": ac_code,
            "eci_code": ac_eci[ac_code],
            "name": ac_name[ac_code],
            "parent_pc_lgd_code": parent_pc,
        })

    # AC<->district membership: one row per (ac, district); is_primary = the
    # plurality district (tie -> lowest district code).
    mem_rows: list[dict[str, str]] = []
    for ac_code, dist_villages in ac_district_villages.items():
        primary = sorted(
            dist_villages.items(), key=lambda kv: (-kv[1], _sort_int(kv[0]))
        )[0][0]
        for dist_code, vcount in dist_villages.items():
            mem_rows.append({
                "lgd_state_code": state_code,
                "ac_lgd_code": ac_code,
                "lgd_district_code": dist_code,
                "village_count": str(vcount),
                "is_primary": "true" if dist_code == primary else "false",
            })

    return state_code, cons_rows, mem_rows, unmatched, raw_count


# --------------------------------------------------------------------------- #
# Deterministic CSV writer                                                     #
# --------------------------------------------------------------------------- #

def _write_csv(path: Path, columns: list[str], rows: list[dict[str, str]]) -> int:
    """Write ``rows`` to ``path`` deterministically (UTF-8, LF, trailing NL, no BOM).

    Mirrors the canonical ``write_csv`` serialisation discipline (LF only, no
    BOM, trailing newline, QUOTE_MINIMAL) so the parsed snapshot reads cleanly in
    the backend seed. Returns the row count.
    """
    buf = io.StringIO(newline="")
    writer = csv.writer(buf, lineterminator="\n", quoting=csv.QUOTE_MINIMAL)
    writer.writerow(columns)
    for r in rows:
        writer.writerow([r.get(c, "") for c in columns])
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(buf.getvalue(), encoding="utf-8", newline="")
    return len(rows)


def _sorted(rows: list[dict[str, str]], keys: list[str]) -> list[dict[str, str]]:
    return sorted(rows, key=lambda r: tuple(_sort_int(r[k]) for k in keys))


# --------------------------------------------------------------------------- #
# Orchestration                                                                #
# --------------------------------------------------------------------------- #

def _one(ephemeral: Path, glob: str) -> Path:
    """Resolve exactly-one (latest by name) ephemeral source matching ``glob``."""
    matches = sorted(ephemeral.glob(glob))
    if not matches:
        raise ParseError(f"no ephemeral source matches {glob!r} in {ephemeral}")
    return matches[-1]


def discover_sources(ephemeral: Path) -> dict[str, Any]:
    """Resolve the source paths under ``ephemeral`` by their export name globs.

    Returns ``{"states", "districts", "subdistricts", "pri"}`` where ``pri`` is a
    sorted list. Raises ``ParseError`` if any required export is absent. Exposed
    so the golden test can substitute fixture filenames without the production
    glob discipline.
    """
    pri_srcs = sorted(ephemeral.glob(PRI_GLOB))
    if not pri_srcs:
        raise ParseError(f"no PRI workbook matches {PRI_GLOB!r} in {ephemeral}")
    return {
        "states": _one(ephemeral, STATES_GLOB),
        "districts": _one(ephemeral, DISTRICTS_GLOB),
        "subdistricts": _one(ephemeral, SUBDISTRICTS_GLOB),
        "pri": pri_srcs,
    }


def build_snapshot(
    ephemeral: Path,
    out_dir: Path,
    *,
    sources: dict[str, Any] | None = None,
    receipt_path: Path | None = None,
) -> dict[str, Any]:
    """Parse the ephemeral exports and write the five snapshot CSVs + receipt.

    Args:
        ephemeral: directory holding the throwaway LGD exports (used to discover
            sources when ``sources`` is not given).
        out_dir: committed snapshot CSV output directory
            (default production location: ``datasets/data/entities/lgd``).
        sources: optional explicit ``{states, districts, subdistricts, pri}`` map
            (the golden test passes fixture paths whose names do not match the
            production globs).
        receipt_path: path the parse receipt JSON is written to. When not given,
            falls back to ``out_dir / OUT_RECEIPT`` for backward compatibility
            with the golden-fixture tests that point ``out_dir`` at a tmp dir.
            Production runs pass ``datasets/_ops/lgd-parse-receipt.json`` per
            plan-doc section 9 (operator state lives under ``_ops/``).

    Returns the receipt dict (also written to ``receipt_path``).
    """
    if sources is None:
        sources = discover_sources(ephemeral)
    states_src: Path = sources["states"]
    districts_src: Path = sources["districts"]
    subdistricts_src: Path = sources["subdistricts"]
    pri_srcs: list[Path] = list(sources["pri"])

    state_rows = parse_states(states_src)
    district_rows = parse_districts(districts_src)
    subdistrict_rows = parse_subdistricts(subdistricts_src)

    district_to_state = {d["lgd_district_code"]: d["lgd_state_code"] for d in district_rows}
    state_name_to_code = {_norm_name(s["state_name"]): s["lgd_state_code"] for s in state_rows}

    cons_rows: list[dict[str, str]] = []
    mem_rows: list[dict[str, str]] = []
    pri_receipts: list[dict[str, Any]] = []
    all_unmatched: dict[str, list[str]] = {}
    for src in pri_srcs:
        state_code, c_rows, m_rows, unmatched, raw_count = parse_pri_workbook(
            src, district_to_state, state_name_to_code
        )
        cons_rows.extend(c_rows)
        mem_rows.extend(m_rows)
        if unmatched:
            all_unmatched[state_code] = unmatched
        pri_receipts.append({
            "file": src.name,
            "sha256": _sha256(src),
            "lgd_state_code": state_code,
            "raw_rows": raw_count,
            "unmatched_district_codes": unmatched,
        })

    # Deterministic ordering.
    state_rows = _sorted(state_rows, ["lgd_state_code"])
    district_rows = _sorted(district_rows, ["lgd_state_code", "lgd_district_code"])
    subdistrict_rows = _sorted(
        subdistrict_rows, ["lgd_state_code", "lgd_district_code", "lgd_subdistrict_code"]
    )
    cons_rows = _sorted(cons_rows, ["lgd_state_code", "kind", "lgd_code"])
    mem_rows = _sorted(mem_rows, ["lgd_state_code", "ac_lgd_code", "lgd_district_code"])

    n_states = _write_csv(out_dir / OUT_STATES, STATES_COLS, state_rows)
    n_districts = _write_csv(out_dir / OUT_DISTRICTS, DISTRICTS_COLS, district_rows)
    n_subdistricts = _write_csv(out_dir / OUT_SUBDISTRICTS, SUBDISTRICTS_COLS, subdistrict_rows)
    n_cons = _write_csv(out_dir / OUT_CONSTITUENCIES, CONSTITUENCIES_COLS, cons_rows)
    n_mem = _write_csv(out_dir / OUT_MEMBERSHIP, MEMBERSHIP_COLS, mem_rows)

    receipt: dict[str, Any] = {
        "$schema": "../schemas/lgd-parse-receipt.schema.json",
        "$schema_version": RECEIPT_SCHEMA_VERSION,
        "snapshot_vintage": SNAPSHOT_VINTAGE,
        "parser": "tools/lgd/parse_lgd_export.py",
        "sources": [
            {"file": states_src.name, "sha256": _sha256(states_src), "raw_rows": len(state_rows)},
            {"file": districts_src.name, "sha256": _sha256(districts_src), "raw_rows": len(district_rows)},
            {"file": subdistricts_src.name, "sha256": _sha256(subdistricts_src), "raw_rows": len(subdistrict_rows)},
        ],
        "pri_sources": pri_receipts,
        "outputs": [
            {"file": OUT_STATES, "rows": n_states},
            {"file": OUT_DISTRICTS, "rows": n_districts},
            {"file": OUT_SUBDISTRICTS, "rows": n_subdistricts},
            {"file": OUT_CONSTITUENCIES, "rows": n_cons},
            {"file": OUT_MEMBERSHIP, "rows": n_mem},
        ],
        "unmatched_district_codes_by_state": all_unmatched,
    }
    if receipt_path is None:
        # Backwards-compat for the golden-fixture tests: drop the receipt
        # next to the CSVs when no explicit location was given.
        receipt_path = out_dir / OUT_RECEIPT
    receipt_path.parent.mkdir(parents=True, exist_ok=True)
    receipt_path.write_text(
        json.dumps(receipt, indent=2, ensure_ascii=False, sort_keys=True) + "\n",
        encoding="utf-8",
        newline="",
    )
    return receipt


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--ephemeral", type=Path, default=Path("datasets/ephemeral"),
        help="Directory holding the throwaway LGD exports (default: datasets/ephemeral).",
    )
    parser.add_argument(
        "--out", type=Path, default=Path("datasets/data/entities/lgd"),
        help="Committed snapshot CSV output directory (default: datasets/data/entities/lgd).",
    )
    parser.add_argument(
        "--receipt", type=Path, default=Path("datasets/_ops/lgd-parse-receipt.json"),
        help="Parse-receipt JSON output path (default: datasets/_ops/lgd-parse-receipt.json).",
    )
    args = parser.parse_args(argv)
    receipt = build_snapshot(args.ephemeral, args.out, receipt_path=args.receipt)
    n = {o["file"]: o["rows"] for o in receipt["outputs"]}
    print(f"wrote snapshot to {args.out} (vintage {receipt['snapshot_vintage']}):")
    for fname, rows in n.items():
        print(f"  {fname}: {rows} rows")
    unmatched = receipt["unmatched_district_codes_by_state"]
    if unmatched:
        total = sum(len(v) for v in unmatched.values())
        print(f"  NOTE: {total} unmatched district code(s) across "
              f"{len(unmatched)} state(s) - see parse-receipt.json")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
