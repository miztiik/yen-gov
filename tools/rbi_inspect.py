"""Inspect an RBI XLSX: list sheets, dump headers, show a few sample rows.

Usage:
    python tools/rbi_inspect.py path/to/file.xlsx [--sheet SHEET]
"""
from __future__ import annotations

import argparse
import io
import sys
from pathlib import Path

# Force UTF-8 stdout (Windows defaults to cp1252 which chokes on ₹/em-dash).
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("xlsx", type=Path)
    p.add_argument("--sheet", default=None)
    p.add_argument("--rows", type=int, default=20)
    args = p.parse_args()

    wb = load_workbook(args.xlsx, data_only=True, read_only=True)
    print(f"== file: {args.xlsx}")
    print(f"== sheets ({len(wb.sheetnames)}):")
    for s in wb.sheetnames:
        ws = wb[s]
        print(f"   - {s!r}  ({ws.max_row} rows x {ws.max_column} cols)")

    target = args.sheet or wb.sheetnames[0]
    if target not in wb.sheetnames:
        print(f"!! sheet {target!r} not found", flush=True)
        return 2

    print(f"\n== first {args.rows} rows of {target!r}:")
    ws = wb[target]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= args.rows:
            break
        cells = [
            ("" if v is None else str(v)).strip()[:40]
            for v in row
        ]
        # Trim trailing blanks
        while cells and not cells[-1]:
            cells.pop()
        print(f"  r{i:02d}: {cells}")

    wb.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
