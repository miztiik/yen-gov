"""Correlate state-AE UNK candidacies against ECI Statement-X Detailed Results xlsx.

The ECI Statistical Report Statement-X (10 - Detailed Results) is the
publisher's authoritative per-(AC, candidate) listing for an assembly
election. It carries the publisher's ``PARTY`` (abbreviation) for every
candidate. Our state assembly candidacies.csv has ~660 rows where
``party_id == parties.IN.UNK`` because the TCPD publisher label on
``party_short_raw`` does not match any row in
``datasets/data/entities/parties.csv``.

This tool joins each UNK candidacy row against the ECI Detailed Results
xlsx by ``(constituency_name, candidate_name)`` (normalised + numeric
prefix stripped), then resolves the joined ECI abbr against parties.csv:

1. ``NOTA`` / ``IND``: alias-add to the corresponding sentinel pid.
2. Existing ``short`` / ``aliases`` match (case-insensitive): alias-add.
3. No match: ``mint-new`` with ``parties.IN.<abbr>``. If that pid already
   exists with a different full, disambiguate to
   ``<abbr>_AE<rep_year>_<rep_state_iso>``.

Per-label aggregation: rows are grouped by ``party_short_raw`` across all
target state-years; if a label's joined ECI rows disagree (e.g. publisher
emits same label for two genuinely-different parties), pick the dominant
ECI label when it covers >= 80% of joined rows; otherwise skip the label
entirely with reason ``eci-internal-collision``.

xlsx shape (discovered via Phase 1 recon, 2026-06-12):
  - All 11 files have a single sheet (``Worksheet``); header row at index
    3; data starts at row 4.
  - Priority 6 files (2022 + 2023 cycles, 14 cols): STATE/UT NAME, AC NO.,
    AC NAME, CANDIDATE NAME, SEX, AGE, CATEGORY, PARTY, SYMBOL, GENERAL,
    POSTAL, TOTAL, % VOTES POLLED, TOTAL ELECTORS
  - Opportunistic 5 files (May 2021 cycle, 15 cols): same as above with
    GENDER instead of SEX and an extra "OVER VALID VOTES + NOTA" column.
  - Candidate names carry a leading position prefix like ``"1 Umar Ali
    Khan"`` (the numeric is the within-constituency position) - stripped
    in ``normalise()``.
  - Constituency names sometimes carry ``"(SC)"`` / ``"(ST)"`` suffixes -
    stripped in ``normalise()``.
  - Single-state files: each xlsx is one state's full AE returns.

Year derivation: 3-tier fallback (filename year -> xlsx-title year -> per-
file hardcoded default in ``STATE_XLSX_MAP``). For real files, 6 of 11
hit tier 1 (filename), the rest fall to tier 3 (hardcoded). Tier 2 is
exercised by fixture tests only.

Output: emits ``verdict.csv`` + ``skipped.csv`` under
``datasets/ephemeral/party-parity/eci-state-xlsx/<run-id>/``. The
verdict.csv uses the PR #952 schema consumed by
``tools.correlate_unk_apply``; the ``tcpd_*`` column names are historical
(the apply tool was authored for PR-Q2 TCPD work) and carry ECI
``PARTY`` here.

Run from the repo root:

    python -m tools.correlate_unk_via_eci_state_xlsx
    python -m tools.correlate_unk_via_eci_state_xlsx --files 2022_uttar_pradesh_10-Detailed\\ Results.xlsx
    python -m tools.correlate_unk_via_eci_state_xlsx --limit 100

Apply the verdict via the existing seam:

    python -m tools.correlate_unk_apply --verdict-csv <path-to-verdict>
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[2]
PARTIES_CSV = REPO_ROOT / "datasets" / "data" / "entities" / "parties.csv"
EPHEMERAL_DIR = REPO_ROOT / "datasets" / "ephemeral"
VERDICT_ROOT = (
    REPO_ROOT / "datasets" / "ephemeral" / "party-parity" / "eci-state-xlsx"
)

# filename -> (state_slug, hardcoded_year_fallback)
# Priority 6: filename carries year (2022/2023); hardcoded redundant but
# kept for the tier-3 fallback path.
# Opportunistic 5: filename has NO year; hardcoded 2021 is the source of
# truth (these are all from the May 2021 5-state cycle).
STATE_XLSX_MAP: dict[str, tuple[str, int]] = {
    "2022_uttar_pradesh_10-Detailed Results.xlsx": ("uttar-pradesh", 2022),
    "2022_gujarat_10-Detailed Results.xlsx": ("gujarat", 2022),
    "2022_punjab_10.Detailed Results.xlsx": ("punjab", 2022),
    "2022_state_himachal_pradesh_10-Detailed Results.xlsx": (
        "himachal-pradesh",
        2022,
    ),
    "2023_state_karnataka_10-Detailed Results.xlsx": ("karnataka", 2023),
    "2023_rajasthan_Detailed_Results.xlsx": ("rajasthan", 2023),
    "assam_10-Detailed_Results_1778163955.xlsx": ("assam", 2021),
    "kerala_10-Detailed_Results_1778164525.xlsx": ("kerala", 2021),
    "pondy_10-Detailed_Results_1778164807.xlsx": ("puducherry", 2021),
    "tn_10-Detailed_Results_1778165153.xlsx": ("tamil-nadu", 2021),
    "wb_10-Detailed_Results_1779879116.xlsx": ("west-bengal", 2021),
}

# State slug -> ISO 3166-2 short for mint disambiguation suffix.
# Only the 11 targeted states need to be mapped; other states never
# disambiguate via this tool (they don't appear in STATE_XLSX_MAP).
STATE_SLUG_TO_ISO: dict[str, str] = {
    "uttar-pradesh": "UP",
    "gujarat": "GJ",
    "punjab": "PB",
    "himachal-pradesh": "HP",
    "karnataka": "KA",
    "rajasthan": "RJ",
    "assam": "AS",
    "kerala": "KL",
    "puducherry": "PY",
    "tamil-nadu": "TN",
    "west-bengal": "WB",
}

# Verdict CSV schema (matches PR #952 / correlate_unk_apply consumer).
# The ``tcpd_*`` column names are historical: the apply tool was authored
# for TCPD work but reads these columns as opaque source-publisher data,
# so we reuse the same shape for ECI state xlsx here.
VERDICT_FIELDNAMES = [
    "action",
    "proposed_party_id",
    "party_short_raw",
    "tcpd_frequent_abbrev",
    "tcpd_party_name",
    "tcpd_party_type",
    "state",
    "tcpd_start_year",
    "tcpd_last_year",
]

SKIPPED_FIELDNAMES = ["party_short_raw", "n_rows", "reason", "detail"]

NOTA_SENTINEL = "parties.IN.NOTA"
IND_SENTINEL = "parties.IN.IND"

# Threshold for picking the dominant ECI abbr when a publisher's UNK
# label maps to multiple ECI abbreviations across constituencies.
DOMINANT_THRESHOLD = 0.80

HEADER_PROBE_DEPTH = 10

# Year-derivation regex: 4-digit year in 2010-2029 range. Uses non-digit
# lookaround instead of ``\b`` because ``_`` is a regex word character,
# so ``\b`` would not fire between ``2022`` and ``_uttar_pradesh_...``
# (which is the dominant filename shape for the priority 6 files).
YEAR_RE = re.compile(r"(?<!\d)(20[12]\d)(?!\d)")


# --- normalisation ----------------------------------------------------------


def normalise(s: str) -> str:
    """NFKD-ASCII + uppercase + strip-(SC/ST) suffix + collapse punctuation.

    Also strips the ECI ``"<position> "`` numeric prefix on candidate
    names (e.g. ``"10 MUNNA LAL"`` -> ``"MUNNA LAL"``).
    """
    s = unicodedata.normalize("NFKD", s or "")
    # Drop combining marks (e.g. acute that decomposes off of "e\u0301").
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    # Collapse any remaining non-ASCII run (en-dash, em-dash, non-Latin
    # scripts, etc.) to a single space so it does not get silently
    # concatenated across the missing-character gap.
    s = re.sub(r"[^\x00-\x7F]+", " ", s)
    s = s.upper()
    # Strip leading ECI position-prefix like "1 ", "10 ", "117 ".
    s = re.sub(r"^\s*\d+\s+", "", s)
    # Strip ECI category-suffix like " (SC)", " (ST)", " (sc/st)".
    s = re.sub(r"\s*\(\s*(SC|ST|SC/ST|ST/SC)\s*\)\s*$", "", s)
    s = re.sub(r"[^A-Z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _slugify_abbr(abbr: str) -> str:
    """Make ``abbr`` slug-safe (uppercase, ``[A-Z0-9_]`` only)."""
    s = re.sub(r"[^A-Za-z0-9_]+", "_", (abbr or "").upper()).strip("_")
    return s or "UNK"


# --- xlsx loading -----------------------------------------------------------


def derive_year_for_file(
    path: Path, filename: str, hardcoded: int
) -> tuple[int, str]:
    """3-tier year derivation: filename -> xlsx-title -> hardcoded.

    Returns ``(year, source)`` where ``source`` is one of ``"filename"``,
    ``"xlsx-title"``, or ``"hardcoded"`` for the verdict log.
    """
    m = YEAR_RE.search(filename)
    if m:
        return int(m.group(1)), "filename"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > HEADER_PROBE_DEPTH:
                break
            for cell in row:
                if cell is None:
                    continue
                m2 = YEAR_RE.search(str(cell))
                if m2:
                    return int(m2.group(1)), "xlsx-title"
    finally:
        wb.close()
    return hardcoded, "hardcoded"


def _detect_header_row(ws) -> tuple[int, list[str]]:
    """Find the header row containing 'AC NAME' + 'CANDIDATE NAME' + 'PARTY'.

    Returns ``(row_idx, normalised_header_cells)``. Raises ``ValueError``
    if no header is found within ``HEADER_PROBE_DEPTH`` rows.
    """
    rows_seen: list[list[str]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > HEADER_PROBE_DEPTH:
            break
        cells = [str(c or "").strip().upper() for c in row]
        rows_seen.append(cells)
        if (
            "AC NAME" in cells
            and "CANDIDATE NAME" in cells
            and "PARTY" in cells
        ):
            return i, cells
    raise ValueError(
        f"no ECI xlsx header found in first {HEADER_PROBE_DEPTH} rows; "
        f"probed rows: {rows_seen!r}"
    )


def load_eci_xlsx_index(path: Path) -> dict[tuple[str, str], str]:
    """Build ``(norm_ac_name, norm_candidate_name) -> ECI PARTY`` index.

    Duplicate keys keep the first value (within a single xlsx this should
    be vanishingly rare; cross-constituency duplicates resolve correctly
    because the AC name is part of the key).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        header_idx, header = _detect_header_row(ws)
        col_ac = header.index("AC NAME")
        col_cand = header.index("CANDIDATE NAME")
        col_party = header.index("PARTY")
        max_col = max(col_ac, col_cand, col_party)
        out: dict[tuple[str, str], str] = {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i <= header_idx:
                continue
            if len(row) <= max_col:
                continue
            ac = str(row[col_ac] or "").strip()
            cand = str(row[col_cand] or "").strip()
            party = str(row[col_party] or "").strip()
            if not ac or not cand or not party:
                continue
            key = (normalise(ac), normalise(cand))
            out.setdefault(key, party)
    finally:
        wb.close()
    return out


# --- parties.csv loading ----------------------------------------------------


def load_parties_index(
    parties_csv: Path,
) -> tuple[set[str], dict[str, str], dict[str, str]]:
    """Return ``(existing_pids, short_to_pid, alias_to_pid)`` keyed UPPER."""
    existing_pids: set[str] = set()
    short_to_pid: dict[str, str] = {}
    alias_to_pid: dict[str, str] = {}
    with parties_csv.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            pid = (row.get("party_id") or "").strip()
            if not pid:
                continue
            existing_pids.add(pid)
            short = (row.get("short") or "").strip().upper()
            if short:
                short_to_pid[short] = pid
            for a in (row.get("aliases") or "").split("|"):
                a_u = a.strip().upper()
                if a_u:
                    alias_to_pid[a_u] = pid
    return existing_pids, short_to_pid, alias_to_pid


def load_unk_rows_for_state_year(
    repo_root: Path, state_slug: str, year: int
) -> list[dict[str, str]]:
    """Walk the per-(state, year) assembly candidacies.csv; UNK rows only."""
    path = (
        repo_root
        / "datasets"
        / "elections"
        / "assembly"
        / f"state={state_slug}"
        / f"election={year}"
        / "candidacies.csv"
    )
    out: list[dict[str, str]] = []
    if not path.exists():
        return out
    with path.open(encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if (row.get("party_id") or "").strip() == "parties.IN.UNK":
                out.append(
                    {**row, "_state_slug": state_slug, "_year": str(year)}
                )
    return out


# --- resolution -------------------------------------------------------------


def resolve_eci_label(
    eci_label: str,
    existing_pids: set[str],
    short_to_pid: dict[str, str],
    alias_to_pid: dict[str, str],
    rep_year: int,
    rep_state_iso: str,
) -> tuple[str, str, bool]:
    """Resolve an ECI publisher label to a parties.csv pid action.

    Returns ``(action, proposed_pid, was_disambiguated)``. ``action`` is
    ``alias-add`` or ``mint-new``. ``was_disambiguated`` is True when a
    naive ``parties.IN.<abbr>`` would have collided with an existing row
    and the proposal carries an ``_AE<year>_<state>`` (or ``_<n>``) suffix.
    """
    eci_u = (eci_label or "").strip().upper()
    if eci_u == "NOTA":
        return ("alias-add", NOTA_SENTINEL, False)
    if eci_u == "IND" or eci_u == "INDEPENDENT":
        return ("alias-add", IND_SENTINEL, False)
    if eci_u in short_to_pid:
        return ("alias-add", short_to_pid[eci_u], False)
    if eci_u in alias_to_pid:
        return ("alias-add", alias_to_pid[eci_u], False)
    # mint-new
    base = f"parties.IN.{_slugify_abbr(eci_label)}"
    if base not in existing_pids:
        return ("mint-new", base, False)
    # disambiguate via AE<year>_<state>
    disamb = f"{base}_AE{rep_year}_{rep_state_iso}"
    if disamb not in existing_pids:
        return ("mint-new", disamb, True)
    for i in range(2, 100):
        cand = f"{disamb}_{i}"
        if cand not in existing_pids:
            return ("mint-new", cand, True)
    raise RuntimeError(f"unable to disambiguate {base!r}")


# --- correlation main -------------------------------------------------------


def _run_id(verdict_rows: list[dict[str, str]]) -> str:
    """9-char SHA256 of the sorted verdict content (deterministic run id)."""
    if not verdict_rows:
        return "empty0000"
    encoded = "\n".join(
        sorted(
            ",".join(r.get(k, "") for k in VERDICT_FIELDNAMES)
            for r in verdict_rows
        )
    )
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()[:9]


def correlate(
    *,
    files: list[str] | None = None,
    limit: int | None = None,
    repo_root: Path = REPO_ROOT,
    parties_csv: Path | None = None,
    ephemeral_dir: Path | None = None,
    verdict_root: Path | None = None,
    state_xlsx_map: dict[str, tuple[str, int]] | None = None,
) -> dict:
    """Run the correlator end-to-end. Returns stats + output paths.

    ``files`` filters STATE_XLSX_MAP to a subset (None = all). ``limit``
    caps total UNK rows scanned. ``state_xlsx_map`` is a hook for tests
    to inject a different mapping.
    """
    parties_csv = parties_csv or (
        repo_root / "datasets" / "data" / "entities" / "parties.csv"
    )
    ephemeral_dir = ephemeral_dir or (repo_root / "datasets" / "ephemeral")
    verdict_root = verdict_root or (
        repo_root / "datasets" / "ephemeral" / "party-parity" / "eci-state-xlsx"
    )
    xlsx_map = state_xlsx_map or STATE_XLSX_MAP

    if files is not None:
        unknown = [f for f in files if f not in xlsx_map]
        if unknown:
            raise ValueError(
                f"unknown filename(s) (not in STATE_XLSX_MAP): {unknown!r}"
            )
        xlsx_map = {k: v for k, v in xlsx_map.items() if k in files}

    existing_pids, short_to_pid, alias_to_pid = load_parties_index(parties_csv)

    # 1) Per-file ECI xlsx index + year derivation; build a
    #    (state_slug, year) -> {(ac, cand): eci_label} map.
    eci_indexes: dict[tuple[str, int], dict[tuple[str, str], str]] = {}
    year_provenance: dict[str, str] = {}
    per_file_meta: dict[str, dict[str, str | int]] = {}
    for fname, (state_slug, hardcoded) in xlsx_map.items():
        path = ephemeral_dir / fname
        if not path.exists():
            raise FileNotFoundError(
                f"ECI xlsx file missing: "
                f"{path.relative_to(repo_root).as_posix()}; "
                f"upload to datasets/ephemeral/ before running"
            )
        year, source = derive_year_for_file(path, fname, hardcoded)
        year_provenance[fname] = source
        per_file_meta[fname] = {
            "state_slug": state_slug,
            "year": year,
            "year_source": source,
        }
        eci_indexes[(state_slug, year)] = load_eci_xlsx_index(path)

    # 2) Walk UNK candidacies for each (state, year); collect per-label entries.
    #    Aggregate by ``party_short_raw`` across all targeted state-years.
    label_to_entries: dict[
        str, list[tuple[str, int, str | None, dict[str, str]]]
    ] = defaultdict(list)
    per_state_in: Counter[tuple[str, int]] = Counter()
    state_year_pairs = sorted({(s, y) for s, y in eci_indexes.keys()})
    seen = 0
    for state_slug, year in state_year_pairs:
        for row in load_unk_rows_for_state_year(repo_root, state_slug, year):
            if limit is not None and seen >= limit:
                break
            per_state_in[(state_slug, year)] += 1
            seen += 1
            key = (
                normalise(row.get("constituency_name", "")),
                normalise(row.get("candidate_name", "")),
            )
            eci_label = eci_indexes[(state_slug, year)].get(key)
            label = (row.get("party_short_raw") or "").strip()
            label_to_entries[label].append((state_slug, year, eci_label, row))

    # 3) Per-label verdict decision
    verdict_rows: list[dict[str, str]] = []
    skipped_rows: list[dict[str, str]] = []
    per_state_stats: dict[tuple[str, int], Counter[str]] = defaultdict(Counter)
    skip_reason_rows: Counter[str] = Counter()

    for label, entries in sorted(label_to_entries.items()):
        if not label:
            skipped_rows.append(
                {
                    "party_short_raw": "",
                    "n_rows": str(len(entries)),
                    "reason": "empty-publisher-label",
                    "detail": (
                        f"{len(entries)} UNK rows with blank party_short_raw"
                    ),
                }
            )
            skip_reason_rows["empty-publisher-label"] += len(entries)
            for s, y, _, _ in entries:
                per_state_stats[(s, y)]["empty-label-skip"] += 1
            continue

        joined = [(s, y, e, r) for s, y, e, r in entries if e]
        n_total = len(entries)
        n_joined = len(joined)
        if n_joined == 0:
            skipped_rows.append(
                {
                    "party_short_raw": label,
                    "n_rows": str(n_total),
                    "reason": "eci-no-match-on-name",
                    "detail": (
                        f"none of {n_total} rows joined against ECI state "
                        f"xlsx by (constituency, candidate)"
                    ),
                }
            )
            skip_reason_rows["eci-no-match-on-name"] += n_total
            for s, y, _, _ in entries:
                per_state_stats[(s, y)]["unmatched-join"] += 1
            continue

        eci_counter: Counter[str] = Counter(e for _, _, e, _ in joined)
        dominant_label, dominant_count = eci_counter.most_common(1)[0]
        fraction = dominant_count / n_joined
        if fraction < DOMINANT_THRESHOLD:
            skipped_rows.append(
                {
                    "party_short_raw": label,
                    "n_rows": str(n_total),
                    "reason": "eci-internal-collision",
                    "detail": (
                        f"no dominant ECI abbr among {n_joined} joined rows "
                        f"(top {dominant_label!r}={dominant_count}); "
                        f"distribution: {dict(eci_counter)}"
                    ),
                }
            )
            skip_reason_rows["eci-internal-collision"] += n_total
            for s, y, _, _ in entries:
                per_state_stats[(s, y)]["collision-skip"] += 1
            continue

        # Representative (state, year) for disambiguation = the (s, y)
        # pair contributing the most joined rows for this label.
        sy_tally: Counter[tuple[str, int]] = Counter(
            (s, y) for s, y, _, _ in joined
        )
        (rep_state, rep_year), _ = sy_tally.most_common(1)[0]
        rep_state_iso = STATE_SLUG_TO_ISO.get(rep_state, "XX")

        action, proposed_pid, was_disambig = resolve_eci_label(
            dominant_label,
            existing_pids,
            short_to_pid,
            alias_to_pid,
            rep_year,
            rep_state_iso,
        )
        # state column: pipe-delim sorted unique slugs from joined rows.
        # Used by apply tool to project home_state_codes via
        # state_iso_seed.csv.
        states_in_label = sorted({s for s, _, _, _ in joined})
        verdict_rows.append(
            {
                "action": action,
                "proposed_party_id": proposed_pid,
                "party_short_raw": label,
                "tcpd_frequent_abbrev": dominant_label,
                "tcpd_party_name": label,
                "tcpd_party_type": "",
                "state": "|".join(states_in_label),
                "tcpd_start_year": "",
                "tcpd_last_year": "",
            }
        )
        # Refresh local indexes so a subsequent verdict in the same run
        # detects the just-proposed mint as a collision target.
        if action == "mint-new":
            existing_pids.add(proposed_pid)
            short_to_pid[dominant_label.upper()] = proposed_pid

        if was_disambig:
            skipped_rows.append(
                {
                    "party_short_raw": label,
                    "n_rows": str(n_total),
                    "reason": "existing-collision-disambiguated",
                    "detail": (
                        f"naive parties.IN.{_slugify_abbr(dominant_label)} "
                        f"already exists; mint as {proposed_pid}"
                    ),
                }
            )
            # Informational only; not counted as a UNK-row skip.
            skip_reason_rows["existing-collision-disambiguated"] += 0

        bucket = "alias-resolved" if action == "alias-add" else "mint-resolved"
        for s, y, _, _ in joined:
            per_state_stats[(s, y)][bucket] += 1
        # Track the UNJOINED minority of an otherwise-resolved label.
        for s, y, e, _ in entries:
            if e is None:
                per_state_stats[(s, y)]["unmatched-within-resolved-label"] += 1

    # 4) Write outputs
    run_id = _run_id(verdict_rows)
    out_dir = verdict_root / run_id
    out_dir.mkdir(parents=True, exist_ok=True)
    verdict_path = out_dir / "verdict.csv"
    skipped_path = out_dir / "skipped.csv"
    with verdict_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(
            f, fieldnames=VERDICT_FIELDNAMES, lineterminator="\n"
        )
        w.writeheader()
        w.writerows(verdict_rows)
    with skipped_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(
            f, fieldnames=SKIPPED_FIELDNAMES, lineterminator="\n"
        )
        w.writeheader()
        w.writerows(skipped_rows)

    return {
        "verdict_path": verdict_path,
        "skipped_path": skipped_path,
        "verdict_rows": len(verdict_rows),
        "skipped_rows": len(skipped_rows),
        "per_state_in": {f"{s}-{y}": n for (s, y), n in per_state_in.items()},
        "per_state_stats": {
            f"{s}-{y}": dict(c) for (s, y), c in per_state_stats.items()
        },
        "skip_reasons": dict(skip_reason_rows),
        "year_provenance": dict(year_provenance),
        "per_file_meta": per_file_meta,
        "run_id": run_id,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--files",
        default=None,
        help=(
            "Comma-separated xlsx basenames to process (default: all 11 in "
            "STATE_XLSX_MAP). Filenames must match keys exactly."
        ),
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Limit total UNK rows scanned (for fast smoke runs).",
    )
    args = parser.parse_args()
    files = None
    if args.files:
        files = [f.strip() for f in args.files.split(",") if f.strip()]

    result = correlate(files=files, limit=args.limit)

    print(
        f"verdict.csv: {result['verdict_path'].relative_to(REPO_ROOT).as_posix()}"
    )
    print(
        f"skipped.csv: {result['skipped_path'].relative_to(REPO_ROOT).as_posix()}"
    )
    print(f"run_id:      {result['run_id']}")
    print()
    print("=== per-file year derivation ===")
    for fname, meta in sorted(result["per_file_meta"].items()):
        print(
            f"  {fname}: state={meta['state_slug']}, year={meta['year']} "
            f"({meta['year_source']})"
        )
    print()
    print("=== per-(state, year) stats ===")
    for sy in sorted(result["per_state_in"].keys()):
        n_in = result["per_state_in"][sy]
        bucket = result["per_state_stats"].get(sy, {})
        bucket_s = ", ".join(f"{k}={v}" for k, v in sorted(bucket.items()))
        print(f"  {sy}: UNK in = {n_in}, {bucket_s}")
    print()
    print(f"  verdict rows: {result['verdict_rows']}")
    print(f"  skipped rows: {result['skipped_rows']}")
    print()
    print("=== skip-reason histogram (in UNK candidacy rows) ===")
    for r, n in sorted(result["skip_reasons"].items()):
        print(f"  {r}: {n}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
