"""Quick shape-inspect of files in datasets/ephemeral_datasets/ for ICED ingest planning.

Per CLAUDE.md anti-patterns this is a tools/ recon script, not part of backend/ pipeline.
"""
from __future__ import annotations
import io
import sys
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

ROOT = Path(__file__).resolve().parents[1]
EPH = ROOT / "datasets" / "ephemeral_datasets"


def inspect(p: Path) -> None:
    print(f"\n=== {p.name} ({p.stat().st_size:,} bytes) ===")
    try:
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    except Exception as e:
        print(f"  OPEN-FAIL: {type(e).__name__}: {e}")
        return
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        print(f"  sheet={sn!r}  rows={len(rows)}  cols={ws.max_column}")
        for i, r in enumerate(rows[:6]):
            cells = [("" if c is None else str(c))[:25] for c in r[:8]]
            print(f"    r{i}: {cells}")
        if len(rows) > 6:
            print(f"    ... +{len(rows)-6} more")


def main() -> None:
    files = sorted(EPH.glob("*.xlsx"))
    print(f"Found {len(files)} xlsx files under datasets/ephemeral_datasets/")
    for f in files:
        inspect(f)


if __name__ == "__main__":
    main()
