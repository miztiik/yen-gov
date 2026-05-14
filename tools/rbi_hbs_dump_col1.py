"""Dump full per-row contents of RBI HBS-IE Table 5 sheet to map base-year sections."""
from __future__ import annotations
import io, sys
from pathlib import Path
import openpyxl
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
xlsx = Path(sys.argv[1])
wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"\n=== {xlsx.name} :: {sname} ===")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        c0 = row[1] if len(row) > 1 else None
        print(f"  {i}: {c0!r}")
wb.close()
