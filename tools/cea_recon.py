"""Recon for CEA installed-capacity workbook."""
from __future__ import annotations

import io
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

PATH = Path(__file__).resolve().parents[1] / ".runtime" / "raw" / "cea" / "installed_capacity_2026_03.xlsx"
wb = load_workbook(PATH, data_only=True)
print("Sheets:", wb.sheetnames)
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"\n=== {sn} (dims={ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column}) ===")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80))):
        vals = [c.value for c in row]
        if any(v is not None for v in vals):
            print(i, vals[:14])
