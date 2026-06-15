"""ECI Form 10 "Detailed Results" multi-state Assembly Election ingest.

Reads the ECI-published per-state Form-10 Detailed Results xlsx workbooks
(operator-dropped under ``datasets/ephemeral/``) and emits canonical per-event
candidacies.csv + summary.csv into ``datasets/elections/assembly/state=<slug>/
election=<year>/``.

Lights up 14 Pending state-events from the /t/elections firehose:
  - 2023: Telangana, Madhya Pradesh, Mizoram, Chhattisgarh
  - 2024: Andhra Pradesh, Arunachal Pradesh, Sikkim, Odisha, Jammu & Kashmir,
          Haryana, Jharkhand, Maharashtra
  - 2025: Delhi, Bihar

Schema of the upstream xlsx (verified 2026-06-14 against 13 files):

  Row 1: title "10 - Detailed Results"
  Row 2: blank
  Row 3: section headers (TOTAL VALID VOTES POLLED / VALID VOTES POLLED ...)
  Row 4: column header
  Rows 5+: candidate rows + "TURN OUT" marker rows

Column layout (14 or 15 cols depending on vintage):
  STATE/UT NAME | AC NO. | AC NAME | CANDIDATE NAME | GENDER|SEX | AGE |
  CATEGORY | PARTY | SYMBOL | GENERAL | POSTAL | TOTAL | <pct cols> |
  [TOTAL ELECTORS]

CANDIDATE NAME is the publisher form "<serial> <name>"
  (e.g. "1 AMSHYA FULJI PADVI") - we strip the leading numeric serial.

NOTA rows: party=="NOTA", GENDER/AGE/CATEGORY are blank. Excluded from
candidacies (per Rajasthan precedent; NOTA aggregates flow through
summary.electors/votes_polled separately).

TURN OUT rows: STATE/UT NAME=="TURN OUT" - marker for per-AC totals. We
detect these to carry per-AC TOTAL ELECTORS / votes_polled / turnout
that don't otherwise appear on candidate rows.

Per-AC summary derivation uses canonical
``yen_gov.canonical.reingest.assembly_results.recompute_summary_row`` so
the G15 contract test (``summary == recompute(candidacies)``) passes by
construction. Aggregate fields (electors, votes_polled, turnout_pct)
come from the ECI per-candidate TOTAL ELECTORS column + TURN OUT marker
row.
"""

from __future__ import annotations

import csv
import json
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Final

import openpyxl

from yen_gov.canonical.citation import derive_source_id
from yen_gov.canonical.party_resolver import load_resolver
from yen_gov.canonical.processing_quality import derive_processing
from yen_gov.canonical.reingest.assembly_results import recompute_summary_row


ECI_VINTAGE: Final[str] = "2026-06-14"
ECI_PRODUCER: Final[str] = "Election Commission of India"
ECI_TITLE_FMT: Final[str] = (
    "Form 10 Detailed Results - {state_display} Legislative Assembly Election {year}"
)


@dataclass(frozen=True)
class EciEventSpec:
    """One ECI ingest job: ephemeral file path + canonical event metadata."""

    file_name: str  # under datasets/ephemeral/
    state_slug: str  # canonical electoral.csv state slug
    state_code: str  # S/U code for election_events.json catalogue
    state_display: str  # for the source.csv title field
    expected_state_name: str  # value in xlsx col 1 (e.g. "Bihar", "Jammu & Kashmir")
    election_year: int
    event_id: str  # catalogue event_id (e.g. "assembly-2024")
    polled_on: str  # ISO date matched against the catalogue entry
    period_label: str  # for citation cross-link (e.g. "AcGenNov2024")


# Catalogue of all ECI ingest jobs. Ordered roughly newest -> oldest so the
# operator-visible CLI output reads chronologically.
JOBS: tuple[EciEventSpec, ...] = (
    EciEventSpec(
        "2025_BIHAR_10-Detailed_Results_1763549630.xlsx",
        "bihar", "S04", "Bihar", "Bihar",
        2025, "assembly-2025", "2025-11-11", "AcGenNov2025",
    ),
    EciEventSpec(
        "2025_DL_10-Detailed_Results_1744913508.xlsx",
        "delhi", "U05", "NCT of Delhi", "NCT OF Delhi",
        2025, "assembly-2025", "2025-02-05", "AcGenFeb2025",
    ),
    EciEventSpec(
        "2024_MH_10-Detailed_Results_1744893339.xlsx",
        "maharashtra", "S13", "Maharashtra", "Maharashtra",
        2024, "assembly-2024", "2024-11-20", "AcGenNov2024",
    ),
    EciEventSpec(
        "2024_jharkhand_10-Detailed_Results_1744892172.xlsx",
        "jharkhand", "S27", "Jharkhand", "Jharkhand",
        2024, "assembly-2024", "2024-11-20", "AcGenNov2024",
    ),
    EciEventSpec(
        "2024_haryana_10-Detailed-Results.xlsx",
        "haryana", "S07", "Haryana", "Haryana",
        2024, "assembly-2024", "2024-10-05", "AcGenOct2024",
    ),
    EciEventSpec(
        "2024_jk_10-Detailed-Results.xlsx",
        "jammu-and-kashmir", "U08", "Jammu and Kashmir (UT)", "Jammu & Kashmir",
        2024, "assembly-2024", "2024-10-01", "AcGenOct2024",
    ),
    EciEventSpec(
        "2024_odisha_10-Detailed-Results.xlsx",
        "odisha", "S18", "Odisha", "Odisha",
        2024, "assembly-2024", "2024-06-01", "AcGenJun2024",
    ),
    EciEventSpec(
        "2024_AP_10-Detailed-Results.xlsx",
        "andhra-pradesh", "S01", "Andhra Pradesh", "Andhra Pradesh",
        2024, "assembly-2024", "2024-05-13", "AcGenJun2024",
    ),
    EciEventSpec(
        "2024_Arunachal_10-Detailed-Results.xlsx",
        "arunachal-pradesh", "S02", "Arunachal Pradesh", "Arunachal Pradesh",
        2024, "assembly-2024", "2024-04-19", "AcGenJun2024",
    ),
    EciEventSpec(
        "2024_sikkim_10-Detailed-Results.xlsx",
        "sikkim", "S21", "Sikkim", "Sikkim",
        2024, "assembly-2024", "2024-04-19", "AcGenJun2024",
    ),
    EciEventSpec(
        "2023_telengana_Detailed_Results.xlsx",
        "telangana", "S29", "Telangana", "Telangana",
        2023, "assembly-2023", "2023-11-30", "AcGenNov2023",
    ),
    EciEventSpec(
        "2023_MadhyaPrashesh_Detailed_Results.xlsx",
        "madhya-pradesh", "S12", "Madhya Pradesh", "Madhya Pradesh",
        2023, "assembly-2023", "2023-11-17", "AcGenNov2023",
    ),
    EciEventSpec(
        "2023_mizoram_Detailed_Results.xlsx",
        "mizoram", "S16", "Mizoram", "Mizoram",
        2023, "assembly-2023", "2023-11-07", "AcGenNov2023",
    ),
    EciEventSpec(
        "2023_Chattisgargh_Detailed_Results.xlsx",
        "chhattisgarh", "S26", "Chhattisgarh", "Chhattisgarh",
        2023, "assembly-2023", "2023-11-07", "AcGenNov2023",
    ),
)


# Pattern matching candidate name strings like "1 AMSHYA FULJI PADVI" or
# "10 Surendra Prasad" - the leading serial is publisher metadata, not name.
_CAND_SERIAL_RE = re.compile(r"^\s*(\d+)\s+(.+?)\s*$")


def _norm_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _to_int(v: Any) -> int:
    s = _norm_str(v)
    if not s:
        return 0
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 0


def _to_float_or_none(v: Any) -> float | None:
    s = _norm_str(v)
    if not s or s == "-":
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _strip_serial(candidate_name_cell: str) -> str:
    """Strip leading publisher serial from ``"1 AMSHYA FULJI PADVI"`` -> ``"AMSHYA FULJI PADVI"``."""
    m = _CAND_SERIAL_RE.match(candidate_name_cell)
    if m:
        return m.group(2).strip()
    return candidate_name_cell.strip()


def _normalise_sex(v: Any) -> str:
    """ECI uses MALE/FEMALE/OTHERS/blank. Map to single letter per Rajasthan precedent."""
    s = _norm_str(v).upper()
    if s.startswith("M"):
        return "M"
    if s.startswith("F"):
        return "F"
    if s.startswith("O"):
        return "O"
    return "U"


def _normalise_category(v: Any) -> str:
    """ECI: GENERAL/SC/ST/blank. Map to per-candidate categorisation tier."""
    s = _norm_str(v).upper()
    if s in ("SC", "ST", "GENERAL", "GEN"):
        return s if s != "GEN" else "GENERAL"
    return ""


def _load_electoral_entities(
    root: Path, state: str, delim_year: str
) -> dict[int, list[tuple[str, str]]]:
    """Index electoral.csv: eci_no -> [(entity_id, name), ...] for (state, delim, ac)."""
    csv_path = root / "datasets" / "data" / "entities" / "electoral.csv"
    result: dict[int, list[tuple[str, str]]] = defaultdict(list)

    with csv_path.open(encoding="utf-8", newline="") as fh:
        for row in csv.DictReader(fh):
            if (row.get("entity_kind") or "").strip() != "ac":
                continue
            if (row.get("state") or "").strip() != state:
                continue
            if (row.get("delim_year") or "").strip() != delim_year:
                continue
            try:
                eci_no = int(_norm_str(row.get("eci_no")))
            except ValueError:
                continue
            entity_id = _norm_str(row.get("entity_id"))
            name = _norm_str(row.get("name"))
            if entity_id and name:
                result[eci_no].append((entity_id, name))

    return dict(result)


def _resolve_entity(
    eci_no: int,
    publisher_name: str,
    entities: dict[int, list[tuple[str, str]]],
) -> tuple[str, str] | None:
    """Disambiguate (eci_no) -> single (entity_id, name) via name match."""
    candidates = entities.get(eci_no, [])
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]
    norm_pub = re.sub(r"\s+", "", publisher_name.upper())
    norm_pub = re.sub(r"\(.*\)", "", norm_pub)  # strip (ST)/(SC) qualifier
    for entity_id, name in candidates:
        norm_cand = re.sub(r"\s+", "", name.upper())
        if (
            norm_pub == norm_cand
            or norm_pub.startswith(norm_cand)
            or norm_cand.startswith(norm_pub)
        ):
            return (entity_id, name)
    return None


def _get_or_mint_eci_source(root: Path, job: EciEventSpec) -> str:
    """Return source_id for the ECI Form-10 citation triple for this event."""
    csv_path = root / "datasets" / "data" / "entities" / "source.csv"
    rows: list[dict] = []
    if csv_path.exists():
        with csv_path.open(encoding="utf-8", newline="") as fh:
            rows = list(csv.DictReader(fh))

    title = ECI_TITLE_FMT.format(state_display=job.state_display, year=job.election_year)
    for row in rows:
        if (
            _norm_str(row.get("producer")) == ECI_PRODUCER
            and _norm_str(row.get("title")) == title
            and _norm_str(row.get("vintage")) == ECI_VINTAGE
        ):
            return _norm_str(row.get("source_id"))

    source_id = derive_source_id(
        producer=ECI_PRODUCER,
        title=title,
        vintage=ECI_VINTAGE,
    )
    new_row = {
        "source_id": source_id,
        "producer": ECI_PRODUCER,
        "title": title,
        "vintage": ECI_VINTAGE,
        "url": "https://results.eci.gov.in/",
    }
    with csv_path.open("a", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(
            fh, fieldnames=["source_id", "producer", "title", "vintage", "url"]
        )
        writer.writerow(new_row)
    return source_id


# ---------------------------------------------------------------------------
# Stub-minting for AC entities missing from electoral.csv (Path B3, 2026-06-15
# Hans + Max + Fowler persona-debate verdict). When the ECI Form 10 XLSX
# carries an AC NO. that has no matching row in
# datasets/data/entities/electoral.csv for (state, delim_year=2008), we
# auto-mint a stub electoral entity_id of the form
# ``IN-AC-2008-<state>-eci<N>`` (the convention already in production for
# AP, Assam, Karnataka pre-2026-06-15).
#
# - name: XLSX publisher AC NAME with the trailing (SC)/(ST) reservation
#   suffix stripped.
# - reservation: derived from the stripped suffix; defaults to "GEN".
# - parent: empty (PC parent code unknown until LGD master directory
#   reconciles; the schema permits empty here).
# - aliases: empty.
#
# This preserves writer idempotence: a stub minted on the first ingest is
# resolved by ``_load_electoral_entities`` on a subsequent run, so
# ``missing_from_entities`` is empty and no second mint occurs.
#
# Every stub-mint also appends a row to
# ``datasets/_ops/electoral-stub-mints.csv`` as an operator-visible
# receipt naming the LGD gap and the XLSX source.
# ---------------------------------------------------------------------------


_RESERVATION_SUFFIX_RE = re.compile(r"\s*\((SC|ST)\)\s*$", re.IGNORECASE)


def _mint_stub_entities(
    missing_eci_nos: list[int],
    by_ac: dict[int, list[dict]],
    state_slug: str,
    delim_year: str,
) -> list[dict]:
    """Build electoral.csv stub rows for AC eci_nos missing from electoral.csv.

    Pure function: takes the list of XLSX eci_nos that aren't in
    electoral.csv and returns a list of dicts ready to be appended.
    Mints under the established ``IN-AC-2008-<state>-eci<N>`` convention.
    """
    stubs: list[dict] = []
    for eci_no in sorted(missing_eci_nos):
        cands = by_ac.get(eci_no, [])
        if not cands:
            continue  # defensive; should not happen if eci_no was in by_ac.keys()
        raw_name = cands[0]["publisher_ac_name"].strip()
        reservation = "GEN"
        match = _RESERVATION_SUFFIX_RE.search(raw_name)
        if match:
            reservation = match.group(1).upper()
        clean_name = _RESERVATION_SUFFIX_RE.sub("", raw_name).strip()
        stubs.append({
            "entity_id": f"IN-AC-{delim_year}-{state_slug}-eci{eci_no}",
            "name": clean_name,
            "entity_kind": "ac",
            "delim_year": delim_year,
            "state": state_slug,
            "parent": "",
            "eci_no": str(eci_no),
            "aliases": "",
            "reservation": reservation,
        })
    return stubs


def _append_stubs_to_electoral_csv(stubs: list[dict], root: Path) -> None:
    """Append stub rows to datasets/data/entities/electoral.csv. No-op if empty."""
    if not stubs:
        return
    csv_path = root / "datasets" / "data" / "entities" / "electoral.csv"
    fields = [
        "entity_id", "name", "entity_kind", "delim_year", "state",
        "parent", "eci_no", "aliases", "reservation",
    ]
    with csv_path.open("a", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writerows(stubs)


def _append_stubs_to_receipt(
    stubs: list[dict],
    job: EciEventSpec,
    root: Path,
) -> None:
    """Append operator-visible receipt rows to datasets/_ops/electoral-stub-mints.csv.

    The receipt records every stub minted: which eci_no, AC NAME, source
    XLSX, mint date. Used by the curator-reconciliation flow when the LGD
    master directory next syncs.
    """
    if not stubs:
        return
    receipt_path = root / "datasets" / "_ops" / "electoral-stub-mints.csv"
    receipt_path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "entity_id", "eci_no", "ac_name", "reservation",
        "mint_date", "source_xlsx", "gap_reason",
    ]
    from datetime import date

    today = date.today().isoformat()
    existed = receipt_path.exists()
    with receipt_path.open("a", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        if not existed:
            writer.writeheader()
        for stub in stubs:
            writer.writerow({
                "entity_id": stub["entity_id"],
                "eci_no": stub["eci_no"],
                "ac_name": stub["name"],
                "reservation": stub["reservation"],
                "mint_date": today,
                "source_xlsx": job.file_name,
                "gap_reason": (
                    "LGD master directory lacks this AC for the 2008 delim; "
                    "pending upstream sync"
                ),
            })


@dataclass
class IngestResult:
    state_slug: str
    election_year: int
    event_id: str
    file_name: str
    status: str  # "ok" | "skip" | "fail"
    n_candidacies: int = 0
    n_summary: int = 0
    n_unresolved_winners: int = 0
    n_missing_acs: int = 0
    expected_acs: int = 0
    reason: str = ""


def _ingest_one(root: Path, job: EciEventSpec) -> IngestResult:
    """Ingest one ECI Form-10 file. Returns IngestResult either way."""
    xlsx_path = root / "datasets" / "ephemeral" / job.file_name
    if not xlsx_path.exists():
        return IngestResult(
            job.state_slug, job.election_year, job.event_id, job.file_name,
            status="fail", reason=f"file not found: {xlsx_path}",
        )

    electoral_entities = _load_electoral_entities(root, job.state_slug, "2008")
    if not electoral_entities:
        return IngestResult(
            job.state_slug, job.election_year, job.event_id, job.file_name,
            status="fail",
            reason=f"electoral.csv has no AC entities for state={job.state_slug} delim=2008",
        )

    parties_csv = root / "datasets" / "data" / "entities" / "parties.csv"
    resolver = load_resolver(parties_csv)
    source_id = _get_or_mint_eci_source(root, job)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    # Group rows by AC number. Skip header rows (1-4), TURN OUT marker rows.
    by_ac: dict[int, list[dict]] = defaultdict(list)
    ac_turn_out: dict[int, dict] = {}
    current_ac: int | None = None
    n_cols = ws.max_column

    for row in ws.iter_rows(min_row=5, values_only=True):
        cells = list(row) + [None] * (15 - len(row))  # pad to 15 for safe indexing
        state_cell = _norm_str(cells[0])
        ac_cell = _norm_str(cells[1])

        # TURN OUT marker row: "TURN OUT" | "" | "" | "" | "" | "" | "TOTAL:" | "" | "" | general | postal | total | "-" | turnout_pct | ""
        if state_cell.upper().startswith("TURN OUT") or state_cell.upper().startswith("TURNOUT"):
            if current_ac is not None:
                # Capture totals from the TURN OUT row
                general = _to_int(cells[9])
                postal = _to_int(cells[10])
                total = _to_int(cells[11])
                # Turnout % cell varies: pos 12 (14-col) or pos 13 (15-col)
                # Try last col first then fall back
                turnout_pct = _to_float_or_none(cells[13]) if n_cols >= 15 else _to_float_or_none(cells[12])
                ac_turn_out[current_ac] = {
                    "votes_polled": total if total > 0 else general + postal,
                    "turnout_pct": turnout_pct,
                }
            current_ac = None
            continue

        if not ac_cell:
            continue

        try:
            ac_no = int(ac_cell)
        except ValueError:
            continue

        current_ac = ac_no
        cand_cell = _norm_str(cells[3])
        if not cand_cell:
            continue

        # NOTA candidate row: party=="NOTA". Treat as a non-candidate vote
        # bucket - exclude from candidacies but capture for completeness.
        party_raw = _norm_str(cells[7])
        if party_raw.upper() in ("NOTA", "NONE OF THE ABOVE"):
            continue

        candidate_name = _strip_serial(cand_cell)
        # Some files have leading capitalised "1 Surendra Prasad" - preserve as-is
        sex = _normalise_sex(cells[4])
        age = _to_int(cells[5])
        category = _normalise_category(cells[6])
        symbol = _norm_str(cells[8])
        general_votes = _to_int(cells[9])
        postal_votes = _to_int(cells[10])
        total_votes = _to_int(cells[11])
        if total_votes == 0:
            total_votes = general_votes + postal_votes

        # TOTAL ELECTORS: last column (col index 14 for 15-col files, col 13
        # for 14-col files). The 14-col files put it at position 13.
        electors_raw = cells[n_cols - 1]
        electors = _to_int(electors_raw)

        by_ac[ac_no].append({
            "candidate_name": candidate_name,
            "party_raw": party_raw,
            "sex": sex,
            "age": age,
            "category": category,
            "symbol": symbol,
            "votes": total_votes,
            "general_votes": general_votes,
            "postal_votes": postal_votes,
            "electors": electors,
            "publisher_ac_name": _norm_str(cells[2]),
        })

    wb.close()

    if not by_ac:
        return IngestResult(
            job.state_slug, job.election_year, job.event_id, job.file_name,
            status="fail", reason="zero candidate rows parsed",
        )

    # Auto-mint electoral.csv stubs for any XLSX eci_no that has no row in
    # electoral.csv for (state, delim_year=2008). Path B3 of the 2026-06-15
    # Hans + Max + Fowler persona-debate verdict; reuses the
    # ``IN-AC-2008-<state>-eci<N>`` convention already in production for AP,
    # Assam, Karnataka. Idempotent: a re-run finds the stubs in
    # electoral.csv on the next ``_load_electoral_entities`` call so
    # ``missing_from_entities`` is empty and no second mint occurs.
    missing_from_entities = sorted(
        ac_no for ac_no in by_ac if ac_no not in electoral_entities
    )
    if missing_from_entities:
        stubs = _mint_stub_entities(
            missing_from_entities, by_ac, job.state_slug, "2008"
        )
        if stubs:
            _append_stubs_to_electoral_csv(stubs, root)
            _append_stubs_to_receipt(stubs, job, root)
            electoral_entities = _load_electoral_entities(
                root, job.state_slug, "2008"
            )

    candidacies_rows: list[dict] = []
    summary_rows: list[dict] = []
    unresolved_winners = 0
    missing_acs: list[int] = []

    for ac_no in sorted(by_ac.keys()):
        cands = by_ac[ac_no]
        publisher_name = cands[0]["publisher_ac_name"]
        resolved = _resolve_entity(ac_no, publisher_name, electoral_entities)
        if resolved is None:
            missing_acs.append(ac_no)
            continue

        entity_id, ac_name = resolved

        # Sort by vote total desc; build candidacies rows
        ranked = sorted(cands, key=lambda d: -d["votes"])
        non_nota_total = sum(c["votes"] for c in ranked)

        ac_cand_rows = []
        for position, cand in enumerate(ranked, start=1):
            party_id = resolver.resolve(
                party_short=cand["party_raw"],
                eci_code=None,
                is_nota=False,
                is_independent=cand["party_raw"].upper() == "IND"
                or cand["party_raw"].upper() == "INDEPENDENT",
            )
            if position == 1 and party_id.endswith(".UNK"):
                unresolved_winners += 1

            share = (cand["votes"] / non_nota_total * 100) if non_nota_total > 0 else 0.0
            proc_level, proc_note = derive_processing(party_id, cand["party_raw"])

            row = {
                "entity_id": entity_id,
                "state": job.state_slug,
                "election_year": job.election_year,
                "constituency_no": ac_no,
                "constituency_name": ac_name,
                "candidate_name": cand["candidate_name"],
                "party_id": party_id,
                "party_short_raw": cand["party_raw"],
                "votes": cand["votes"],
                "vote_share_pct": round(share, 4),
                "position": position,
                "result": "won" if position == 1 else "lost",
                "sex": cand["sex"],
                "age": cand["age"] if cand["age"] > 0 else "",
                "education": "",
                "profession": "",
                "candidate_type": "challenger",
                "source_id": source_id,
                "processing_level": proc_level,
                "processing_note": proc_note,
            }
            candidacies_rows.append(row)
            ac_cand_rows.append(row)

        # Per-AC summary via canonical reducer
        first_cand_electors = next((c["electors"] for c in cands if c["electors"] > 0), 0)
        turn_out = ac_turn_out.get(ac_no, {})
        votes_polled = turn_out.get("votes_polled") or sum(c["votes"] for c in cands)
        turnout_pct = turn_out.get("turnout_pct")
        if turnout_pct is None and first_cand_electors > 0:
            turnout_pct = round((votes_polled / first_cand_electors) * 100, 2)

        ac_facts = {
            "electors": first_cand_electors if first_cand_electors > 0 else None,
            "votes_polled": votes_polled if votes_polled > 0 else None,
            "turnout_pct": turnout_pct,
        }

        summary_row = recompute_summary_row(
            entity_id=entity_id,
            state_slug=job.state_slug,
            election_year=job.election_year,
            candidacy_rows=ac_cand_rows,
            ac_facts=ac_facts,
            source_id=source_id,
        )
        summary_rows.append(summary_row)

    # Write per-event CSVs
    out_dir = (
        root
        / "datasets"
        / "elections"
        / "assembly"
        / f"state={job.state_slug}"
        / f"election={job.election_year}"
    )
    out_dir.mkdir(parents=True, exist_ok=True)

    cand_fields = [
        "entity_id", "state", "election_year", "constituency_no",
        "constituency_name", "candidate_name", "party_id", "party_short_raw",
        "votes", "vote_share_pct", "position", "result", "sex", "age",
        "education", "profession", "candidate_type", "source_id",
        "processing_level", "processing_note",
    ]
    with (out_dir / "candidacies.csv").open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=cand_fields)
        writer.writeheader()
        writer.writerows(candidacies_rows)

    sum_fields = [
        "entity_id", "state", "election_year", "constituency_name",
        "electors", "votes_polled", "turnout_pct", "winner_candidate",
        "winner_party_id", "winner_party_short_raw", "winner_votes",
        "winner_share_pct", "runnerup_candidate", "runnerup_party_id",
        "runnerup_party_short_raw", "runnerup_votes", "margin_votes",
        "margin_pct", "source_id", "processing_level", "processing_note",
    ]
    with (out_dir / "summary.csv").open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=sum_fields)
        writer.writeheader()
        writer.writerows(summary_rows)

    return IngestResult(
        state_slug=job.state_slug,
        election_year=job.election_year,
        event_id=job.event_id,
        file_name=job.file_name,
        status="ok",
        n_candidacies=len(candidacies_rows),
        n_summary=len(summary_rows),
        n_unresolved_winners=unresolved_winners,
        n_missing_acs=len(missing_acs),
        expected_acs=max(by_ac.keys()),
    )


def _flip_catalogue(root: Path, results: list[IngestResult]) -> int:
    """Flip data_status to 'complete' for every successfully-ingested event.

    Returns count of catalogue entries flipped.
    """
    events_path = root / "datasets" / "taxonomy" / "election_events.json"
    with events_path.open(encoding="utf-8") as fh:
        events_data = json.load(fh)

    # Build (state_code, event_id, polled_on) -> ingest_status map
    job_by_state_event: dict[tuple[str, str, str], str] = {}
    for job in JOBS:
        for r in results:
            if (
                r.state_slug == job.state_slug
                and r.election_year == job.election_year
                and r.event_id == job.event_id
            ):
                job_by_state_event[(job.state_code, job.event_id, job.polled_on)] = r.status
                break

    flipped = 0
    for state_code, events in events_data.get("states", {}).items():
        for ev in events:
            key = (state_code, ev.get("event_id", ""), ev.get("polled_on", ""))
            if key in job_by_state_event and job_by_state_event[key] == "ok":
                if ev.get("data_status") == "pending_upstream":
                    ev["data_status"] = "complete"
                    flipped += 1

    if flipped > 0:
        with events_path.open("w", encoding="utf-8") as fh:
            json.dump(events_data, fh, indent=2, ensure_ascii=False)
            fh.write("\n")

    return flipped


def ingest_all(root: Path) -> tuple[list[IngestResult], int]:
    """Ingest all 13 ECI Form-10 files; flip catalogue; return (results, flipped_count)."""
    results: list[IngestResult] = []
    for job in JOBS:
        try:
            r = _ingest_one(root, job)
        except Exception as e:
            r = IngestResult(
                state_slug=job.state_slug,
                election_year=job.election_year,
                event_id=job.event_id,
                file_name=job.file_name,
                status="fail",
                reason=f"{type(e).__name__}: {e}",
            )
        results.append(r)

    flipped = _flip_catalogue(root, results)
    return results, flipped
