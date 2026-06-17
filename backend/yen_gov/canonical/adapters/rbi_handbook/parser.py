"""RBI Handbook of Statistics on Indian States - wide-table parser.

The Handbook publishes one ``state x period`` matrix per indicator
(Birth Rate, Death Rate, Total Fertility Rate, Infant Mortality Rate,
Life Expectancy, ...). Each XLSX sheet has:

  - a banner / title row or two,
  - a header row whose first labelled cell reads "State" / "State/UT"
    and whose subsequent cells are period labels (calendar years like
    "2016", or multi-year windows like "2016-20"),
  - one data row per state / UT, the first cell carrying an optional
    ordinal prefix ("1. Andhra Pradesh"),
  - optional aggregate / note / source rows at the bottom.

This module is layout-driven by :class:`HbsTableSpec`: the header row
and the period columns are AUTO-DETECTED, so adding a new table means
writing one spec (see ``registry.py``), not editing the parser. The
melt to long format is one generic function that serves every
``state x period`` Handbook table across all sections (Social and
Demographic today; State Domestic Product, Fiscal, Banking, Prices
tomorrow).

No network. Reads operator-staged workbook bytes only (parent plan
section 21.4: ingest reads local source files; the network fetcher was
deleted in the rip).
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field

from openpyxl import load_workbook

from .resolver import COUNTRY_ENTITY_ID, StateResolver, normalise_label

__all__ = [
    "HbsTableSpec",
    "LongRow",
    "RbiHbsShapeError",
    "parse_hbs_workbook",
]

# Calendar-year header: exactly four digits ("2016", "2024").
_CAL_YEAR_RE = re.compile(r"^\s*(\d{4})\s*$")
# Multi-year window header: "2016-20", "2016-2020", "2016 - 20", "2016/20".
_WINDOW_RE = re.compile(r"(\d{4})\s*[-/]\s*(\d{2,4})")
# Leading signed number with optional decimals (tolerates a trailing
# publisher footnote like "70.3 P" / "1.5*").
_LEADING_NUM_RE = re.compile(r"^[-+]?\d+(?:\.\d+)?")

# Cell contents that mean "no observation" -> dropped (sparse-safe).
_NA_MARKERS: frozenset[str] = frozenset(
    {
        "",
        "-",
        "--",
        "n.a.",
        "na",
        "n.a",
        "na.",
        "neg",
        "neg.",
        "nr",
        "...",
        "..",
    }
)

# Time-axis interpretation modes a spec can select.
TIME_CALENDAR_YEAR = "calendar_year"
TIME_INTERVAL_WINDOW_END = "interval_window_end"
_TIME_KINDS = frozenset({TIME_CALENDAR_YEAR, TIME_INTERVAL_WINDOW_END})


class RbiHbsShapeError(ValueError):
    """The workbook no longer matches its spec.

    Raised loud (never emit zero rows) so a layout drift surfaces to the
    operator instead of silently dropping a state-year cell - a silent
    coverage drop would lie to the citizen.
    """


@dataclass(frozen=True)
class HbsTableSpec:
    """One RBI Handbook table -> one canonical indicator.

    A single spec carries everything three downstream surfaces need: the
    workbook layout (parser), the catalogue rows (``variables.csv`` +
    ``concepts.csv``), and the citation triple (``source.csv``). A future
    agent adds a new Handbook table by appending one spec - no parser
    edits.
    """

    # --- identity / output (the variables.csv + concepts.csv rows) ---
    indicator_id: str          # flat kebab; = the datapoint filename stem
    name: str                  # citizen-facing label (variables.csv.name)
    concept_id: str            # FK -> concepts.csv
    concept_noun: str          # concepts.csv.noun
    concept_description: str   # concepts.csv.description (one honest caveat)
    unit: str                  # variables.csv.unit (display)
    unit_canonical: str        # concepts.csv.unit_canonical
    normalisation: str         # concepts.csv enum: absolute|per_capita|per_area|share|ratio|index
    topic: str                 # FK -> topics.csv
    entity_kinds: str          # concepts.csv.entity_kinds (space-joined, e.g. "country state")
    update_period_days: int    # publisher refresh cadence

    # --- provenance (source.csv row; source_id is DERIVED, never set) ---
    source_producer: str       # source-of-origin agency (SRS/ORGI/Census), NOT "RBI"
    source_title: str          # names the Handbook as the access surface
    source_vintage: str        # Handbook edition (e.g. "2024-25")
    source_url: str            # source-of-origin landing page

    # --- workbook layout (operator stages the file; path injected) ---
    staging_filename: str      # filename the operator saves under the staging dir
    sheet: str | None = None   # sheet name / case-insensitive substring; None = active
    state_label_match: str = "state"   # header-row anchor (normalised substring)
    time_kind: str = TIME_CALENDAR_YEAR
    value_scale: float = 1.0   # multiplier (e.g. lakh -> absolute); default identity
    # Banded (two-row) header support. When set, the parser reads the row
    # immediately below the period row as a sub-header and, for each period
    # band (e.g. "2018-22" spanning Male / Female / Total sub-columns), keeps
    # ONLY the sub-column whose label matches this value (normalised
    # substring; e.g. "total"). This is how Life Expectancy (M/F/T per window)
    # ingests as ONE Total series instead of fragmenting into three files.
    # None (default) = single-value layout: each period column IS the value.
    value_sub_label: str | None = None
    skip_labels: tuple[str, ...] = ()        # aggregate / note rows to drop
    all_india_labels: tuple[str, ...] = ()   # extra labels mapping to "IN"

    def __post_init__(self) -> None:
        if self.time_kind not in _TIME_KINDS:
            raise ValueError(
                f"{self.indicator_id}: unknown time_kind {self.time_kind!r} "
                f"(expected one of {sorted(_TIME_KINDS)})"
            )
        if "__" in self.indicator_id:
            raise ValueError(
                f"indicator_id must not contain '__': {self.indicator_id!r}"
            )


@dataclass(frozen=True)
class LongRow:
    """One melted observation: ``(entity_id, time, value)``."""

    entity_id: str
    time: int
    value: float


def parse_hbs_workbook(
    workbook_bytes: bytes,
    spec: HbsTableSpec,
    resolver: StateResolver,
) -> list[LongRow]:
    """Parse one staged Handbook workbook into long-format rows.

    Auto-detects the header row (first row carrying the state-label
    anchor AND at least one parseable period column) and melts every
    ``state x period`` cell to a :class:`LongRow`. Blank / N.A. cells are
    dropped (sparse-safe). All-India rows resolve to ``"IN"``.

    Raises:
        RbiHbsShapeError: the sheet is missing, the header cannot be
            located, a state label is unmatched (and not skip-listed),
            a value cell is unparseable, a ``(entity, time)`` key
            duplicates, or no rows are produced.
    """
    workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    try:
        sheet = _select_sheet(workbook, spec.sheet)
        grid = [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    data_start_idx, state_col, period_cols = _detect_header(grid, spec)

    skip = {normalise_label(label) for label in spec.skip_labels}
    extra_all_india = {normalise_label(label) for label in spec.all_india_labels}

    rows: list[LongRow] = []
    seen: set[tuple[str, int]] = set()
    unmatched: list[str] = []

    for raw_cells in grid[data_start_idx:]:
        label = raw_cells[state_col] if state_col < len(raw_cells) else None
        if label is None or not str(label).strip():
            continue
        key = normalise_label(label)
        # Resolve FIRST: an all-India aggregate ('All India' / 'India')
        # resolves to the country entity and is KEPT; only labels that do
        # not resolve are tested against the skip list (footnote / sub-sum
        # rows) before a fail-loud unmatched error.
        entity_id = resolver.resolve(label)
        if entity_id is None and key in extra_all_india:
            entity_id = COUNTRY_ENTITY_ID
        if entity_id is None:
            if _is_skip(key, skip):
                continue
            unmatched.append(str(label).strip())
            continue
        for col, time in period_cols.items():
            cell = raw_cells[col] if col < len(raw_cells) else None
            value = _coerce_value(cell)
            if value is None:
                continue
            pk = (entity_id, time)
            if pk in seen:
                raise RbiHbsShapeError(
                    f"{spec.indicator_id}: duplicate (entity_id, time) {pk} "
                    f"- the workbook lists {label!r} more than once or a "
                    f"period column repeats"
                )
            seen.add(pk)
            rows.append(LongRow(entity_id, time, value * spec.value_scale))

    if unmatched:
        raise RbiHbsShapeError(
            f"{spec.indicator_id}: {len(unmatched)} unmatched state label(s) "
            f"- add them to the resolver dialect map or spec.skip_labels: "
            f"{sorted(set(unmatched))}"
        )
    if not rows:
        raise RbiHbsShapeError(
            f"{spec.indicator_id}: header located but no data rows parsed"
        )

    rows.sort(key=lambda r: (r.entity_id, r.time))
    return rows


def _select_sheet(workbook, sheet: str | None):
    if sheet is None:
        return workbook.active
    if sheet in workbook.sheetnames:
        return workbook[sheet]
    needle = sheet.lower()
    for name in workbook.sheetnames:
        if needle in name.lower():
            return workbook[name]
    raise RbiHbsShapeError(
        f"sheet {sheet!r} not found; available sheets: {workbook.sheetnames}"
    )


def _detect_header(
    grid: list[list[object]], spec: HbsTableSpec
) -> tuple[int, int, dict[int, int]]:
    """Locate the data-row start, the state column, and the period columns.

    Returns ``(data_start_idx, state_col, period_cols)`` where
    ``period_cols`` maps a value-column index to its integer ``time`` and
    ``data_start_idx`` is the first grid row carrying observations.

    Single-value layout (``value_sub_label is None``): the header row is the
    first row carrying the state anchor plus >= 1 parseable period column;
    data starts on the next row.

    Banded layout (``value_sub_label`` set): the matched row is the PERIOD
    row (its period labels may span several sub-columns, blank-filled by
    merged cells); the next row is the SUB-HEADER (e.g. Male / Female /
    Total). Only sub-columns whose label matches ``value_sub_label`` are
    kept, each governed by the nearest period label at or left of it; data
    starts two rows below the period row.
    """
    state_needle = normalise_label(spec.state_label_match)
    for idx, cells in enumerate(grid):
        state_col = _find_state_col(cells, state_needle)
        if state_col is None:
            continue
        if spec.value_sub_label is None:
            period_cols: dict[int, int] = {}
            for col in range(state_col + 1, len(cells)):
                time = _parse_period(cells[col], spec.time_kind)
                if time is not None:
                    period_cols[col] = time
            if period_cols:
                return idx + 1, state_col, period_cols
            continue
        # Banded: this row is the period row; the next row is the sub-header.
        if idx + 1 >= len(grid):
            continue
        filled = _forward_fill_periods(cells, state_col, spec.time_kind)
        if not filled:
            continue
        sub_needle = normalise_label(spec.value_sub_label)
        sub_cells = grid[idx + 1]
        banded_cols: dict[int, int] = {}
        for col in range(state_col + 1, len(sub_cells)):
            if col in filled and sub_needle and sub_needle in normalise_label(sub_cells[col]):
                banded_cols[col] = filled[col]
        if banded_cols:
            return idx + 2, state_col, banded_cols
    anchor = spec.value_sub_label or spec.time_kind
    raise RbiHbsShapeError(
        f"{spec.indicator_id}: could not locate a header row carrying the "
        f"state label {spec.state_label_match!r} plus at least one period "
        f"column (sub-label anchor: {anchor!r})"
    )


def _forward_fill_periods(
    period_row: list[object], state_col: int, time_kind: str
) -> dict[int, int]:
    """Map every value-side column to its governing period time.

    Walks columns left -> right from ``state_col + 1``, carrying the most
    recent parseable period forward across blank (merged-cell) gaps. This
    resolves which window/year governs each Male/Female/Total sub-column in
    a banded header.
    """
    out: dict[int, int] = {}
    current: int | None = None
    for col in range(state_col + 1, len(period_row)):
        time = _parse_period(period_row[col], time_kind)
        if time is not None:
            current = time
        if current is not None:
            out[col] = current
    return out


def _find_state_col(cells: list[object], state_needle: str) -> int | None:
    for col, cell in enumerate(cells):
        if cell is None:
            continue
        if state_needle and state_needle in normalise_label(cell):
            return col
    return None


def _is_skip(key: str, skip_prefixes: set[str]) -> bool:
    """True when a normalised label is an aggregate / footnote row to drop.

    Prefix match (not equality) so trailing text like 'Source: SRS 2024'
    or 'Notes on the table' is dropped while a real state name is never
    accidentally skipped (no Indian state starts with these tokens).
    """
    return any(key == p or key.startswith(p + " ") for p in skip_prefixes)


def _parse_period(cell: object, time_kind: str) -> int | None:
    if cell is None:
        return None
    if isinstance(cell, (int, float)) and time_kind == TIME_CALENDAR_YEAR:
        year = int(cell)
        return year if 1900 <= year <= 2100 else None
    text = str(cell).strip()
    if time_kind == TIME_CALENDAR_YEAR:
        match = _CAL_YEAR_RE.match(text)
        return int(match.group(1)) if match else None
    if time_kind == TIME_INTERVAL_WINDOW_END:
        match = _WINDOW_RE.search(text)
        if not match:
            return None
        start, end = match.group(1), match.group(2)
        # "2016-20" -> end year 2020 (borrow the century from the start).
        if len(end) == 2:
            end = start[:2] + end
        return int(end)
    raise RbiHbsShapeError(f"unknown time_kind: {time_kind!r}")


def _coerce_value(cell: object) -> float | None:
    if cell is None:
        return None
    if isinstance(cell, bool):
        # A boolean in a value cell is a publisher anomaly, not a number.
        raise RbiHbsShapeError(f"unexpected boolean value cell: {cell!r}")
    if isinstance(cell, (int, float)):
        return float(cell)
    text = str(cell).strip()
    if text.lower() in _NA_MARKERS:
        return None
    cleaned = text.replace(",", "")
    match = _LEADING_NUM_RE.match(cleaned)
    if match:
        return float(match.group(0))
    raise RbiHbsShapeError(f"unparseable value cell: {cell!r}")
