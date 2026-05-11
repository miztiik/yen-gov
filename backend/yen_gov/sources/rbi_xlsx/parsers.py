"""RBI State Finances XLSX parser — per-Statement, wide per-state layout.

Real-world recon (Jan 23, 2026 edition) showed RBI publishes ONE XLSX
PER Statement / Appendix Table on the
``State Finances: A Study of Budgets`` page, not one combined workbook.
Each per-state Statement has a fixed layout:

  - one sheet (e.g. ``ST_20``, ``ST_2``, ``APPT_3``)
  - a header row whose first labelled cell is ``State`` / ``State/UT``
    and whose subsequent cells are year periods
    (bare ``YYYY`` for stock indicators like outstanding liabilities,
    ``YYYY-YY`` fiscal-year spans for flow indicators)
  - an optional column-index row (``1 2 3 …``) immediately below
  - one data row per state, with column 1 = ``"<ordinal>. <Name>"``
    and subsequent columns = values
  - trailing aggregate row (``All States and UTs #``) + notes + source

This module is layout-driven via :class:`IndicatorSpec`. Adding a new
fiscal indicator means writing one spec + pinning one URL in
``urls.py`` — no parser-logic edits.

Pure: the orchestrator in :mod:`ingest` does network + writes.

See ``docs/architecture/backend/sources-rbi.md`` for the data contract.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook import Workbook


# ---------------------------------------------------------------------------
# Errors
# ---------------------------------------------------------------------------


class RBIWorkbookShapeError(ValueError):
    """The workbook's shape no longer matches the spec.

    Carries enough context for a maintainer to update the spec or
    re-run recon. We raise loud rather than emit zero rows: a silent
    coverage drop would lie to the citizen.
    """


# ---------------------------------------------------------------------------
# State-name normalisation
# ---------------------------------------------------------------------------


# RBI prefixes state rows with ordinals: ``"1. Andhra Pradesh"``,
# ``"29. Jammu and Kashmir"``. Normalise to canonical ECI state code.
_STATE_NAME_TO_ECI: dict[str, str] = {
    "andhra pradesh": "S01",
    "arunachal pradesh": "S02",
    "assam": "S03",
    "bihar": "S04",
    "chhattisgarh": "S26",
    "goa": "S05",
    "gujarat": "S06",
    "haryana": "S07",
    "himachal pradesh": "S08",
    "jammu and kashmir": "S09",
    "j&k": "S09",
    "jharkhand": "S27",
    "karnataka": "S10",
    "kerala": "S11",
    "madhya pradesh": "S12",
    "maharashtra": "S13",
    "manipur": "S14",
    "meghalaya": "S15",
    "mizoram": "S16",
    "nagaland": "S17",
    "odisha": "S18",
    "orissa": "S18",
    "punjab": "S19",
    "rajasthan": "S20",
    "sikkim": "S21",
    "tamil nadu": "S22",
    "telangana": "S29",
    "tripura": "S23",
    "uttar pradesh": "S24",
    "uttarakhand": "S28",
    "uttaranchal": "S28",
    "west bengal": "S25",
    # UTs with legislatures (RBI tables include them).
    "delhi": "U05",
    "nct delhi": "U05",
    "nct of delhi": "U05",
    "puducherry": "U07",
}


# Recognise and strip RBI's ordinal prefix: "1. Andhra Pradesh" → "Andhra Pradesh".
_ORDINAL_RE = re.compile(r"^\s*\d{1,2}\s*[.)]\s*")


def normalise_state_label(raw: str | None) -> str | None:
    """ECI state code for a raw RBI state label, or ``None``."""
    if not raw:
        return None
    s = str(raw).strip()
    s = _ORDINAL_RE.sub("", s)
    key = " ".join(s.lower().split())
    return _STATE_NAME_TO_ECI.get(key)


# ---------------------------------------------------------------------------
# Period normalisation
# ---------------------------------------------------------------------------


# Bare 4-digit year, optionally with a qualifier in parens or after a space.
# Matches: "2025", "2025 (RE)", "2025 (BE)", "2025-26", "2025-26 (BE)".
_YEAR_RE = re.compile(
    r"(?P<y1>\d{4})(?:\s*[-–]\s*(?P<y2>\d{2,4}))?"
    r"(?:\s*\(?\s*(?P<qual>RE|BE|Accounts|A|B|R)\s*\)?)?",
    re.IGNORECASE,
)


def _normalise_period(raw: str, *, period_kind: str) -> tuple[str, str | None]:
    """Convert a header cell to ``(time, qualifier)``.

    ``period_kind``:
      - ``"fy_end_year"``: bare year is END of Indian fiscal year
        (e.g. ``2026`` = end-March 2026 = FY 2025-26). Encoded as
        ``YYYY-03``.
      - ``"fy_span"``: ``YYYY-YY`` span (e.g. ``2022-23``). Encoded as
        ``YYYY-04`` (April-start of fiscal year).

    ``qualifier`` is ``None`` for accounts data, ``"RE"`` for revised,
    ``"BE"`` for budget estimates.
    """
    m = _YEAR_RE.search(raw)
    if not m:
        raise RBIWorkbookShapeError(f"unparseable period header: {raw!r}")
    y1 = m.group("y1")
    y2 = m.group("y2")
    qual = m.group("qual")
    if qual:
        q = qual.upper()
        if q in {"A", "ACCOUNTS"}:
            qual = None
        elif q == "R":
            qual = "RE"
        elif q == "B":
            qual = "BE"
        else:
            qual = q

    if period_kind == "fy_span":
        return f"{y1}-04", qual
    if period_kind == "fy_end_year":
        if y2 and not qual:
            return f"{y1}-04", None
        return f"{y1}-03", qual
    raise RBIWorkbookShapeError(f"unknown period_kind: {period_kind!r}")


# ---------------------------------------------------------------------------
# Spec + result dataclasses
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class IndicatorSpec:
    """Locate one fiscal indicator within ONE RBI Statement workbook.

    Args:
        indicator_id: stable id, e.g.
            ``"in.fiscal.outstanding_debt_pct_gsdp"``.
        sheet_match: case-insensitive substring of the sheet name
            (e.g. ``"ST_20"``). Empty ``""`` means "first sheet".
        header_label_match: case-insensitive substring identifying the
            header row's first labelled cell (e.g. ``"state"``).
        period_kind: ``"fy_end_year"`` (stocks) or ``"fy_span"`` (flows).
        sign: ``+1`` normally; ``-1`` for sign-convention flips.
        value_column_label: when the workbook stacks multiple value
            sub-columns under each period header (e.g. Statement 17's
            ``Gross | Net*`` pair under each fiscal year), the
            case-insensitive substring of the desired sub-header.
            ``None`` (default) = no sub-header; the period header column
            IS the value column (the Statement 20 / outstanding-debt
            shape). When set, the parser reads the row immediately
            below the period header as a sub-header row and shifts each
            period's column to the matching sub-cell.
    """

    indicator_id: str
    sheet_match: str
    header_label_match: str
    period_kind: str
    sign: int = 1
    value_column_label: str | None = None


@dataclass(frozen=True)
class ParsedRow:
    """One indicator-schema row: ``(entity, time, value, facet)``."""

    entity_id: str
    time: str
    value: float | None
    facet: str | None = None


@dataclass
class ParsedIndicator:
    indicator_id: str
    rows: list[ParsedRow] = field(default_factory=list)
    unmatched_states: list[str] = field(default_factory=list)
    sheet_name: str = ""
    period_columns: int = 0


# ---------------------------------------------------------------------------
# Cell coercion
# ---------------------------------------------------------------------------


_NULL_TOKENS = {"", "—", "-", "–", "N.A.", "NA", "n.a.", "na", "..", "..."}


def _coerce_value(raw: Any, sign: int) -> float | None:
    """RBI uses em-dash / hyphen / "N.A." as the missing sentinel."""
    if raw is None:
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        return float(raw) * sign
    s = str(raw).strip().replace("\xa0", "").replace(",", "")
    if s in _NULL_TOKENS:
        return None
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    try:
        v = float(s)
    except ValueError:
        return None
    if neg:
        v = -v
    return v * sign


# ---------------------------------------------------------------------------
# Sheet introspection
# ---------------------------------------------------------------------------


def _find_sheet(wb: Workbook, sheet_match: str) -> str:
    if not sheet_match:
        return wb.sheetnames[0]
    needle = sheet_match.lower().strip()
    for name in wb.sheetnames:
        if needle in name.lower():
            return name
    raise RBIWorkbookShapeError(
        f"no sheet matched {sheet_match!r}; saw {wb.sheetnames!r}"
    )


def _find_header_row(
    rows: list[list[Any]],
    header_label_match: str,
    period_kind: str,
) -> tuple[int, int]:
    """Return ``(row_idx, label_col_idx)`` of the header row.

    A row qualifies only if BOTH conditions hold:
      - it contains the ``header_label_match`` substring in some cell
        (case-insensitive); the first such cell is the label column.
      - the row also contains at least 2 period columns to the right
        of the label column.

    Without the second check we would falsely match the workbook's
    free-text title row (``Statement 20: Outstanding Liabilities …``)
    which contains the word ``State`` but no year columns.
    """
    needle = header_label_match.lower().strip()
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            if cell is None:
                continue
            s = str(cell).strip().lower()
            if needle not in s:
                continue
            # Candidate. Now check that there are >=2 period columns to the right.
            periods = _extract_period_columns(row, j, period_kind)
            if len(periods) >= 2:
                return i, j
            break  # this row's first labelled cell didn't win — try next row
    raise RBIWorkbookShapeError(
        f"no header row contained {header_label_match!r} with >=2 period columns"
    )


def _extract_period_columns(
    header_row: list[Any],
    label_col: int,
    period_kind: str,
) -> list[tuple[int, str, str | None]]:
    """``[(col_idx, time, qualifier), …]`` for every header cell after
    ``label_col`` that parses as a period."""
    out: list[tuple[int, str, str | None]] = []
    for j in range(label_col + 1, len(header_row)):
        cell = header_row[j]
        if cell is None:
            continue
        text = str(cell).strip()
        if not text or not _YEAR_RE.search(text):
            continue
        try:
            time, qual = _normalise_period(text, period_kind=period_kind)
        except RBIWorkbookShapeError:
            continue
        out.append((j, time, qual))
    return out


def _is_aggregate_label(label: str) -> bool:
    """RBI footers like ``"All States and UTs #"`` and trailing notes
    must be skipped so they never enter per-state output."""
    s = label.lower()
    return (
        "all state" in s
        or "all-india" in s
        or "all india" in s
        or s.startswith("notes")
        or s.startswith("note ")
        or s.startswith("source")
        or s.startswith("re:")
        or s.startswith("be:")
        or s.startswith("#:")
    )


def _apply_subcolumn_offset(
    periods: list[tuple[int, str, str | None]],
    sub_header_row: list[Any],
    value_column_label: str,
) -> list[tuple[int, str, str | None]]:
    """Shift each period's column to its matching sub-header sub-cell.

    For Statement 17 each fiscal-year header (e.g. ``"2023-24 (Accounts)"``
    at col 2) spans two value cells: ``Gross`` at col 2, ``Net*`` at col 3.
    Given ``value_column_label="Net*"``, return periods with col bumped
    to the matching sub-cell column. The search window for each period
    is ``[period_col, next_period_col)`` so we never cross a year boundary.
    """
    needle = value_column_label.lower().strip()
    out: list[tuple[int, str, str | None]] = []
    for i, (col, time, qual) in enumerate(periods):
        end = periods[i + 1][0] if i + 1 < len(periods) else len(sub_header_row)
        match_col: int | None = None
        for j in range(col, end):
            if j >= len(sub_header_row):
                break
            cell = sub_header_row[j]
            if cell is None:
                continue
            if needle in str(cell).strip().lower():
                match_col = j
                break
        if match_col is None:
            raise RBIWorkbookShapeError(
                f"sub-header {value_column_label!r} not found under period "
                f"{time!r} (search cols {col}..{end - 1})"
            )
        out.append((match_col, time, qual))
    return out


def _is_index_row_cell(cell: Any) -> bool:
    """Detect the ``"1 2 3 …"`` column-index row RBI inserts below
    headers. The first labelled cell is a small integer string."""
    if cell is None:
        return False
    s = str(cell).strip()
    if not s:
        return False
    return s.isdigit() and len(s) <= 3


# ---------------------------------------------------------------------------
# Per-indicator parser
# ---------------------------------------------------------------------------


def _parse_one(
    wb: Workbook,
    spec: IndicatorSpec,
) -> ParsedIndicator:
    sheet_name = _find_sheet(wb, spec.sheet_match)
    sheet = wb[sheet_name]
    rows: list[list[Any]] = [list(r) for r in sheet.iter_rows(values_only=True)]

    header_idx, label_col = _find_header_row(
        rows, spec.header_label_match, spec.period_kind
    )
    periods = _extract_period_columns(
        rows[header_idx], label_col, spec.period_kind
    )
    if len(periods) < 2:
        raise RBIWorkbookShapeError(
            f"only {len(periods)} period columns found in header of "
            f"{sheet_name!r} for {spec.indicator_id}; expected >=2"
        )

    # Optional: shift each period's column to its sub-header sub-cell
    # (Statement 17's stacked Gross|Net under each fiscal year). When
    # set, the sub-header row is the FIRST row at or below header+1
    # that contains the value_column_label substring in any cell;
    # intermediate decoration rows (e.g. Statement 17's
    # ``Col.4/Col.2`` variation-formula row) are skipped. The parser
    # then resumes data-row scanning past the sub-header.
    sub_header_idx: int | None = None
    if spec.value_column_label is not None:
        needle = spec.value_column_label.lower().strip()
        for j in range(header_idx + 1, len(rows)):
            for cell in rows[j]:
                if cell is None:
                    continue
                if needle in str(cell).strip().lower():
                    sub_header_idx = j
                    break
            if sub_header_idx is not None:
                break
        if sub_header_idx is None:
            raise RBIWorkbookShapeError(
                f"value_column_label {spec.value_column_label!r} not found "
                f"in any row below header in {sheet_name!r}"
            )
        periods = _apply_subcolumn_offset(
            periods, rows[sub_header_idx], spec.value_column_label
        )

    out = ParsedIndicator(
        indicator_id=spec.indicator_id,
        sheet_name=sheet_name,
        period_columns=len(periods),
    )

    # Data starts after: the header row, any rows up to AND including
    # the sub-header row (when present), and zero-or-more "1 2 3 …"
    # column-index rows. Skip them all.
    start = (sub_header_idx if sub_header_idx is not None else header_idx) + 1
    while start < len(rows):
        row = rows[start]
        if not row or label_col >= len(row):
            start += 1
            continue
        if _is_index_row_cell(row[label_col]):
            start += 1
            continue
        break

    seen_unknown: set[str] = set()
    for i in range(start, len(rows)):
        row = rows[i]
        if not row or label_col >= len(row):
            continue
        raw_label = row[label_col]
        if raw_label is None:
            continue
        label = str(raw_label).strip()
        if not label:
            continue
        if _is_aggregate_label(label):
            continue

        eci = normalise_state_label(label)
        if not eci:
            # State names are short. Long labels are almost always trailer
            # prose — RBI footnotes start with ordinals ("2. As reported …")
            # so the aggregate-label heuristic misses them. Cap by length
            # to keep ``unmatched_states`` clean (it surfaces in the
            # artifact's ``notes`` field).
            if label not in seen_unknown and len(label) <= 50:
                out.unmatched_states.append(label)
                seen_unknown.add(label)
            continue

        for col, time, qual in periods:
            cell = row[col] if col < len(row) else None
            v = _coerce_value(cell, spec.sign)
            out.rows.append(
                ParsedRow(entity_id=eci, time=time, value=v, facet=qual)
            )

    if not out.rows:
        raise RBIWorkbookShapeError(
            f"no per-state rows materialised for {spec.indicator_id} in "
            f"sheet {sheet_name!r}; layout may have shifted (re-run recon)."
        )
    return out


def parse_workbook(
    xlsx_bytes: bytes,
    spec: IndicatorSpec,
) -> ParsedIndicator:
    """Parse one indicator from one RBI Statement workbook."""
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    try:
        return _parse_one(wb, spec)
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Spec registry — Phase C ships ONE indicator end-to-end.
#
# Statement 20 ("Total Outstanding Liabilities of State Governments —
# As per cent of GSDP") is the cleanest per-state %GSDP series RBI
# publishes: ~31 states/UTs × 19 fiscal years (end-March 2008 → 2026 BE).
# It anchors the fiscal-capacity narrative on StateOverview.
#
# Subsequent indicators (revenue deficit, fiscal deficit, interest
# payments, central transfers, …) require either downloading additional
# Statement workbooks (Statement 2, 3, 13, 17, 33, …) or computing
# ratios across two Statements. Those land in Phase D — see
# docs/architecture/backend/sources-rbi.md.
# ---------------------------------------------------------------------------


SHIPPED_SPECS: tuple[IndicatorSpec, ...] = (
    IndicatorSpec(
        indicator_id="fiscal/outstanding_debt_pct_gsdp",
        sheet_match="ST_20",
        header_label_match="state",
        period_kind="fy_end_year",
    ),
    # Statement 17: net devolution + grants from Centre to each state.
    # Stacked Gross | Net* sub-columns under each fiscal year; we take
    # Net (RBI's net-of-adjustments figure). Coverage is currently 3
    # fiscal years from one edition; historical depth requires scraping
    # prior editions (deferred per ADR-0022 ingest gate).
    IndicatorSpec(
        indicator_id="fiscal/net_transfers_from_centre",
        sheet_match="ST_17",
        header_label_match="state/ut",
        period_kind="fy_span",
        value_column_label="Net",
    ),
)
