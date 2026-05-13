"""Pure parsers for the CEA Installed Capacity workbook.

The "IC" sheet of the CEA monthly Executive Summary lays out installed
generation capacity by **state × ownership × fuel mode**. The structure is:

    Region header (e.g. "INSTALLED CAPACITY ... NORTHERN REGION")
    Header rows (3 lines: Region|State, Ownership, Fuel-mode tree)
    State block 1:
        Col B = state name, Col C = "State"      | values
        Col B = None,       Col C = "Private"    | values
        Col B = None,       Col C = "Central"    | values
        Col B = None,       Col C = "Sub-Total"  | aggregated values
    State block 2: ...
    "Central - Unallocated" row (ignored — no entity)
    "Total (X Region)" block (ignored — region totals)

Fuel columns (1-based per Excel):
    D Coal | E Lignite | F Gas | G Diesel | H Total Thermal
    I Nuclear | J Hydro | K RES (MNRE) | L Total Renewable
    M Grand Total

This parser walks the sheet, finds each state's "Sub-Total" row, and emits
one row per ``(entity_id, fuel)`` for the snapshot date carried in the file.

No I/O, no network. ``parse_workbook(content, snapshot)`` takes the raw
XLSX bytes and returns a dict ``{indicator_id: list[ParsedRow]}``.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# State-name normalisation
# ---------------------------------------------------------------------------

# CEA state labels → ECI state code. CEA bundles "Jammu & Kashmir and Ladakh"
# into one row; we attribute the combined value to U08 (J&K UT) and document
# this in the per-indicator notes (the alternative — splitting unknown
# fractions across U08/U09 — would be fabrication).
_STATE_NAME_TO_ECI: dict[str, str] = {
    # Northern Region
    "delhi": "U05",
    "haryana": "S07",
    "himachal pradesh": "S08",
    "jammu & kashmir and ladakh": "U08",
    "jammu and kashmir and ladakh": "U08",
    "punjab": "S19",
    "rajasthan": "S20",
    "uttar pradesh": "S24",
    "uttarakhand": "S28",
    "chandigarh": "U02",
    # Western Region
    "goa": "S05",
    "gujarat": "S06",
    "madhya pradesh": "S12",
    "chhattisgarh": "S26",
    "maharashtra": "S13",
    "dadra and nagar haveli and daman and diu": "U03",
    # Southern Region
    "andhra pradesh": "S01",
    "telangana": "S29",
    "karnataka": "S10",
    "kerala": "S11",
    "tamil nadu": "S22",
    "puducherry": "U07",
    # Eastern Region
    "bihar": "S04",
    "jharkhand": "S27",
    "west bengal": "S25",
    "odisha": "S18",
    "sikkim": "S21",
    # North-Eastern Region
    "assam": "S03",
    "arunachal pradesh": "S02",
    "meghalaya": "S15",
    "tripura": "S23",
    "manipur": "S14",
    "nagaland": "S17",
    "mizoram": "S16",
    # Islands
    "andaman & nicobar": "U01",
    "andaman and nicobar": "U01",
    "lakshadweep": "U04",
}

# Labels that appear in column B but aren't state-attributable. Includes
# the central-pool corporations CEA lists alongside states (NLC, DVC) and
# the residual "Central - Unallocated" share that the workbook can't pin
# to a single state.
_NON_STATE_LABELS = {
    "nlc",
    "dvc",
    "central - unallocated",
}


def _normalise_state_label(raw: object) -> str | None:
    """Return ECI state code for a CEA state-name cell, or ``None``."""
    if raw is None:
        return None
    # Workbook frequently has trailing whitespace and newlines from merged
    # cell formatting; collapse them before lookup.
    key = " ".join(str(raw).strip().lower().split())
    if not key or key in _NON_STATE_LABELS:
        return None
    if key.startswith("total ") or key.startswith("all india"):
        return None
    return _STATE_NAME_TO_ECI.get(key)


# ---------------------------------------------------------------------------
# Fuel-column spec
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class FuelColumn:
    """One fuel column in the IC sheet, with the indicator id it feeds."""

    indicator_id: str  # e.g. "energy/installed_capacity_coal_mw"
    column_index: int  # 0-based offset into the row tuple
    title: str         # human-readable fuel name (used in error messages)


# Column indexes correspond to ``ws.iter_rows(values_only=True)`` tuples
# (0-based). The IC sheet uses Col A as a margin (None), so:
#   B=1 State, C=2 Ownership, D=3 Coal, E=4 Lignite, F=5 Gas, G=6 Diesel,
#   H=7 Thermal Total, I=8 Nuclear, J=9 Hydro, K=10 RES (MNRE),
#   L=11 Renewable Total, M=12 Grand Total.
SHIPPED_COLUMNS: tuple[FuelColumn, ...] = (
    FuelColumn("energy/installed_capacity_total_mw",      12, "Grand Total"),
    FuelColumn("energy/installed_capacity_thermal_mw",     7, "Total Thermal"),
    FuelColumn("energy/installed_capacity_coal_mw",        3, "Coal"),
    FuelColumn("energy/installed_capacity_gas_mw",         5, "Gas"),
    FuelColumn("energy/installed_capacity_nuclear_mw",     8, "Nuclear"),
    FuelColumn("energy/installed_capacity_hydro_mw",       9, "Hydro"),
    FuelColumn("energy/installed_capacity_renewable_mw",  10, "RES (MNRE)"),
)


# ---------------------------------------------------------------------------
# Output shape
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class ParsedRow:
    entity_id: str  # ECI state/UT code (e.g. "S22", "U05")
    time: str       # "YYYY-MM" — snapshot month
    value: float    # MW


@dataclass(frozen=True)
class ParsedWorkbook:
    """All indicators extracted from one workbook."""

    snapshot_period: str  # "YYYY-MM"
    rows_by_indicator: dict[str, list[ParsedRow]]
    state_count: int      # how many distinct state labels resolved


# ---------------------------------------------------------------------------
# Snapshot date detection
# ---------------------------------------------------------------------------

# CEA stamps "(As on 31.03.2026)" in the "IC" sheet header. We pull the
# month from there (period = "YYYY-MM") rather than trusting a filename.
_AS_ON_RE = re.compile(r"(?i)as\s*on\s*\d{1,2}[./-](\d{1,2})[./-](\d{4})")


def _detect_snapshot_period(values: list[object]) -> str | None:
    for cell in values:
        if cell is None:
            continue
        m = _AS_ON_RE.search(str(cell))
        if m:
            return f"{int(m.group(2)):04d}-{int(m.group(1)):02d}"
    return None


# ---------------------------------------------------------------------------
# Value coercion
# ---------------------------------------------------------------------------

_NULL_TOKENS = {"", "—", "-", "–", "N.A.", "NA", "n.a.", "na", ".."}


def _coerce(raw: object) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip()
    if s in _NULL_TOKENS:
        return None
    s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Main parse
# ---------------------------------------------------------------------------


def parse_workbook(content: bytes) -> ParsedWorkbook:
    """Parse a CEA Installed Capacity workbook into per-indicator rows.

    Walks the "IC" sheet, identifies each state block by its "Sub-Total"
    ownership row, and emits one row per (state, fuel-column) pair.
    """
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    if "IC" not in wb.sheetnames:
        raise ValueError(f"workbook missing 'IC' sheet — found {wb.sheetnames!r}")
    ws = wb["IC"]

    # First pass: find snapshot period from the header band (first ~10 rows).
    header_band: list[object] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 12:
            break
        header_band.extend(row)
    snapshot = _detect_snapshot_period(header_band)
    if snapshot is None:
        raise ValueError("could not detect snapshot date in IC sheet header")

    rows_by_indicator: dict[str, list[ParsedRow]] = {
        col.indicator_id: [] for col in SHIPPED_COLUMNS
    }
    seen_states: set[str] = set()

    # Second pass: walk all rows, tracking the current state context.
    # A state is "active" from when its name appears in col B (and resolves
    # to an ECI code) until the "Sub-Total" row for that state is hit.
    current_state_eci: str | None = None
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 13:
            continue
        label = row[1]
        ownership = row[2]

        # New state row: col B has a state name AND col C has "State"
        # (the first ownership tier in each state block).
        if label is not None and ownership is not None:
            ownership_str = str(ownership).strip().lower()
            if ownership_str == "state":
                current_state_eci = _normalise_state_label(label)
                if current_state_eci is not None:
                    seen_states.add(current_state_eci)
                continue

        # Sub-Total row: emit values for the active state, then close.
        if (
            current_state_eci is not None
            and ownership is not None
            and str(ownership).strip().lower() in {"sub-total", "sub total"}
        ):
            for col in SHIPPED_COLUMNS:
                v = _coerce(row[col.column_index])
                if v is None:
                    continue
                rows_by_indicator[col.indicator_id].append(
                    ParsedRow(entity_id=current_state_eci, time=snapshot, value=v)
                )
            current_state_eci = None
            continue

        # Reset on non-state labels (region totals, footnotes) so no stale
        # context bleeds into the next block.
        if label is not None and ownership is None:
            current_state_eci = None

    return ParsedWorkbook(
        snapshot_period=snapshot,
        rows_by_indicator=rows_by_indicator,
        state_count=len(seen_states),
    )
