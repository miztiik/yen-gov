"""Adapter for RBI HBS-IE state SDP tables.

No network. Reads the cached HBS-IE 2024-25 state SDP workbooks from
``.runtime/raw/rbi/handbook_economy_2024_25/`` and writes the canonical
long-format CSV variables under ``datasets/data/datapoints/geo/``.

Replaces the retired ``tools/rbi_hbs_ingest_state_gdp.py`` script. Under
B4-pt3 the legacy folded-indicator JSON write path was retired (per umbrella
plan O1 - no strangler-fig); CSV is the only canonical artifact this adapter
emits today.
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl

from yen_gov.canonical.citation import derive_source_id
from yen_gov.canonical.csv_writer import write_csv
from yen_gov.sources.rbi_hbs import (
    ALL_INDIA_NAMES,
    HBS_IE_LANDING,
    NAME_TO_ECI,
    coerce_value,
    fy_label_to_time,
)

CACHE_RELDIR = Path(".runtime/raw/rbi/handbook_economy_2024_25")

SNAPSHOT_URLS: dict[str, str] = {
    "T05": "https://rbidocs.rbi.org.in/rdocs/Publications/DOCs/05T_2908202556D0D1A9FA0C4615A7889EC1F025BACE.XLSX",
    "T06": "https://rbidocs.rbi.org.in/rdocs/Publications/DOCs/06T_290820258621D22235014AC19EE27C859382FEAF.XLSX",
    "T09": "https://rbidocs.rbi.org.in/rdocs/Publications/DOCs/09T_2908202582956A1D380840F5870B18841EEEF815.XLSX",
    "T10": "https://rbidocs.rbi.org.in/rdocs/Publications/DOCs/10T_290820257B1763CF72624007ABF4F4E48CC941B3.XLSX",
}

BASE_PRIORITY = ("2011-12", "2004-05", "1999-2000", "1993-94")


class RBIHBSIEStateSDPCacheMissing(RuntimeError):
    """No cached HBS-IE state SDP workbook to read."""


@dataclass(frozen=True)
class TableSpec:
    table_key: str
    filename: str
    table_label: str


@dataclass(frozen=True)
class IndicatorSpec:
    indicator_id: str
    table: TableSpec


@dataclass(frozen=True)
class WorkbookData:
    content: bytes
    fetched_at: datetime
    url: str


@dataclass(frozen=True)
class IndicatorIngestResult:
    """Per-CSV emit receipt. ``variable_id`` is the kebab-case CSV filename stem."""
    variable_id: str
    csv_path: Path
    row_count: int
    entity_count: int
    period_start: int
    period_end: int


@dataclass(frozen=True)
class IngestResult:
    indicators: tuple[IndicatorIngestResult, ...]


TABLE_T05 = TableSpec(
    table_key="T05",
    filename="T05_NSDP_Statewise_Current.xlsx",
    table_label="Table 5: Net State Domestic Product - State-wise (At Current Prices)",
)
TABLE_T06 = TableSpec(
    table_key="T06",
    filename="T06_NSDP_Statewise_Constant.xlsx",
    table_label="Table 6: Net State Domestic Product - State-wise (At Constant Prices)",
)
TABLE_T09 = TableSpec(
    table_key="T09",
    filename="T09_PCNSDP_Statewise_Current.xlsx",
    table_label="Table 9: Per Capita Net State Domestic Product - State-wise (At Current Prices)",
)
TABLE_T10 = TableSpec(
    table_key="T10",
    filename="T10_PCNSDP_Statewise_Constant.xlsx",
    table_label="Table 10: Per Capita Net State Domestic Product - State-wise (At Constant Prices)",
)

PER_CAPITA_SPECS: tuple[IndicatorSpec, ...] = (
    IndicatorSpec(
        indicator_id="economy/per_capita_nsdp_current_inr",
        table=TABLE_T09,
    ),
    IndicatorSpec(
        indicator_id="economy/per_capita_nsdp_constant_inr",
        table=TABLE_T10,
    ),
)


def _is_year(value: object) -> str | None:
    if not isinstance(value, str):
        return None
    label = value.strip()
    if len(label) == 7 and label[4] == "-" and label[:4].isdigit() and label[5:].isdigit():
        return label
    return None


# ---------------------------------------------------------------------------
# Canonical long-format CSV emission (B1.5.5)
# ---------------------------------------------------------------------------

# All rbi_hbs_ie_state_sdp specs come from one publication
# (Handbook of Statistics on Indian Economy). Per ADR-0042 the
# vintage is the publisher edition string. fk-validator stays dark on
# the source_id hashes until entities/source.csv lands (B2a), by
# design. Per-table title (Table 5 / 6 / 9 / 10) yields a distinct
# triple per variable_id, matching the sibling rbi_hbs_ie_centre_deficits
# precedent.
_CSV_SOURCE_PRODUCER = "Reserve Bank of India"
_CSV_SOURCE_VINTAGE = "2024-25"
_CSV_FILE_CLASS = "datasets/data/datapoints/geo/*.csv"
_CSV_OUT_REL_DIR = "datasets/data/datapoints/geo"

# Per-(legacy indicator_id, facet) -> kebab-case variable_id per
# ADR-0044 (no grain prefix) and parent plan section 21.6 / 21.12
# (no `__`). The nsdp-inr-crore family splits into per-price-basis
# variables because the writer does not yet support facet columns
# (csv_writer.py top-of-file note); per sub-plan point 7 we split each
# facet-bearing series into per-facet variable_ids.
_VARIABLE_ID_NSDP_CURRENT = "nsdp-inr-crore-current-prices"
_VARIABLE_ID_NSDP_CONSTANT = "nsdp-inr-crore-constant-prices"
_VARIABLE_ID_PER_CAPITA_CURRENT = "per-capita-nsdp-inr-current-prices"
_VARIABLE_ID_PER_CAPITA_CONSTANT = "per-capita-nsdp-inr-constant-prices"

# Source title per table; vintage shared across the family.
_CSV_SOURCE_TITLE_BY_TABLE: dict[str, str] = {
    "T05": "Handbook of Statistics on Indian Economy, Table 5",
    "T06": "Handbook of Statistics on Indian Economy, Table 6",
    "T09": "Handbook of Statistics on Indian Economy, Table 9",
    "T10": "Handbook of Statistics on Indian Economy, Table 10",
}


def _fy_start_year(time_str: str) -> int:
    """Lift the FY start year (int) from a parser period stamp.

    ``collapse_to_long`` emits ``YYYY-04`` for FY starts via
    ``fy_label_to_time``. The canonical CSV file class declares
    ``time`` as integer; we lift YYYY. Raises ``ValueError`` on
    malformed input - fail fast at the boundary (no silent coercion).
    """
    head, _, _tail = time_str.partition("-")
    return int(head)


def build_csv_rows(
    rows: list[dict[str, Any]],
    *,
    source_id: str,
) -> list[dict[str, Any]]:
    """Project parser rows onto the canonical four-column CSV shape.

    Each row carries the four canonical columns declared on file class
    ``datasets/data/datapoints/geo/*.csv``: ``entity_id``, ``time``,
    ``value``, ``source_id``. ``entity_id`` is the ECI state code (or
    ``IN`` for All-India) lifted from the parser; parent plan section
    22.4 #6 preserves LGD/ECI key separation. Rows sorted by
    (entity_id, time) per file-class contract.
    """
    out = [
        {
            "entity_id": str(r["entity_id"]),
            "time": _fy_start_year(str(r["time"])),
            "value": r["value"],
            "source_id": source_id,
        }
        for r in rows
    ]
    out.sort(key=lambda row: (row["entity_id"], row["time"]))
    return out


def emit_csv_variable(
    *, repo_root: Path, variable_id: str, rows: list[dict[str, Any]]
) -> Path:
    """Write one variable_id to `datasets/data/datapoints/geo/<id>.csv`."""
    out_dir = repo_root / _CSV_OUT_REL_DIR
    return write_csv(
        path=out_dir / f"{variable_id}.csv",
        file_class=_CSV_FILE_CLASS,
        rows=rows,
    )


def _csv_source_id_for_table(table_key: str) -> str:
    return derive_source_id(
        _CSV_SOURCE_PRODUCER,
        _CSV_SOURCE_TITLE_BY_TABLE[table_key],
        _CSV_SOURCE_VINTAGE,
    )


def _is_base_marker(value: object) -> str | None:
    if not isinstance(value, str):
        return None
    if "Base Year" not in value and "Base :" not in value:
        return None
    inside = value.replace("(", "").replace(")", "").strip()
    parts = inside.split(":")
    if len(parts) < 2:
        return None
    return parts[-1].strip()


def parse_workbook(content: bytes) -> dict[str, dict[str, dict[str, float]]]:
    """Return ``{entity_id: {time: {base: value}}}`` for every state column."""
    out: dict[str, dict[str, dict[str, float]]] = {}
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            header_idx = None
            for idx, row in enumerate(rows[:8]):
                cell = row[1] if len(row) > 1 else None
                if isinstance(cell, str) and cell.strip() in (
                    "Year",
                    "State / Union Territory",
                ):
                    header_idx = idx
                    break
            if header_idx is None:
                continue

            header = rows[header_idx]
            col_to_entity: dict[int, str] = {}
            for col_idx, cell in enumerate(header):
                if not isinstance(cell, str):
                    continue
                name = cell.strip()
                if name in NAME_TO_ECI:
                    col_to_entity[col_idx] = NAME_TO_ECI[name]
                elif name in ALL_INDIA_NAMES:
                    col_to_entity[col_idx] = "IN"

            current_base: str | None = None
            for row in rows[header_idx + 1 :]:
                label = row[1] if len(row) > 1 else None
                base = _is_base_marker(label)
                if base:
                    current_base = base
                    continue
                year = _is_year(label)
                if not year or current_base is None:
                    continue
                time = fy_label_to_time(year)
                if time is None:
                    continue
                for col_idx, entity_id in col_to_entity.items():
                    value = coerce_value(row[col_idx]) if col_idx < len(row) else None
                    if value is None:
                        continue
                    out.setdefault(entity_id, {}).setdefault(time, {})[
                        current_base
                    ] = value
    finally:
        wb.close()
    return out


def collapse_to_long(parsed: dict[str, dict[str, dict[str, float]]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for entity_id in sorted(parsed):
        for time in sorted(parsed[entity_id]):
            by_base = parsed[entity_id][time]
            chosen_base = next((base for base in BASE_PRIORITY if base in by_base), None)
            if chosen_base is None:
                continue
            rows.append(
                {
                    "entity_id": entity_id,
                    "time": time,
                    "value": by_base[chosen_base],
                    "vintage": f"Base {chosen_base}",
                }
            )
    return rows


def _resolve_workbook(*, repo_root: Path, table: TableSpec) -> WorkbookData:
    relpath = CACHE_RELDIR / table.filename
    path = repo_root / relpath
    if not path.exists():
        raise RBIHBSIEStateSDPCacheMissing(
            f"No cached RBI HBS-IE state SDP workbook at {relpath.as_posix()}.\n"
            f"  (a) Open {HBS_IE_LANDING}\n"
            f"  (b) Pick the latest Handbook of Statistics on Indian Economy edition.\n"
            f"  (c) Download {table.table_label} as XLSX.\n"
            f"  (d) Save it as {relpath.as_posix()} relative to the repo root.\n"
            f"For the 2024-25 edition, the pinned direct URL is {SNAPSHOT_URLS[table.table_key]}."
        )
    fetched_at = datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).replace(
        microsecond=0
    )
    return WorkbookData(
        content=path.read_bytes(),
        fetched_at=fetched_at,
        url=SNAPSHOT_URLS[table.table_key],
    )


def _coverage_csv(rows: list[dict[str, Any]]) -> tuple[int, int, int]:
    """Return (period_start, period_end, entity_count) over CSV-shape rows.

    ``rows`` here are the canonical 4-column CSV rows after
    ``build_csv_rows`` projection (``time`` is int FY-start year).
    """
    times = sorted({int(row["time"]) for row in rows})
    entities = sorted({str(row["entity_id"]) for row in rows})
    if not times:
        raise ValueError("cannot build coverage for empty row set")
    return times[0], times[-1], len(entities)


def ingest(*, repo_root: Path) -> IngestResult:
    """Read cached workbooks and write canonical CSV variables.

    Returns one IndicatorIngestResult per emitted CSV variable (4 today:
    nsdp current + nsdp constant + per-capita current + per-capita
    constant).
    """
    current_workbook = _resolve_workbook(repo_root=repo_root, table=TABLE_T05)
    constant_workbook = _resolve_workbook(repo_root=repo_root, table=TABLE_T06)
    current_rows = collapse_to_long(parse_workbook(current_workbook.content))
    constant_rows = collapse_to_long(parse_workbook(constant_workbook.content))
    if not current_rows or not constant_rows:
        raise ValueError("empty rows for economy/nsdp_inr_crore")

    results: list[IndicatorIngestResult] = []

    # NSDP splits into two per-price-basis variables because the writer
    # does not yet support facet columns (sub-plan B1.5.1..5 point 7).
    for variable_id, parser_rows, table_key in (
        (_VARIABLE_ID_NSDP_CURRENT, current_rows, "T05"),
        (_VARIABLE_ID_NSDP_CONSTANT, constant_rows, "T06"),
    ):
        csv_rows = build_csv_rows(
            parser_rows, source_id=_csv_source_id_for_table(table_key)
        )
        csv_path = emit_csv_variable(
            repo_root=repo_root, variable_id=variable_id, rows=csv_rows
        )
        start, end, entity_count = _coverage_csv(csv_rows)
        results.append(
            IndicatorIngestResult(
                variable_id=variable_id,
                csv_path=csv_path,
                row_count=len(csv_rows),
                entity_count=entity_count,
                period_start=start,
                period_end=end,
            )
        )

    for spec in PER_CAPITA_SPECS:
        workbook = _resolve_workbook(repo_root=repo_root, table=spec.table)
        rows = collapse_to_long(parse_workbook(workbook.content))
        if not rows:
            raise ValueError(f"empty rows for {spec.indicator_id}")
        variable_id = (
            _VARIABLE_ID_PER_CAPITA_CURRENT
            if spec.table.table_key == "T09"
            else _VARIABLE_ID_PER_CAPITA_CONSTANT
        )
        csv_rows = build_csv_rows(
            rows, source_id=_csv_source_id_for_table(spec.table.table_key)
        )
        csv_path = emit_csv_variable(
            repo_root=repo_root, variable_id=variable_id, rows=csv_rows
        )
        start, end, entity_count = _coverage_csv(csv_rows)
        results.append(
            IndicatorIngestResult(
                variable_id=variable_id,
                csv_path=csv_path,
                row_count=len(csv_rows),
                entity_count=entity_count,
                period_start=start,
                period_end=end,
            )
        )

    return IngestResult(indicators=tuple(results))