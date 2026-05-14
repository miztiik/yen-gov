"""Pure parser for the RBI Appendix Table 1 (Major Deficit Indicators).

Workbook layout (verified against ``AppT1_MajorDeficitIndicators_2026.xlsx``,
edition January 2026 = State Finances 2025-26):

  - Single sheet ``APPT_1``.
  - Row 1 = title ("Appendix Table 1: Major Deficit Indicators of …").
  - Row 2 = unit annotation ("(₹ Crore)").
  - Row 3 = header: col 1 = "Year"; cols 2..6 = deficit indicator labels
    ("Gross Fiscal Deficit", "Revenue Deficit", "Primary Deficit",
    "Primary Revenue Deficit", "Net RBI Credit to States").
  - Row 4 = column-index row ("1 2 3 …") — skipped.
  - Rows 5..N: pairs of rows per fiscal year. The first row of each pair
    has the year string in col 1 (e.g. "2007-08", "2024-25 (BE)$") and
    the absolute ₹ Crore values in cols 2..6. The second row has a blank
    col 1 and the % of GDP values in cols 2..6 (negatives wrapped in
    parens, e.g. "(-2.9)").

We ship the **₹ Crore** values (the absolute series). The % GDP series
is not ingested in this version — it would need its own indicator family
with ``value_kind="percent"`` and a separate normaliser.

Layout-driven via :class:`DeficitSpec`. Add a new indicator → add a one-
line spec, no parser-logic edits.

Pure: I/O lives in ``ingest.py``.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook


class RBIAppT1ShapeError(ValueError):
    """The workbook's shape no longer matches the spec.

    Raised loudly rather than emitting zero rows: a silent coverage
    drop would lie to the citizen.
    """


# Match a fiscal-year label. Accepts:
#   "2007-08"
#   "2024-25 (BE)$"        — Budget Estimate, may carry trailing footnote markers
#   "2023-24 (RE)"         — Revised Estimate
#   "2025-26 (Budget Estimates)"
_PERIOD_RE = re.compile(
    r"^\s*(\d{4})-\d{2}(?:\s*\(([^)]+)\))?[\s$*#@]*$"
)


def _parse_period(raw: Any) -> tuple[int, str | None] | None:
    """Return ``(start_year, qualifier)`` or ``None`` if not a period label."""
    if raw is None:
        return None
    text = str(raw).strip()
    m = _PERIOD_RE.match(text)
    if not m:
        return None
    qualifier = (m.group(2) or "").strip().lower() or None
    return int(m.group(1)), qualifier


def _fy_to_period(start_year: int) -> str:
    """Canonical period for a fiscal year, start-of-FY (e.g. 2007 -> 2007-04)."""
    return f"{start_year:04d}-04"


@dataclass(frozen=True)
class DeficitSpec:
    """Locate one deficit indicator in the AppT1 workbook by column.

    Args:
        indicator_id: stable id, e.g. ``fiscal/national_gross_fiscal_deficit``.
        column_label_match: case-insensitive substring of the row-3 header
            cell. Must uniquely identify one column among the indicator
            header cells.
        sign: ``+1`` normally; ``-1`` for sign-convention flips (RBI reports
            Revenue Deficit as positive when revenue receipts < revenue
            expenditure; we keep RBI's sign convention as published).
    """

    indicator_id: str
    column_label_match: str
    sign: int = 1


@dataclass(frozen=True)
class ParsedRow:
    entity_id: str
    time: str
    value: float | None


@dataclass
class ParsedIndicator:
    indicator_id: str
    rows: list[ParsedRow] = field(default_factory=list)
    period_count: int = 0


# Ship the four citizen-facing deficit indicators. Net RBI Credit to States
# is a niche monetary-policy series and is intentionally NOT shipped from
# this source family.
SHIPPED_SPECS: tuple[DeficitSpec, ...] = (
    DeficitSpec(
        indicator_id="fiscal/national_gross_fiscal_deficit",
        column_label_match="gross fiscal deficit",
    ),
    DeficitSpec(
        indicator_id="fiscal/national_revenue_deficit",
        column_label_match="revenue deficit",
    ),
    DeficitSpec(
        indicator_id="fiscal/national_primary_deficit",
        column_label_match="primary deficit",
    ),
    DeficitSpec(
        indicator_id="fiscal/national_primary_revenue_deficit",
        column_label_match="primary revenue deficit",
    ),
)


_NULL_TOKENS = {"", "—", "-", "–", "N.A.", "NA", "n.a.", "na", "..", "...", "*"}


def _coerce(raw: Any, sign: int = 1) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw) * sign
    text = str(raw).strip()
    if text in _NULL_TOKENS:
        return None
    text = text.replace(",", "")
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1].lstrip("-")
    try:
        return float(text) * sign
    except ValueError:
        return None


def _norm(text: Any) -> str:
    return " ".join(str(text or "").split()).lower()


def _find_header(ws) -> tuple[int, int, dict[int, str]]:
    """Locate the header row.

    Returns ``(row_index, year_column, {column_index: header_label})``
    where column indices are 1-based (openpyxl convention) and
    ``header_label`` is the lowercased indicator name. The "Year"
    column itself is excluded from the indicator dict.
    """
    for row_index, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20)), start=1
    ):
        cells = list(row)
        first_label = next(
            (c for c in cells if c.value is not None and str(c.value).strip()),
            None,
        )
        if first_label is None or _norm(first_label.value) != "year":
            continue
        year_column = first_label.column
        # Collect indicator headers in subsequent columns.
        headers: dict[int, str] = {}
        for c in cells:
            if c.column == year_column:
                continue
            label = _norm(c.value)
            if label and label != "year":
                headers[c.column] = label
        # Need at least 2 indicator columns to be a credible header.
        if len(headers) >= 2:
            return row_index, year_column, headers
    raise RBIAppT1ShapeError(
        f"sheet {ws.title!r}: no header row with 'Year' + >= 2 indicator "
        f"columns found in the first 20 rows"
    )


def _resolve_column(
    headers: dict[int, str], match: str
) -> int:
    """Find the single column whose header contains ``match`` (case-insensitive).

    Uses **exact equality first** then substring match — needed because
    ``"primary deficit"`` is a substring of ``"primary revenue deficit"``.
    """
    needle = match.lower().strip()
    exact = [col for col, label in headers.items() if label == needle]
    if len(exact) == 1:
        return exact[0]
    if len(exact) > 1:
        raise RBIAppT1ShapeError(
            f"column label {match!r} matched {len(exact)} columns exactly: "
            f"{exact} — refine the spec"
        )
    matches = [col for col, label in headers.items() if needle in label]
    if len(matches) == 1:
        return matches[0]
    if not matches:
        raise RBIAppT1ShapeError(
            f"no column matched label {match!r}; headers were {headers!r}"
        )
    raise RBIAppT1ShapeError(
        f"column label {match!r} ambiguous — matched columns {matches} "
        f"({[headers[c] for c in matches]!r})"
    )


def parse_workbook(
    content: bytes,
    specs: tuple[DeficitSpec, ...] = SHIPPED_SPECS,
) -> dict[str, ParsedIndicator]:
    """Parse the AppT1 workbook for every spec; return ``{indicator_id: ParsedIndicator}``."""
    wb = load_workbook(io.BytesIO(content), data_only=True)
    if not wb.sheetnames:
        raise RBIAppT1ShapeError("workbook has no sheets")

    # AppT1 ships as a single-sheet workbook. If RBI ever splits it across
    # multiple sheets we'll iterate; for now use the first sheet.
    ws = wb[wb.sheetnames[0]]

    header_row, year_column, headers = _find_header(ws)
    cols_by_indicator = {
        spec.indicator_id: _resolve_column(headers, spec.column_label_match)
        for spec in specs
    }

    # Build per-indicator parsed rows. Walk rows below the header; for every
    # row whose year-column cell parses as a fiscal-year, emit a ParsedRow
    # per indicator from cells in that row. The very next row carries % GDP
    # which we explicitly skip (the year-row has the absolute values).
    parsed: dict[str, ParsedIndicator] = {
        spec.indicator_id: ParsedIndicator(indicator_id=spec.indicator_id)
        for spec in specs
    }
    period_count = 0

    # Skip the column-index row (row of "1 2 3 ..." digits) which sits
    # immediately below the header. Detect it by inspecting the first
    # indicator column (which holds "1" / "2" / etc., or "1.0" if openpyxl
    # auto-typed it). The year column is empty on that row.
    start_row = header_row + 1
    first_indicator_col = min(headers)
    after = ws.cell(row=start_row, column=first_indicator_col).value
    if isinstance(after, (int, float)) and int(after) == 1:
        start_row += 1
    elif isinstance(after, str) and after.strip() in {"1", "1.0"}:
        start_row += 1

    for row_index in range(start_row, ws.max_row + 1):
        first_cell = ws.cell(row=row_index, column=year_column).value
        period = _parse_period(first_cell)
        if period is None:
            continue
        period_count += 1
        time_str = _fy_to_period(period[0])
        for spec in specs:
            col = cols_by_indicator[spec.indicator_id]
            raw = ws.cell(row=row_index, column=col).value
            value = _coerce(raw, sign=spec.sign)
            parsed[spec.indicator_id].rows.append(
                ParsedRow(entity_id="IN", time=time_str, value=value)
            )

    if period_count == 0:
        raise RBIAppT1ShapeError(
            f"sheet {ws.title!r}: no fiscal-year rows found below header "
            f"(row {header_row}); workbook layout may have changed"
        )

    for ind in parsed.values():
        ind.period_count = period_count

    return parsed
