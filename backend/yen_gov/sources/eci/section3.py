"""Parser for ECI Statistical Report Section 3 — *List of Political Parties Participated*.

Each Statistical Report ships a separate XLSX named like
``3_-_List_Of_Political_Parties_Participated_*.xlsx`` containing a single
worksheet with the shape::

    R01: ['PARTY TYPE', 'ABBREVIATION', 'PARTY']
    R02: ['NATIONAL PARTIES']                                 <- group header
    R03: ['1.0', 'AAAP', 'Aam Aadmi Party']
    ...
    R08: ['STATE PARTIES']                                    <- group header
    ...
    R19: ['REGISTERED(Unrecognised) PARTIES']                 <- group header

Important shape facts confirmed against Assam-2026:
* The first column is a *display row counter*, NOT the canonical ECI
  numeric party code. The numeric eci_code (e.g. 742 for BJP) is published
  ONLY on the live-results portal partywise page and is NOT recoverable
  from any Statistical Report XLSX.
* Group headers occupy a single non-empty column; data rows have three.
* Independents are not listed (they are not parties).

This module returns a flat list of ``ParticipatingParty`` rows annotated
with their group label, leaving eci_code resolution to a future reference
dataset (see TODO/ECI-MULTI-STATE-INGEST-PLAN.md, follow-up to N2).
"""

from __future__ import annotations

import io
from dataclasses import dataclass

import openpyxl


@dataclass(frozen=True)
class ParticipatingParty:
    party_type: str  # "NATIONAL PARTIES", "STATE PARTIES", "STATE PARTIES - OTHER STATES", "REGISTERED(Unrecognised) PARTIES"
    short_name: str
    full_name: str


def parse_section3_parties(xlsx_bytes: bytes) -> list[ParticipatingParty]:
    """Parse the bytes of a Section 3 XLSX into a list of participating parties."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    if not wb.sheetnames:
        raise ValueError("Section 3 workbook has no sheets")
    ws = wb[wb.sheetnames[0]]

    parties: list[ParticipatingParty] = []
    current_type = ""
    seen_header = False
    for row in ws.iter_rows(values_only=True):
        cells = [
            (str(v).strip() if v is not None else "")
            for v in row
        ]
        # Drop trailing empties.
        while cells and not cells[-1]:
            cells.pop()
        if not cells:
            continue

        non_empty = [c for c in cells if c]

        # First non-empty row = column header. Skip it.
        if not seen_header:
            if non_empty[:1] == ["PARTY TYPE"] or "ABBREVIATION" in non_empty:
                seen_header = True
            continue

        # Group header rows have exactly one non-empty cell, which is the
        # type label (e.g. "NATIONAL PARTIES").
        if len(non_empty) == 1:
            current_type = non_empty[0]
            continue

        # Data rows: [counter, abbreviation, full_name]. The counter is
        # discarded (it's a display-row index, not the ECI code).
        if len(cells) >= 3 and cells[1] and cells[2]:
            parties.append(ParticipatingParty(
                party_type=current_type,
                short_name=cells[1],
                full_name=cells[2],
            ))

    return parties
