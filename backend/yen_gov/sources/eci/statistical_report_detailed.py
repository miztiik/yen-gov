"""Parse ECI Statistical Report Section 10 — "Detailed Results" XLSX.

Section 10 is the per-AC per-candidate vote breakdown — the richest single
sheet in the Statistical Report family. Two layouts ECI has shipped under
this title since 2023, both supported by header-name-based column
resolution (so a future column tweak fails loudly rather than silently
shifting a column index):

  2024+ (15 cols, e.g. May-2026 cohort):
    STATE/UT NAME | AC NO. | AC NAME | CANDIDATE NAME | GENDER | AGE |
    CATEGORY | PARTY | SYMBOL | GENERAL | POSTAL | TOTAL |
    OVER VALID VOTES + NOTA | OVER TOTAL ELECTORS | TOTAL ELECTORS

  2023 (14 cols, MP / Chhattisgarh / Mizoram / Telangana cohort):
    STATE/UT NAME | AC NO. | AC NAME | CANDIDATE NAME | SEX | AGE |
    CATEGORY | PARTY | SYMBOL | GENERAL | POSTAL | TOTAL |
    % VOTES POLLED | TOTAL ELECTORS

Semantically the 2023 single "% VOTES POLLED" column is the same as the
2024+ "OVER VALID VOTES + NOTA" (candidate vote share over valid+NOTA).
The 2024+ extra "OVER TOTAL ELECTORS" column is unused on candidate rows
but is what the 2024+ TURN OUT row reads as turnout %; in 2023 the only
% column doubles as turnout on the TURN OUT row.

Rows after the header alternate between:

  - Candidate rows (one per contesting candidate, plus a NOTA row), with
    party SHORT codes (TVK, ADMK, INC, IND, NOTA, ...). Vote share columns
    are pre-computed by ECI; we use them as-authoritative per the authority
    hierarchy (sources-eci.md).
  - A "TURN OUT" sentinel row at the end of each AC, carrying the totals
    used elsewhere in the report (general + postal polled, turnout %).
  - A final "GRAND TOTAL:" row + Disclaimer rows at the bottom of the
    sheet — both ignored by the parser.

Two-step convention (mirrors HTML parsers in this package):

  - parse_detailed_results(xlsx_bytes) -> DetailedResultsRaw
        Pure XLSX -> data. Knows nothing about the schema, the election
        coordinates, or processing knobs.
  - to_constituency_results(raw, *, election, state, top_n, collapse_others,
                             sources)
        Emits one ConstituencyResult per AC, applying top_n / others
        collapsing the same way the HTML mapper does.

Party codes: Section 10 carries SHORT codes only (no numeric ECI code).
The parser leaves `party_eci_code` empty; the mapper (`to_constituency_results`)
optionally backfills it from a `{short_name → eci_code}` lookup, typically
built from the live-results partywise.htm snapshot via
`pipeline.compose.eci_code_by_short_from_partywise`. Without the lookup,
`party_eci_code` is None — the schema permits it.
"""

from __future__ import annotations

import io
from dataclasses import dataclass

from openpyxl import load_workbook

from yen_gov.core.models import (
    CandidateResult,
    ConstituencyResult,
    NotaResult,
    OthersBucket,
    ResultTotals,
    SourceRef,
    WinnerInfo,
)


_NOTA_TOKENS = {"nota", "none of the above"}
_INDEPENDENT_TOKENS = {"ind", "independent"}


@dataclass(frozen=True)
class CandidateRow:
    rank: int  # filled by mapper, not the parser
    name: str
    party_short: str  # raw text from the PARTY column (already short)
    is_nota: bool
    is_independent: bool
    votes_general: int
    votes_postal: int
    votes_total: int
    vote_share_pct: float  # over valid + NOTA, per ECI's own column
    # Per-candidate demographics carried by every Layout-A Section-10 sheet.
    # None on NOTA rows (where ECI leaves the cells blank) and on Layout-B
    # pre-2018 sheets where the columns existed under different names but the
    # parser still extracts them where possible (gender/age/category present).
    gender: str | None = None       # "M" / "F" / "O" / None
    age: int | None = None
    category: str | None = None     # "GEN" / "SC" / "ST" / None


@dataclass(frozen=True)
class AcSection:
    state_name: str  # e.g. "Tamil Nadu"
    eci_no: int
    constituency_name: str  # uppercase per ECI; reservation suffix kept
    candidates: list[CandidateRow]  # includes NOTA, in sheet order
    polled_general: int  # from TURN OUT row (col 10)
    polled_postal: int  # from TURN OUT row (col 11)
    polled_total: int  # from TURN OUT row (col 12); == sum of candidate totals
    turnout_pct: float | None  # from TURN OUT row (col 14); None if "-"
    total_electors: int | None  # from any candidate row (col 15); constant per AC


@dataclass(frozen=True)
class DetailedResultsRaw:
    state_name: str
    sections: list[AcSection]


@dataclass(frozen=True)
class _ColumnMap:
    """Resolved column indices for one header row.

    Built by :func:`_resolve_columns` from the actual header text so the
    same parser handles both the 2023 14-col and 2024+ 15-col layouts. A
    missing required column raises at parse-time with the offending
    header text in the message — same fail-loud posture as the rest of
    the package.
    """

    state: int
    eci_no: int
    ac_name: int
    candidate: int
    party: int
    general: int
    postal: int
    total: int
    vote_share: int
    total_electors: int
    # 2024+ only: separate "OVER TOTAL ELECTORS" % column. None for 2023
    # (which carries turnout in the single % column on the TURN OUT row).
    over_electors_pct: int | None
    # Optional demographic columns (Layout A always has them; Layout B has
    # equivalents under slightly different headers, resolved by the same
    # routine via substring matching).
    sex: int | None = None      # "SEX" or "GENDER" or "Candidate Sex"
    age_col: int | None = None  # "AGE" or "Candidate Age"
    category_col: int | None = None  # "CATEGORY" or "Candidate Category"

    @property
    def turnout_on_turn_out_row(self) -> int:
        """Column to read turnout_pct from on a TURN OUT sentinel row."""
        return self.over_electors_pct if self.over_electors_pct is not None else self.vote_share


# ---------------------------------------------------------------------------
# parsing
# ---------------------------------------------------------------------------

def parse_detailed_results(content: bytes) -> DetailedResultsRaw:
    """Parse Section 10 XLSX bytes into AC sections.

    Dispatches on the actual sheet shape between three observed layouts:

      - **Layout A** (2019+ Statistical Reports, including the May-2026 cohort):
        wide header row starting with ``STATE/UT NAME | AC NO. | AC NAME | ...``,
        candidate rows interleaved with a ``TURN OUT`` / ``TURNOUT`` sentinel
        per AC. Both the 2023 14-col and 2024+ 15-col variants are handled by
        the same code path via header-name resolution.

      - **Layout B** (2016–2017 Statistical Reports): narrower header
        ``Constituency No. | Constituency Name | Candidate Name | Candidate Sex |
        Candidate Age | Candidate Category | Party Name | VALID VOTES POLLED in
        General | ... in Postal | Total Valid Votes | Total Electors | Total
        Votes``. No ``STATE/UT NAME`` column (state name is in the title row),
        no ``% VOTES POLLED`` column (computed), no per-AC ``TURN OUT``
        sentinel (AC boundary detected by ``Constituency No.`` change), and
        ``Total Votes`` is replicated on every row of an AC as the polled total.
\n        State name is taken from the caller (the on-disk filename / event
        registry); the file itself does not carry it.

      - **Layout C** (pre-2019 ``.xls`` / 2018 Karnataka): no header row.
        Each AC is announced by a row whose first cell is ``Constituency`` and
        carries ``TOTAL ELECTORS :`` further along. Candidate rows then run
        ``serial | NAME | SEX | AGE | CATEGORY | PARTY | SYMBOL | GENERAL |
        POSTAL | TOTAL | %SHARE``. Sentinel row starts ``TURNOUT`` /
        ``TURN OUT``.

    All three layouts produce the same ``DetailedResultsRaw`` shape; downstream
    mapping (``to_constituency_results``) is layout-agnostic.

    Raises ValueError on structural surprises (header missing, AC section
    without a TURN OUT row, no sections at all). Same fail-loud posture as
    the HTML parsers — silent partials would let a sheet-shape change go
    unnoticed.
    """
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    layout = _detect_layout(rows)
    if layout == "A":
        return _parse_layout_a(rows)
    if layout == "B":
        return _parse_layout_b(rows)
    if layout == "C":
        return _parse_layout_c(rows)
    # _detect_layout raises with diagnostic detail if no layout fits; reaching
    # this line is impossible, but keep the assertion for type-checkers.
    raise AssertionError(f"unreachable layout {layout!r}")


def _detect_layout(rows: list[tuple]) -> str:
    """Scan the first ~20 rows and decide which layout this sheet is in.

    Layout A: a row starts with ``STATE/UT NAME`` (or ``STATE NAME``).
    Layout B: a row starts with ``Constituency No.`` (pre-2018 reports).
    Layout C: a row starts with ``Constituency`` AND contains ``TOTAL ELECTORS``
        (Karnataka 2018 marker-row style).
    """
    for r in rows[:20]:
        if not r:
            continue
        first = r[0]
        if not isinstance(first, str):
            continue
        head = first.strip().upper()
        if head.startswith("STATE"):
            # Look for AC NO. somewhere nearby to be sure
            joined = " | ".join(str(c or "").strip().upper() for c in r)
            if "AC NO" in joined:
                return "A"
        if head.startswith("CONSTITUENCY NO"):
            return "B"
        if head == "CONSTITUENCY" and any(
            isinstance(c, str) and "TOTAL ELECTORS" in c.upper() for c in r
        ):
            return "C"
    raise ValueError(
        "Section 10 sheet did not match any known layout (A=2019+, "
        "B=2016/2017, C=pre-2019 marker-row). First 5 non-empty rows: "
        f"{[r for r in rows[:20] if r and any(c is not None for c in r)][:5]!r}"
    )


def _parse_layout_a(rows: list[tuple]) -> DetailedResultsRaw:
    """Parse the 2019+ Statistical Report layout (see parse_detailed_results)."""
    header_idx = _find_header_row(rows)
    cols = _resolve_columns(rows[header_idx])
    sections: list[AcSection] = []
    state_name: str | None = None
    current: list[CandidateRow] = []
    current_eci_no: int | None = None
    current_ac_name: str | None = None
    current_state: str | None = None
    current_electors: int | None = None

    for raw_row in rows[header_idx + 1:]:
        if raw_row is None or all(c is None for c in raw_row):
            continue
        first = raw_row[cols.state]

        # End-of-AC sentinel. ECI ships two cosmetic spellings across cohorts:
        #   - "TURN OUT" (with internal space) on 2023+ Statistical Reports
        #   - "TURNOUT"  (no space)            on 2019/2020 Statistical Reports
        # Both mean the same thing — collapse internal whitespace and compare.
        first_norm = (
            "".join(first.split()).upper()
            if isinstance(first, str) else ""
        )
        if first_norm == "TURNOUT":
            if current_eci_no is None or not current:
                raise ValueError(
                    f"TURN OUT row with no preceding candidates "
                    f"(eci_no={current_eci_no})"
                )
            polled_general = _to_int(raw_row[cols.general])
            polled_postal = _to_int(raw_row[cols.postal])
            polled_total = _to_int(raw_row[cols.total])
            turnout_pct = _to_float_or_none(raw_row[cols.turnout_on_turn_out_row])
            sections.append(
                AcSection(
                    state_name=current_state or "",
                    eci_no=current_eci_no,
                    constituency_name=current_ac_name or "",
                    candidates=current,
                    polled_general=polled_general,
                    polled_postal=polled_postal,
                    polled_total=polled_total,
                    turnout_pct=turnout_pct,
                    total_electors=current_electors,
                )
            )
            current = []
            current_eci_no = None
            current_ac_name = None
            current_state = None
            current_electors = None
            continue

        # State-level grand total or disclaimer at the bottom — stop.
        if isinstance(first, str) and (
            first.strip().upper().startswith("GRAND TOTAL")
            or first.strip().lower().startswith("disclaimer")
            or first.strip().lower().startswith("this report is based on")
        ):
            break

        # Candidate row
        ac_no = raw_row[cols.eci_no]
        if not isinstance(ac_no, int):
            continue  # skip stray header repeats / spacer rows
        if state_name is None:
            state_name = str(first).strip() if first else ""
        if current_eci_no is None:
            current_eci_no = ac_no
            current_ac_name = str(raw_row[cols.ac_name]).strip() if raw_row[cols.ac_name] else ""
            current_state = str(first).strip() if first else state_name
        elif current_eci_no != ac_no:
            raise ValueError(
                f"AC #{ac_no} starts before the previous TURN OUT row "
                f"(was building #{current_eci_no})"
            )

        cand = _row_to_candidate(raw_row, cols)
        current.append(cand)
        electors_cell = raw_row[cols.total_electors]
        if current_electors is None and isinstance(electors_cell, (int, float)):
            current_electors = int(electors_cell)

    if current:
        raise ValueError(
            f"trailing AC #{current_eci_no} has no TURN OUT row"
        )
    if not sections:
        raise ValueError("no AC sections parsed from Section 10 XLSX")

    return DetailedResultsRaw(
        state_name=state_name or sections[0].state_name,
        sections=sections,
    )


# ---------------------------------------------------------------------------
# Layout B — 2016/2017 Statistical Reports
# ---------------------------------------------------------------------------

# Column resolver for Layout B works on the actual pre-2018 headers. Lookups
# are exact match against trimmed/upper text so a future column rename fails
# loud just like Layout A.
_LAYOUT_B_HEADER_TOKENS = {
    "CONSTITUENCY NO.", "CONSTITUENCY NO",
}


def _parse_layout_b(rows: list[tuple]) -> DetailedResultsRaw:
    """Parse the pre-2018 Statistical Report layout (2016 / 2017 reports).

    No STATE column, no % VOTES POLLED column, no per-AC TURN OUT sentinel.
    The polled total is what ECI repeats as ``Total Votes`` on every candidate
    row of an AC; ``Total Electors`` is likewise replicated per-row. AC
    boundary is detected when the ``Constituency No.`` integer changes.

    Vote share is computed (``votes / polled_total * 100``) because the source
    sheet does not carry it. We round to 2 dp to match Layout A's precision.

    State name is NOT in the file — caller (``to_constituency_results``) will
    set it; ``DetailedResultsRaw.state_name`` is left empty here.
    """
    header_idx = _find_layout_b_header(rows)
    cols = _resolve_layout_b_columns(rows[header_idx])

    sections: list[AcSection] = []
    current: list[CandidateRow] = []
    current_eci_no: int | None = None
    current_ac_name: str | None = None
    current_electors: int | None = None
    current_polled_total: int | None = None

    def flush_section() -> None:
        nonlocal current, current_eci_no, current_ac_name, current_electors, current_polled_total
        if current_eci_no is None or not current:
            return
        polled_total = current_polled_total or sum(c.votes_total for c in current)
        polled_general = sum(c.votes_general for c in current)
        polled_postal = sum(c.votes_postal for c in current)
        turnout_pct = (
            round(polled_total / current_electors * 100, 2)
            if current_electors else None
        )
        # Backfill vote_share_pct on each candidate from the now-known total.
        denom = polled_total or 1
        rebuilt = [
            CandidateRow(
                rank=c.rank, name=c.name, party_short=c.party_short,
                is_nota=c.is_nota, is_independent=c.is_independent,
                votes_general=c.votes_general, votes_postal=c.votes_postal,
                votes_total=c.votes_total,
                vote_share_pct=round(c.votes_total / denom * 100, 2),
                gender=c.gender, age=c.age, category=c.category,
            )
            for c in current
        ]
        sections.append(AcSection(
            state_name="",  # caller fills via to_constituency_results
            eci_no=current_eci_no,
            constituency_name=current_ac_name or "",
            candidates=rebuilt,
            polled_general=polled_general,
            polled_postal=polled_postal,
            polled_total=polled_total,
            turnout_pct=turnout_pct,
            total_electors=current_electors,
        ))
        current = []
        current_eci_no = None
        current_ac_name = None
        current_electors = None
        current_polled_total = None

    for raw_row in rows[header_idx + 1:]:
        if raw_row is None or all(c is None for c in raw_row):
            continue
        ac_no = raw_row[cols.eci_no]
        if not isinstance(ac_no, int):
            # Pre-2018 reports sometimes carry stray "Grand Total" / disclaimer
            # rows at the bottom. Bail loud if we hit text where an integer
            # was expected mid-section; otherwise it's the end of data.
            first = raw_row[cols.eci_no]
            if isinstance(first, str) and (
                "GRAND TOTAL" in first.upper()
                or first.lower().startswith("disclaimer")
            ):
                break
            continue

        if current_eci_no is None or current_eci_no != ac_no:
            flush_section()
            current_eci_no = ac_no
            current_ac_name = (
                str(raw_row[cols.ac_name]).strip()
                if raw_row[cols.ac_name] else ""
            )

        cand = _row_to_candidate(raw_row, cols)
        current.append(cand)

        electors_cell = raw_row[cols.total_electors]
        if current_electors is None and isinstance(electors_cell, (int, float)):
            current_electors = int(electors_cell)

        # Layout B carries the AC's polled total replicated on every row
        # (column "Total Votes"); we trust it over our own summation, which
        # would still be correct but rounds the same way ECI does.
        if cols.over_electors_pct is not None:
            # In Layout B we reuse the over_electors_pct slot to mean
            # "polled_total column index" (see _resolve_layout_b_columns).
            pol = raw_row[cols.over_electors_pct]
            if current_polled_total is None and isinstance(pol, (int, float)):
                current_polled_total = int(pol)

    flush_section()
    if not sections:
        raise ValueError("no AC sections parsed from Section 10 Layout-B XLSX")

    return DetailedResultsRaw(state_name="", sections=sections)


def _find_layout_b_header(rows: list[tuple]) -> int:
    for i, r in enumerate(rows[:20]):
        if not r:
            continue
        first = r[0]
        if isinstance(first, str) and first.strip().upper().startswith("CONSTITUENCY NO"):
            return i
    raise ValueError("Layout-B header row (Constituency No. | ...) not found")


def _resolve_layout_b_columns(header_row: tuple) -> _ColumnMap:
    """Map Layout-B headers onto the same _ColumnMap structure used by Layout A.

    Layout B has no STATE column (we set state=eci_no, knowing the parser will
    not consume it — Layout B's parse loop never reads ``cols.state``).
    Layout B has no ``% VOTES POLLED`` column; we point ``vote_share`` at the
    ``Total Votes`` column instead (a harmless dummy — the loop ignores
    vote_share for Layout B and computes shares on flush).
    Layout B has no ``OVER TOTAL ELECTORS`` column; we *repurpose* the
    ``over_electors_pct`` slot to carry the index of ``Total Votes`` so the
    Layout-B parse loop can read the per-AC polled total from every row.
    """
    cells = [
        (str(c).strip().upper() if c is not None else "") for c in header_row
    ]

    def find(predicates: list[str], label: str) -> int:
        for i, c in enumerate(cells):
            for p in predicates:
                if p in c:
                    return i
        raise ValueError(
            f"Layout-B header missing column for {label!r}; "
            f"saw: {[c for c in cells if c]!r}"
        )

    eci_no = find(["CONSTITUENCY NO"], "Constituency No.")
    ac_name = find(["CONSTITUENCY NAME"], "Constituency Name")
    candidate = find(["CANDIDATE NAME"], "Candidate Name")
    party = find(["PARTY"], "Party Name")
    general = find(["VOTES POLLED IN GENERAL"], "VALID VOTES POLLED in General")
    postal = find(["VOTES POLLED IN POSTAL"], "VALID VOTES POLLED in Postal")
    total = find(["TOTAL VALID VOTES"], "Total Valid Votes")
    total_electors = find(["TOTAL ELECTORS"], "Total Electors")
    # Pre-2018 reports ship two minor header variants for the replicated
    # AC-level polled total: 2016 = "Total Votes"; 2017 = "Total valid votes
    # polled +NOTA". Either works; if neither is present, _parse_layout_b
    # falls back to summing candidate totals (mathematically identical).
    polled_total: int | None = None
    for i, c in enumerate(cells):
        if c == "TOTAL VOTES" or "TOTAL VALID VOTES POLLED" in c:
            polled_total = i
            break
    # Demographics — present in Layout B under "Candidate Sex / Age / Category".
    sex_idx: int | None = None
    age_idx: int | None = None
    cat_idx: int | None = None
    for i, c in enumerate(cells):
        if "CANDIDATE SEX" in c:
            sex_idx = i
        elif "CANDIDATE AGE" in c:
            age_idx = i
        elif "CANDIDATE CATEGORY" in c:
            cat_idx = i

    return _ColumnMap(
        state=eci_no,        # unused on Layout B; dummy alias
        eci_no=eci_no,
        ac_name=ac_name,
        candidate=candidate,
        party=party,
        general=general,
        postal=postal,
        total=total,
        vote_share=total,    # unused on Layout B; computed at flush time
        total_electors=total_electors,
        over_electors_pct=polled_total,  # repurposed: index of "Total Votes"
        sex=sex_idx,
        age_col=age_idx,
        category_col=cat_idx,
    )


# ---------------------------------------------------------------------------
# Layout C — pre-2019 ``.xls`` (2018 Karnataka, repackaged as ``.xlsx``)
# ---------------------------------------------------------------------------

def _parse_layout_c(rows: list[tuple]) -> DetailedResultsRaw:
    """Parse the marker-row layout (no header, ``Constituency``/``TURNOUT``).

    Sheet has no header row. ACs are announced by:
        ['Constituency', <ac_no>, '.', '<ac name>', 'TOTAL ELECTORS :', <n>, ...]
    Candidate rows are then:
        [<serial>, <NAME>, <SEX>, <AGE>, <CATEGORY>, <PARTY>, <SYMBOL>,
         <GENERAL>, <POSTAL>, <TOTAL>, <%SHARE>]
    AC ends with:
        ['TURNOUT', 'TOTAL:', <general>, <postal>, <total>, <turnout_pct>, ...]
    """
    sections: list[AcSection] = []
    current: list[CandidateRow] = []
    current_eci_no: int | None = None
    current_ac_name: str | None = None
    current_electors: int | None = None

    for raw_row in rows:
        if raw_row is None or all(c is None for c in raw_row):
            continue
        first = raw_row[0]
        first_str = str(first).strip().upper() if isinstance(first, str) else ""

        # Per-AC announcement row
        if first_str == "CONSTITUENCY":
            if current_eci_no is not None:
                raise ValueError(
                    f"new Constituency marker before previous AC #{current_eci_no} "
                    "was closed by a TURNOUT row"
                )
            ac_no_raw = raw_row[1]
            # Karnataka 2018 stores the AC number as a string ('1') after the
            # xlsx repackage; older variants used int. Coerce loudly.
            try:
                current_eci_no = int(ac_no_raw) if ac_no_raw is not None else None
            except (TypeError, ValueError):
                current_eci_no = None
            if current_eci_no is None:
                raise ValueError(f"Constituency marker row missing AC #: {raw_row!r}")
            current_ac_name = (
                str(raw_row[3]).strip() if len(raw_row) > 3 and raw_row[3] else ""
            )
            electors_cell = raw_row[5] if len(raw_row) > 5 else None
            current_electors = (
                int(electors_cell) if isinstance(electors_cell, (int, float)) else None
            )
            continue

        # End-of-AC sentinel
        first_norm = "".join(first_str.split())
        if first_norm == "TURNOUT":
            if current_eci_no is None or not current:
                raise ValueError(
                    f"TURNOUT row with no preceding candidates (eci_no={current_eci_no})"
                )
            polled_general = _to_int(raw_row[2])
            polled_postal = _to_int(raw_row[3])
            polled_total = _to_int(raw_row[4])
            turnout_pct = _to_float_or_none(raw_row[5])
            sections.append(AcSection(
                state_name="",  # not in file — caller fills
                eci_no=current_eci_no,
                constituency_name=current_ac_name or "",
                candidates=current,
                polled_general=polled_general,
                polled_postal=polled_postal,
                polled_total=polled_total,
                turnout_pct=turnout_pct,
                total_electors=current_electors,
            ))
            current = []
            current_eci_no = None
            current_ac_name = None
            current_electors = None
            continue

        # Candidate row — first cell is an integer serial
        if not isinstance(first, (int, float)):
            continue
        if current_eci_no is None:
            raise ValueError(
                f"candidate row outside any Constituency section: {raw_row!r}"
            )
        # Build CandidateRow manually (no header-row column map here).
        name = str(raw_row[1]).strip() if raw_row[1] else ""
        party = str(raw_row[5]).strip() if raw_row[5] else ""
        party_lower = party.lower()
        is_nota = party_lower in _NOTA_TOKENS or name.lower() in _NOTA_TOKENS
        is_ind = party_lower in _INDEPENDENT_TOKENS
        current.append(CandidateRow(
            rank=0,
            name=name,
            party_short=party if party else "IND",
            is_nota=is_nota,
            is_independent=is_ind,
            votes_general=_to_int(raw_row[7]),
            votes_postal=_to_int(raw_row[8]),
            votes_total=_to_int(raw_row[9]),
            vote_share_pct=_to_float(raw_row[10]) if len(raw_row) > 10 else 0.0,
            gender=_normalize_gender(raw_row[2]),
            age=_normalize_age(raw_row[3]),
            category=_normalize_category(raw_row[4]),
        ))

    if current_eci_no is not None:
        raise ValueError(f"trailing AC #{current_eci_no} has no TURNOUT row")
    if not sections:
        raise ValueError("no AC sections parsed from Section 10 Layout-C sheet")

    return DetailedResultsRaw(state_name="", sections=sections)


def _find_header_row(rows: list[tuple]) -> int:
    for i, r in enumerate(rows):
        if r and isinstance(r[0], str) and r[0].strip().upper().startswith("STATE"):
            if r[1] and "AC" in str(r[1]).upper():
                return i
    raise ValueError(
        "Section 10 header row (STATE/UT NAME | AC NO. | ...) not found"
    )


def _resolve_columns(header_row: tuple) -> _ColumnMap:
    """Map header text to canonical column indices.

    Required headers (case-insensitive substring match unless noted):
      - STATE / AC NO / AC NAME / CANDIDATE / PARTY (exact, to avoid
        Section 3's "PARTY TYPE") / GENERAL / POSTAL / TOTAL (exact, to
        avoid "TOTAL ELECTORS") / TOTAL ELECTORS
      - vote_share: "% VOTES POLLED" (2023) or "OVER VALID" (2024+)

    Optional 2024+-only header:
      - over_electors_pct: "OVER TOTAL ELECTORS" (drives turnout reading
        on the TURN OUT row when present; 2023 falls back to vote_share)
    """
    cells = [
        str(c).strip().upper() if c is not None else ""
        for c in header_row
    ]

    def find(predicate, label: str) -> int:
        for i, c in enumerate(cells):
            if predicate(c):
                return i
        raise ValueError(
            f"Section 10 header missing column for {label!r}; "
            f"saw: {[c for c in cells if c]!r}"
        )

    state = find(lambda c: c.startswith("STATE"), "STATE/UT NAME")
    eci_no = find(lambda c: c.startswith("AC NO"), "AC NO.")
    ac_name = find(lambda c: c == "AC NAME", "AC NAME")
    candidate = find(lambda c: c == "CANDIDATE NAME", "CANDIDATE NAME")
    party = find(lambda c: c == "PARTY", "PARTY")
    general = find(lambda c: c == "GENERAL", "GENERAL")
    postal = find(lambda c: c == "POSTAL", "POSTAL")
    total = find(lambda c: c == "TOTAL", "TOTAL (candidate vote sum)")
    total_electors = find(lambda c: c == "TOTAL ELECTORS", "TOTAL ELECTORS")

    # vote_share column: 2023 uses "% VOTES POLLED", 2024+ uses
    # "OVER VALID VOTES + NOTA". Both must NOT match "OVER TOTAL ELECTORS".
    vote_share = find(
        lambda c: ("% VOTES POLLED" in c) or ("OVER VALID" in c),
        "vote-share %",
    )

    # 2024+ only: separate "OVER TOTAL ELECTORS" column. None for 2023.
    over_electors_pct: int | None = None
    for i, c in enumerate(cells):
        if "OVER TOTAL ELECTORS" in c:
            over_electors_pct = i
            break

    # Optional demographic columns. Required headers above already raise loudly
    # if missing; these three are *additive* (schema 3.2) and tolerated as None
    # so the same _resolve_columns path is reusable from Layout B's resolver.
    sex: int | None = None
    age_col: int | None = None
    category_col: int | None = None
    for i, c in enumerate(cells):
        if c in ("SEX", "GENDER", "CANDIDATE SEX"):
            sex = i
        elif c in ("AGE", "CANDIDATE AGE"):
            age_col = i
        elif c in ("CATEGORY", "CANDIDATE CATEGORY"):
            category_col = i

    return _ColumnMap(
        state=state,
        eci_no=eci_no,
        ac_name=ac_name,
        candidate=candidate,
        party=party,
        general=general,
        postal=postal,
        total=total,
        vote_share=vote_share,
        total_electors=total_electors,
        over_electors_pct=over_electors_pct,
        sex=sex,
        age_col=age_col,
        category_col=category_col,
    )


def _row_to_candidate(raw_row: tuple, cols: _ColumnMap) -> CandidateRow:
    name = str(raw_row[cols.candidate]).strip() if raw_row[cols.candidate] else ""
    party = str(raw_row[cols.party]).strip() if raw_row[cols.party] else ""
    party_lower = party.lower()
    is_nota = party_lower in _NOTA_TOKENS
    is_ind = party_lower in _INDEPENDENT_TOKENS
    return CandidateRow(
        rank=0,
        name=_strip_leading_serial(name),
        party_short=party if party else "IND",
        is_nota=is_nota,
        is_independent=is_ind,
        votes_general=_to_int(raw_row[cols.general]),
        votes_postal=_to_int(raw_row[cols.postal]),
        votes_total=_to_int(raw_row[cols.total]),
        vote_share_pct=_to_float(raw_row[cols.vote_share]),
        gender=_normalize_gender(raw_row[cols.sex]) if cols.sex is not None else None,
        age=_normalize_age(raw_row[cols.age_col]) if cols.age_col is not None else None,
        category=_normalize_category(raw_row[cols.category_col]) if cols.category_col is not None else None,
    )


_GENDER_MAP = {
    "M": "M", "MALE": "M",
    "F": "F", "FEMALE": "F",
    "O": "O", "OTHERS": "O", "OTHER": "O",
    "THIRD GENDER": "O", "TRANSGENDER": "O", "T": "O",
}
_CATEGORY_MAP = {
    "GEN": "GEN", "GENERAL": "GEN",
    "SC": "SC",
    "ST": "ST",
}


def _normalize_gender(value: object) -> str | None:
    if value is None or value == "":
        return None
    key = str(value).strip().upper()
    return _GENDER_MAP.get(key)


def _normalize_age(value: object) -> int | None:
    if value is None or value == "" or value == "-":
        return None
    try:
        n = int(float(value))
    except (TypeError, ValueError):
        return None
    # Constitutional minimum for Assembly candidacy is 25 (Art. 173(b)); the
    # schema's 18 floor is conservative so an off-by-one ECI typo doesn't bin
    # the artifact. Bin obviously bad values to None rather than fail loud.
    if not (18 <= n <= 120):
        return None
    return n


def _normalize_category(value: object) -> str | None:
    if value is None or value == "":
        return None
    key = str(value).strip().upper()
    return _CATEGORY_MAP.get(key)


def _strip_leading_serial(name: str) -> str:
    """ECI prefixes the row's serial (e.g. '1 S.vijayakumar'). Strip it."""
    parts = name.split(" ", 1)
    if len(parts) == 2 and parts[0].isdigit():
        return parts[1].strip()
    return name


def _to_int(value: object) -> int:
    if value is None or value == "" or value == "-":
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value).strip().replace(",", "")
    if not s or s == "-":
        return 0
    return int(float(s))


def _to_float(value: object) -> float:
    if value is None or value == "" or value == "-":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().rstrip("%").replace(",", "")
    if not s or s == "-":
        return 0.0
    return float(s)


def _to_float_or_none(value: object) -> float | None:
    if value is None or value == "" or value == "-":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().rstrip("%").replace(",", "")
    if not s or s == "-":
        return None
    return float(s)


# ---------------------------------------------------------------------------
# raw -> schema-bound model
# ---------------------------------------------------------------------------

def to_constituency_results(
    raw: DetailedResultsRaw,
    *,
    election: str,
    state: str,
    top_n: int,
    collapse_others: bool,
    sources: list[SourceRef],
    party_eci_codes: dict[str, str] | None = None,
) -> list[ConstituencyResult]:
    """Build one ConstituencyResult per AC section.

    Mirrors `constituencywise.to_constituency_result` semantics: rank by
    total votes desc, take top_n as candidates, NOTA broken out, optional
    collapsed `others` bucket, winner derived from rank-1 row with margin
    over runner-up.

    Optional `party_eci_codes` is a {short_name → numeric_code} lookup used
    to backfill `party_eci_code` (Section 10 carries short codes only).
    Independents and unmapped shorts get `None`.

    **Countermanded ACs are skipped silently.** ECI publishes Section 10
    rows for postponed/countermanded constituencies as a single NOTA row
    with all-zero votes and polled_total=0 (e.g. WB 2026 AC #144 FALTA).
    The schema requires `candidates: minItems: 1` and a non-zero winner —
    representing "no result" with a stub would be misleading. Skipping
    leaves a gap in the per-AC file sequence that consumers can detect.
    Caller can compare `len(out)` to `len(raw.sections)` to count skips.
    """
    codes = party_eci_codes or {}
    def _code_for(short: str, is_ind: bool) -> str | None:
        if is_ind:
            return None
        return codes.get(short)
    out: list[ConstituencyResult] = []
    for sec in raw.sections:
        if sec.polled_total == 0:
            continue  # countermanded / postponed — no result to emit
        nota_row = next((c for c in sec.candidates if c.is_nota), None)
        contestants = [c for c in sec.candidates if not c.is_nota]
        contestants.sort(key=lambda c: c.votes_total, reverse=True)
        if not contestants:
            raise ValueError(f"AC #{sec.eci_no} has no non-NOTA candidates")
        if nota_row is None:
            raise ValueError(f"AC #{sec.eci_no} has no NOTA row")

        ranked = [
            CandidateRow(
                rank=i + 1,
                name=c.name, party_short=c.party_short, is_nota=False,
                is_independent=c.is_independent,
                votes_general=c.votes_general, votes_postal=c.votes_postal,
                votes_total=c.votes_total, vote_share_pct=c.vote_share_pct,
                gender=c.gender, age=c.age, category=c.category,
            )
            for i, c in enumerate(contestants)
        ]
        kept, rest = ranked[:top_n], ranked[top_n:]

        candidates_model = [
            CandidateResult(
                rank=row.rank,
                name=row.name,
                party_eci_code=_code_for(row.party_short, row.is_independent),
                party_short="IND" if row.is_independent else row.party_short,
                votes=row.votes_total,
                vote_share_pct=row.vote_share_pct,
                is_winner=(row.rank == 1) or None,
                gender=row.gender,
                age=row.age,
                category=row.category,
            )
            for row in kept
        ]
        others_model: OthersBucket | None = None
        if collapse_others and rest:
            others_model = OthersBucket(
                candidate_count=len(rest),
                votes=sum(r.votes_total for r in rest),
                vote_share_pct=round(sum(r.vote_share_pct for r in rest), 2),
            )

        winner = ranked[0]
        runner_up_votes = ranked[1].votes_total if len(ranked) >= 2 else 0
        margin_votes = winner.votes_total - runner_up_votes
        margin_pct = (
            (margin_votes / sec.polled_total * 100.0)
            if sec.polled_total > 0 else 0.0
        )

        out.append(ConstituencyResult(
            sources=sources,
            election=election,
            state=state,
            body="AC",
            eci_no=sec.eci_no,
            constituency_name=sec.constituency_name,
            totals=ResultTotals(
                electors=sec.total_electors,
                votes_polled=sec.polled_total,
                turnout_pct=sec.turnout_pct,
            ),
            candidates=candidates_model,
            nota=NotaResult(
                votes=nota_row.votes_total,
                vote_share_pct=nota_row.vote_share_pct,
            ),
            others=others_model,
            top_n_cutoff=top_n,
            winner=WinnerInfo(
                name=winner.name,
                party_eci_code=_code_for(winner.party_short, winner.is_independent),
                party_short="IND" if winner.is_independent else winner.party_short,
                votes=winner.votes_total,
                margin_votes=margin_votes,
                margin_pct=round(margin_pct, 2),
            ),
        ))
    return out
