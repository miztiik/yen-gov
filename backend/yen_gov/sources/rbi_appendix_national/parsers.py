"""Pure parser for RBI Appendix Table 2 (national-time-series shape).

Workbook layout (verified against ``02_APP_devolution_transfers.xlsx``,
edition January 2026 = State Finances 2025-26):

  - 1+ sheets (e.g. ``APPT_1``, ``APPT_2``, ``APPT_3``), each holding
    one band of fiscal years (e.g. 2007-08…2013-14, 2014-15…2019-20,
    2020-21…2025-26 BE).
  - Header row: first labelled cell is ``"Item"``; subsequent cells
    are fiscal-year periods (``"2007-08"``, ``"2023-24 (Accounts)"``).
  - Column-index row (``1 2 3 …``) immediately below.
  - Item rows: column-1 label is e.g. ``"VI. Net Transfer of Resources
    from the Centre (IV-V)"``; subsequent cells are values in ₹ Crore.
  - Trailing notes / source rows at the bottom.

The parser is layout-driven via :class:`AppendixSpec`. To add another
indicator from the same publication you write one spec; no parser-logic
edits.

Pure: I/O lives in ``ingest.py``.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook import Workbook


class RBIAppendixShapeError(ValueError):
    """The workbook's shape no longer matches the spec.

    Raised loudly rather than emitting zero rows: a silent coverage
    drop would lie to the citizen.
    """


# Match a fiscal-year period header. Accepts:
#   "2007-08"
#   "2023-24 (Accounts)"
#   "2025-26 (Budget Estimates)"
# Captures the start year so we can normalise to a canonical period.
_PERIOD_RE = re.compile(
    r"^\s*(\d{4})-\d{2}(?:\s*\(([^)]+)\))?\s*$"
)


def _parse_period_header(raw: Any) -> tuple[int, str | None] | None:
    """Return ``(start_year, qualifier)`` or ``None`` if not a period header."""
    if raw is None:
        return None
    text = str(raw).strip()
    m = _PERIOD_RE.match(text)
    if not m:
        return None
    qualifier = (m.group(2) or "").strip().lower() or None
    return int(m.group(1)), qualifier


def _fy_to_period(start_year: int) -> str:
    """Canonical period for a fiscal year, start-of-FY (e.g. 2007 → 2007-04)."""
    return f"{start_year:04d}-04"


@dataclass(frozen=True)
class AppendixSpec:
    """Locate one national-level indicator in an Appendix workbook.

    Args:
        indicator_id: stable id, e.g. ``fiscal/national_centre_transfers_total``.
        item_label_match: case-insensitive substring of the row's
            ``Item`` cell. Must be specific enough to identify ONE row
            in any sheet (e.g. ``"net transfer of resources"`` matches
            row VI uniquely).
        sign: ``+1`` normally; ``-1`` for sign-convention flips.
        prefer_qualifier: when one fiscal year appears multiple times
            with different qualifiers (e.g. 2023-24 has both
            ``Accounts`` and ``Budget Estimates`` in this workbook),
            keep only the entry whose qualifier (lowercased) is in this
            tuple, in priority order. ``Accounts`` is the "actual"
            figure — strongly preferred over BE for historical years.
            For the latest two years RBI ships RE / BE only; those are
            kept as-is.
    """

    indicator_id: str
    item_label_match: str
    sign: int = 1
    prefer_qualifier: tuple[str, ...] = ("accounts",)


@dataclass(frozen=True)
class ParsedRow:
    entity_id: str
    time: str
    value: float | None


@dataclass
class ParsedIndicator:
    indicator_id: str
    rows: list[ParsedRow] = field(default_factory=list)
    sheet_count: int = 0
    period_count: int = 0


_NULL_TOKENS = {"", "—", "-", "–", "N.A.", "NA", "n.a.", "na", "..", "..."}


def _coerce_value(raw: Any, sign: int) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw) * sign
    text = str(raw).strip()
    if text in _NULL_TOKENS:
        return None
    # RBI sometimes wraps negatives in parens: "(-2.9)".
    text = text.replace(",", "")
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1].lstrip("-")
    try:
        return float(text) * sign
    except ValueError:
        return None


def _norm(text: Any) -> str:
    return " ".join(str(text or "").split()).lower()


def _find_header_row(ws) -> tuple[int, list[tuple[int, int, str | None]]]:
    """Locate the header row and return ``(row_index, [(col, year, qualifier)])``.

    Header is the first row whose first non-empty cell normalises to
    ``"item"`` AND that has at least 2 fiscal-year period headers in
    later columns.
    """
    for row_index, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30)), start=1
    ):
        cells = list(row)
        first_label = next(
            (c for c in cells if c.value is not None and str(c.value).strip()),
            None,
        )
        if first_label is None or _norm(first_label.value) != "item":
            continue
        periods: list[tuple[int, int, str | None]] = []
        for c in cells:
            parsed = _parse_period_header(c.value)
            if parsed is not None:
                periods.append((c.column, parsed[0], parsed[1]))
        if len(periods) >= 2:
            return row_index, periods
    raise RBIAppendixShapeError(
        f"sheet {ws.title!r}: no header row with 'Item' + ≥2 fiscal-year columns "
        f"found in the first 30 rows"
    )


def _find_item_row(
    ws, header_row: int, item_match: str
) -> int:
    """Return the row index whose first non-empty label contains ``item_match``."""
    needle = item_match.lower()
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        first_label = next(
            (c for c in row if c.value is not None and str(c.value).strip()),
            None,
        )
        if first_label is None:
            continue
        label_norm = _norm(first_label.value)
        # Ignore notes / source rows ("note:", "source:", "*:").
        if label_norm.startswith(("note", "source", "*")):
            continue
        if needle in label_norm:
            return first_label.row
    raise RBIAppendixShapeError(
        f"sheet {ws.title!r}: no row matching item label {item_match!r} found"
    )


def _select_qualifier(
    candidates: list[tuple[str | None, float | None]],
    preferred: tuple[str, ...],
) -> float | None:
    """Pick one value when a period has multiple qualifier variants."""
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0][1]
    # Try the preference list in order.
    by_qual: dict[str | None, float | None] = {}
    for qual, val in candidates:
        by_qual.setdefault(qual, val)
    for pref in preferred:
        for qual, val in by_qual.items():
            if qual is not None and pref in qual:
                return val
    # No preference matched — keep the first encountered (sheets are
    # iterated in workbook order).
    return candidates[0][1]


def parse_workbook(content: bytes, spec: AppendixSpec) -> ParsedIndicator:
    """Parse all sheets of an Appendix workbook for one indicator."""
    import io as _io

    wb: Workbook = load_workbook(_io.BytesIO(content), data_only=True)

    # Collect (start_year, qualifier, value) across every sheet, then
    # collapse duplicates by qualifier preference.
    by_year: dict[int, list[tuple[str | None, float | None]]] = {}
    period_count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row, periods = _find_header_row(ws)
        item_row = _find_item_row(ws, header_row, spec.item_label_match)
        for col, start_year, qualifier in periods:
            cell = ws.cell(row=item_row, column=col)
            value = _coerce_value(cell.value, spec.sign)
            by_year.setdefault(start_year, []).append((qualifier, value))
            period_count += 1

    rows: list[ParsedRow] = []
    for start_year in sorted(by_year):
        chosen = _select_qualifier(by_year[start_year], spec.prefer_qualifier)
        rows.append(
            ParsedRow(
                entity_id="IN",
                time=_fy_to_period(start_year),
                value=chosen,
            )
        )

    if not rows:
        raise RBIAppendixShapeError(
            f"no rows extracted for indicator {spec.indicator_id!r}"
        )

    return ParsedIndicator(
        indicator_id=spec.indicator_id,
        rows=rows,
        sheet_count=len(wb.sheetnames),
        period_count=period_count,
    )


# ---------------------------------------------------------------------------
# Shipped specs
# ---------------------------------------------------------------------------


SHIPPED_SPECS: tuple[AppendixSpec, ...] = (
    AppendixSpec(
        indicator_id="fiscal/national_centre_transfers_total",
        # Item VI: "Net Transfer of Resources from the Centre (IV-V)".
        # This is RBI's net-of-loan-repayments figure — what actually
        # flows from Centre to all States in aggregate.
        item_label_match="net transfer of resources",
        prefer_qualifier=("accounts",),
    ),
)
