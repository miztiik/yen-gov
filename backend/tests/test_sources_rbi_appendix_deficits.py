"""Pure-parser tests for ``yen_gov.sources.rbi_appendix_deficits.parsers``.

No fixture files; we build minimal in-memory workbooks with openpyxl
that mirror the shape of RBI's ``AppT1_MajorDeficitIndicators_<YYYY>.xlsx``:
empty column A, "Year" header in column B, four indicator headers in
columns C-F, a column-index row, then alternating year-data / %GDP rows.
"""
from __future__ import annotations

import io

import pytest
from openpyxl import Workbook

from yen_gov.sources.rbi_appendix_deficits.parsers import (
    DeficitSpec,
    RBIAppT1ShapeError,
    SHIPPED_SPECS,
    parse_workbook,
)


# Indicator labels used by the real workbook (verified against
# AppT1_MajorDeficitIndicators_2026.xlsx; see tools/rbi_appT1_inspect.py).
_HEADERS = (
    "Year",
    "Gross Fiscal Deficit",
    "Revenue Deficit",
    "Primary Deficit",
    "Primary Revenue Deficit",
    "Net RBI Credit to States",
)


def _wb_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_workbook(rows: list[list[object]]) -> bytes:
    """Build a single-sheet AppT1-shaped workbook.

    ``rows`` is the body — header / index / year-data / %GDP rows in the
    order they should appear. Each row is laid out as
    ``[colA_blank, year_or_blank, gfd, rd, pd, prd, net_rbi]`` (7 cells)
    to mirror the real workbook's empty leading column.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "APPT_1"
    # Two filler rows (real workbook has title in r1, "(₹ Crore)" in r2).
    ws.append([None] * 7)
    ws.append([None] * 7)
    for row in rows:
        ws.append(row)
    return _wb_to_bytes(wb)


def _header_row() -> list[object]:
    return [None, *list(_HEADERS)]  # 7 cells: blank A, Year+5 indicators


def _index_row() -> list[object]:
    return [None, None, 1, 2, 3, 4, 5]  # column-index row


def _year_row(year_label: str, gfd: object, rd: object, pd: object, prd: object, net: object) -> list[object]:
    return [None, year_label, gfd, rd, pd, prd, net]


def _gdp_row(gfd: object, rd: object, pd: object, prd: object, net: object) -> list[object]:
    return [None, None, gfd, rd, pd, prd, net]


# ---------------------------------------------------------------------------
# Happy path
# ---------------------------------------------------------------------------


def test_parse_workbook_emits_one_row_per_year_per_indicator():
    content = _build_workbook(
        [
            _header_row(),
            _index_row(),
            _year_row("2007-08", 75454.7, -42942.7, -24375.9, -142773.4, 1140),
            _gdp_row(-1.5, "(-0.9)", "(-0.5)", "(-2.9)", 0),
            _year_row("2008-09", 134589.3, -12672.2, 31634.5, -115627, -1608),
            _gdp_row(-2.4, "(-0.2)", 0.6, "(-2.1)", "(-0.0)"),
        ]
    )
    parsed = parse_workbook(content)

    # All four shipped specs got parsed.
    assert set(parsed) == {spec.indicator_id for spec in SHIPPED_SPECS}

    gfd = parsed["fiscal/national_gross_fiscal_deficit"]
    assert gfd.period_count == 2
    assert [(r.time, r.value) for r in gfd.rows] == [
        ("2007-04", 75454.7),
        ("2008-04", 134589.3),
    ]

    rd = parsed["fiscal/national_revenue_deficit"]
    assert [(r.time, r.value) for r in rd.rows] == [
        ("2007-04", -42942.7),
        ("2008-04", -12672.2),
    ]

    pd = parsed["fiscal/national_primary_deficit"]
    assert [(r.time, r.value) for r in pd.rows] == [
        ("2007-04", -24375.9),
        ("2008-04", 31634.5),
    ]

    prd = parsed["fiscal/national_primary_revenue_deficit"]
    assert [(r.time, r.value) for r in prd.rows] == [
        ("2007-04", -142773.4),
        ("2008-04", -115627.0),
    ]


def test_parse_workbook_skips_gdp_rows_not_year_data():
    """The %GDP row immediately after each year row must NOT be ingested."""
    content = _build_workbook(
        [
            _header_row(),
            _index_row(),
            _year_row("2010-11", 161461.1, -3050.6, 36640.7, -127870.9, 2515),
            # The %GDP row would otherwise look like another data row — but
            # has no year label in col B, so it's silently skipped.
            _gdp_row(-2.1, "(-0.0)", 0.5, "(-1.7)", 0),
        ]
    )
    parsed = parse_workbook(content)
    gfd = parsed["fiscal/national_gross_fiscal_deficit"]
    assert gfd.period_count == 1
    assert [r.value for r in gfd.rows] == [161461.1]
    # The -2.1 (%GDP) value must not have leaked in.
    assert -2.1 not in [r.value for r in gfd.rows]


def test_parse_workbook_handles_qualified_year_labels():
    """Year labels like '2024-25 (BE)$' should still parse to start year 2024."""
    content = _build_workbook(
        [
            _header_row(),
            _index_row(),
            _year_row("2023-24 (RE)", 877194.6, 93436.9, 370019.2, -413738.5, 5808.3),
            _gdp_row(-2.9, 0.3, 1.2, "(-1.4)", 0),
            _year_row("2024-25 (BE)$", 1039138.1, 80119.5, 475606.6, -483412.1, "-"),
            _gdp_row(-3.2, 0.2, 1.5, "(-1.5)", "-"),
        ]
    )
    parsed = parse_workbook(content)
    gfd = parsed["fiscal/national_gross_fiscal_deficit"]
    assert [(r.time, r.value) for r in gfd.rows] == [
        ("2023-04", 877194.6),
        ("2024-04", 1039138.1),
    ]


# ---------------------------------------------------------------------------
# Coercion edges
# ---------------------------------------------------------------------------


def test_coerce_strips_commas_and_handles_paren_negatives():
    """RBI-style negatives wrapped in parens (and comma-grouped numbers) decode."""
    content = _build_workbook(
        [
            _header_row(),
            _index_row(),
            # Mix: int, str-with-comma, paren-negative-str, plain str number, dash.
            _year_row("2015-16", 420670, "5,381.7", "(-208865.2)", "-3052", "-"),
            _gdp_row(-3.1, 0, "(-1.5)", "(-0.0)", "-"),
        ]
    )
    parsed = parse_workbook(content)
    rows = {ind: parsed[ind].rows[0].value for ind in parsed}
    assert rows["fiscal/national_gross_fiscal_deficit"] == 420670.0
    assert rows["fiscal/national_revenue_deficit"] == 5381.7
    assert rows["fiscal/national_primary_deficit"] == -208865.2  # paren -> negative
    assert rows["fiscal/national_primary_revenue_deficit"] == -3052.0


def test_coerce_returns_none_for_null_tokens():
    content = _build_workbook(
        [
            _header_row(),
            _index_row(),
            _year_row("2025-26 (BE)", "N.A.", "—", "..", None, "-"),
            _gdp_row(None, None, None, None, None),
        ]
    )
    parsed = parse_workbook(content)
    for ind_id, ind in parsed.items():
        assert len(ind.rows) == 1, f"{ind_id} should still emit a row with null value"
        assert ind.rows[0].value is None, f"{ind_id} value should be None"
        assert ind.rows[0].time == "2025-04"


# ---------------------------------------------------------------------------
# Column-resolution disambiguation
# ---------------------------------------------------------------------------


def test_column_resolution_prefers_exact_match_over_substring():
    """'primary deficit' is a substring of 'primary revenue deficit'.

    The resolver must pick the column whose header is exactly
    'primary deficit', not the 'primary revenue deficit' column.
    Otherwise both indicators would point at the same column.
    """
    content = _build_workbook(
        [
            _header_row(),
            _index_row(),
            _year_row("2020-21", 800000, 370000, 417000, -16000, 900),
            _gdp_row(-4.1, 1.9, 2.1, "(-0.1)", 0),
        ]
    )
    parsed = parse_workbook(content)
    pd_val = parsed["fiscal/national_primary_deficit"].rows[0].value
    prd_val = parsed["fiscal/national_primary_revenue_deficit"].rows[0].value
    assert pd_val == 417000.0
    assert prd_val == -16000.0
    assert pd_val != prd_val  # if the substring bug regressed, they'd be equal


# ---------------------------------------------------------------------------
# Shape errors fail loud (no silent zero-row artifact)
# ---------------------------------------------------------------------------


def test_missing_header_row_raises():
    """A workbook without 'Year' + indicator headers in the first 20 rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "APPT_1"
    for _ in range(5):
        ws.append([None, "filler", "filler"])
    with pytest.raises(RBIAppT1ShapeError, match="no header row"):
        parse_workbook(_wb_to_bytes(wb))


def test_no_year_rows_raises():
    """Header is present but no year-labelled body rows follow."""
    content = _build_workbook(
        [
            _header_row(),
            _index_row(),
            _year_row("not-a-year", 1, 2, 3, 4, 5),
            _gdp_row(0, 0, 0, 0, 0),
        ]
    )
    with pytest.raises(RBIAppT1ShapeError, match="no fiscal-year rows"):
        parse_workbook(content)


def test_unknown_column_label_raises():
    """An impossible-to-resolve spec must fail loud, not silently emit zero rows."""
    content = _build_workbook(
        [
            _header_row(),
            _index_row(),
            _year_row("2020-21", 1, 2, 3, 4, 5),
            _gdp_row(0, 0, 0, 0, 0),
        ]
    )
    bogus = (DeficitSpec(indicator_id="fiscal/x", column_label_match="not a real label"),)
    with pytest.raises(RBIAppT1ShapeError, match="no column matched"):
        parse_workbook(content, specs=bogus)


def test_empty_workbook_raises():
    wb = Workbook()
    wb.remove(wb.active)
    # openpyxl needs at least one sheet to save; add a dummy then remove pretext.
    wb.create_sheet("placeholder")
    # Replace with a workbook that has no usable sheet content.
    empty = Workbook()
    empty.active.title = "APPT_1"
    with pytest.raises(RBIAppT1ShapeError):
        parse_workbook(_wb_to_bytes(empty))


# ---------------------------------------------------------------------------
# Spec catalogue invariants
# ---------------------------------------------------------------------------


def test_shipped_specs_are_distinct():
    ids = [spec.indicator_id for spec in SHIPPED_SPECS]
    matches = [spec.column_label_match for spec in SHIPPED_SPECS]
    assert len(ids) == len(set(ids)), "duplicate indicator id in SHIPPED_SPECS"
    assert len(matches) == len(set(matches)), "duplicate column match in SHIPPED_SPECS"


def test_shipped_specs_cover_expected_indicator_set():
    """If a future commit adds/removes a deficit indicator, this test must be updated."""
    assert {spec.indicator_id for spec in SHIPPED_SPECS} == {
        "fiscal/national_gross_fiscal_deficit",
        "fiscal/national_revenue_deficit",
        "fiscal/national_primary_deficit",
        "fiscal/national_primary_revenue_deficit",
    }
