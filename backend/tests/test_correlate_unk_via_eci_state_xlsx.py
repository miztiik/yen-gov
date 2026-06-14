"""Tests for ``tools.correlate_unk_via_eci_state_xlsx``.

Covers header detection, join + resolve, NOTA/IND sentinels, unicode
normalisation, internal collision, existing collision disambiguation,
unmapped state-year handling, and year derivation from xlsx content.

All xlsx fixtures are real (Holy Law #7) built with openpyxl under
``tmp_path``.
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

import openpyxl
import pytest

REPO_ROOT = Path(__file__).resolve().parents[2]
TOOLS_DIR = REPO_ROOT / "tools"
if str(TOOLS_DIR.parent) not in sys.path:
    sys.path.insert(0, str(TOOLS_DIR.parent))

from tools.correlate_unk_via_eci_state_xlsx.__main__ import (  # noqa: E402
    correlate,
    derive_year_for_file,
    load_eci_xlsx_index,
    normalise,
    resolve_eci_label,
    _detect_header_row,
    STATE_SLUG_TO_ISO,
    VERDICT_FIELDNAMES,
    SKIPPED_FIELDNAMES,
)


# ---------------------------------------------------------------------------
# Fixture helpers
# ---------------------------------------------------------------------------


def _write_eci_xlsx(
    path: Path,
    *,
    state_name: str = "Uttar Pradesh",
    rows: list[tuple],
    title: str = "10 - Detailed Results",
) -> None:
    """Write a 14-column ECI Detailed Results xlsx mirroring shape A.

    Header is laid out at row 4 (1-indexed) -> openpyxl row 4 = 0-indexed
    row 3, matching the real ECI files. The 4 ``rows`` columns are
    ``(ac_no, ac_name, candidate_name, party_abbr)`` and the writer
    fills the remaining columns with deterministic placeholders.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Worksheet"
    ws.cell(row=1, column=1, value=title)
    ws.cell(row=3, column=10, value="VALID VOTES POLLED")
    headers = [
        "STATE/UT NAME",
        "AC NO.",
        "AC NAME",
        "CANDIDATE NAME",
        "SEX",
        "AGE",
        "CATEGORY",
        "PARTY",
        "SYMBOL",
        "GENERAL",
        "POSTAL",
        "TOTAL",
        "% VOTES POLLED",
        "TOTAL ELECTORS",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col_idx, value=h)
    for r_idx, (ac_no, ac_name, cand, party) in enumerate(rows, start=5):
        ws.cell(row=r_idx, column=1, value=state_name)
        ws.cell(row=r_idx, column=2, value=ac_no)
        ws.cell(row=r_idx, column=3, value=ac_name)
        ws.cell(row=r_idx, column=4, value=cand)
        ws.cell(row=r_idx, column=5, value="MALE")
        ws.cell(row=r_idx, column=6, value=40)
        ws.cell(row=r_idx, column=7, value="GENERAL")
        ws.cell(row=r_idx, column=8, value=party)
        ws.cell(row=r_idx, column=9, value="Symbol")
        ws.cell(row=r_idx, column=10, value=1000)
        ws.cell(row=r_idx, column=11, value=10)
        ws.cell(row=r_idx, column=12, value=1010)
        ws.cell(row=r_idx, column=13, value=10.0)
        ws.cell(row=r_idx, column=14, value=10000)
    wb.save(path)


def _write_candidacies_csv(
    path: Path,
    rows: list[dict[str, str]],
) -> None:
    """Write a TCPD-shape assembly candidacies.csv under tmp_path."""
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "entity_id",
        "state",
        "election_year",
        "constituency_no",
        "constituency_name",
        "candidate_name",
        "party_id",
        "party_short_raw",
        "votes",
        "vote_share_pct",
        "position",
        "result",
        "sex",
        "age",
        "education",
        "profession",
        "candidate_type",
        "source_id",
        "processing_level",
        "processing_note",
    ]
    defaults = {
        "entity_id": "IN-AC-2008-fixture",
        "votes": "1",
        "vote_share_pct": "0.5",
        "position": "1",
        "result": "lost",
        "sex": "U",
        "age": "40",
        "education": "Graduate",
        "profession": "Self",
        "candidate_type": "challenger",
        "source_id": "src-fixture",
        "processing_level": "minor",
        "processing_note": "",
    }
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, lineterminator="\n")
        w.writeheader()
        for row in rows:
            w.writerow({**defaults, **row})


def _write_parties_csv(path: Path, rows: list[dict[str, str]]) -> None:
    """Write parties.csv with the on-disk 18-col schema."""
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "party_id",
        "short",
        "full",
        "eci_codes",
        "brand_colour",
        "symbol_asset",
        "wikipedia",
        "aliases",
        "recognition_scope",
        "home_state_codes",
        "founded_year",
        "dissolved_year",
        "predecessor_party_ids",
        "successor_party_ids",
        "name_history",
        "claims_to_parent_name",
        "name_native_script",
        "is_sentinel",
    ]
    base = {k: "" for k in fieldnames}
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, lineterminator="\n")
        w.writeheader()
        for row in rows:
            w.writerow({**base, **row})


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_header_detection_uttar_pradesh_2022_shape(tmp_path: Path) -> None:
    """ECI 2022/2023 14-col shape: header at row 4 (openpyxl 0-indexed 3)."""
    xlsx = tmp_path / "fixture.xlsx"
    _write_eci_xlsx(
        xlsx,
        rows=[
            (1, "Behat", "1 Umar Ali Khan", "SP"),
            (1, "Behat", "2 Naresh Saini", "BJP"),
        ],
    )
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_idx, header = _detect_header_row(ws)
    wb.close()
    assert header_idx == 3, f"expected header at row 4 (0-indexed 3), got {header_idx}"
    assert "AC NAME" in header
    assert "CANDIDATE NAME" in header
    assert "PARTY" in header


def test_state_constituency_candidate_join_resolves(tmp_path: Path) -> None:
    """Fixture ECI + UNK candidacies join correctly and produce verdicts."""
    # Build a UP-2022 ECI fixture
    eci_path = tmp_path / "2022_uttar_pradesh_10-Detailed Results.xlsx"
    _write_eci_xlsx(
        eci_path,
        rows=[
            (117, "Dataganj", "10 MUNNA LAL", "RSUP"),
            (1, "Behat", "1 Umar Ali Khan", "SP"),
        ],
    )
    # Build a TCPD candidacies.csv with one UNK matching the ECI row
    cand_path = (
        tmp_path
        / "datasets"
        / "elections"
        / "assembly"
        / "state=uttar-pradesh"
        / "election=2022"
        / "candidacies.csv"
    )
    _write_candidacies_csv(
        cand_path,
        [
            {
                "state": "uttar-pradesh",
                "election_year": "2022",
                "constituency_no": "117",
                "constituency_name": "DATAGANJ",
                "candidate_name": "MUNNA LAL",
                "party_id": "parties.IN.UNK",
                "party_short_raw": "Rashtra Uday Party",
            },
        ],
    )
    # parties.csv WITHOUT RSUP (mint expected)
    parties_path = tmp_path / "datasets" / "data" / "entities" / "parties.csv"
    _write_parties_csv(
        parties_path,
        [
            {"party_id": "parties.IN.UNK", "short": "UNK", "full": "Unknown", "is_sentinel": "true"},
            {"party_id": "parties.IN.NOTA", "short": "NOTA", "full": "NOTA", "is_sentinel": "true"},
            {"party_id": "parties.IN.IND", "short": "IND", "full": "Independent", "is_sentinel": "true"},
            {"party_id": "parties.IN.SP", "short": "SP", "full": "Samajwadi Party"},
        ],
    )
    result = correlate(
        repo_root=tmp_path,
        parties_csv=parties_path,
        ephemeral_dir=tmp_path,
        verdict_root=tmp_path / "verdict",
        state_xlsx_map={
            "2022_uttar_pradesh_10-Detailed Results.xlsx": ("uttar-pradesh", 2022),
        },
    )
    # Read the verdict
    with result["verdict_path"].open(encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    assert len(rows) == 1, f"expected 1 verdict row, got {len(rows)}: {rows!r}"
    v = rows[0]
    assert v["action"] == "mint-new"
    assert v["proposed_party_id"] == "parties.IN.RSUP"
    assert v["party_short_raw"] == "Rashtra Uday Party"
    assert v["tcpd_frequent_abbrev"] == "RSUP"
    assert v["state"] == "uttar-pradesh"


def test_nota_row_maps_to_sentinel(tmp_path: Path) -> None:
    """ECI NOTA row resolves to parties.IN.NOTA (no mint)."""
    eci_path = tmp_path / "2022_uttar_pradesh_10-Detailed Results.xlsx"
    _write_eci_xlsx(
        eci_path,
        rows=[
            (1, "Behat", "6 Nota", "NOTA"),
        ],
    )
    cand_path = (
        tmp_path
        / "datasets"
        / "elections"
        / "assembly"
        / "state=uttar-pradesh"
        / "election=2022"
        / "candidacies.csv"
    )
    _write_candidacies_csv(
        cand_path,
        [
            {
                "state": "uttar-pradesh",
                "election_year": "2022",
                "constituency_no": "1",
                "constituency_name": "Behat",
                "candidate_name": "Nota",
                "party_id": "parties.IN.UNK",
                "party_short_raw": "None Of The Above",
            },
        ],
    )
    parties_path = tmp_path / "datasets" / "data" / "entities" / "parties.csv"
    _write_parties_csv(
        parties_path,
        [
            {"party_id": "parties.IN.UNK", "short": "UNK", "full": "Unknown", "is_sentinel": "true"},
            {"party_id": "parties.IN.NOTA", "short": "NOTA", "full": "NOTA", "is_sentinel": "true"},
            {"party_id": "parties.IN.IND", "short": "IND", "full": "Independent", "is_sentinel": "true"},
        ],
    )
    result = correlate(
        repo_root=tmp_path,
        parties_csv=parties_path,
        ephemeral_dir=tmp_path,
        verdict_root=tmp_path / "verdict",
        state_xlsx_map={
            "2022_uttar_pradesh_10-Detailed Results.xlsx": ("uttar-pradesh", 2022),
        },
    )
    with result["verdict_path"].open(encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    assert len(rows) == 1
    v = rows[0]
    assert v["action"] == "alias-add"
    assert v["proposed_party_id"] == "parties.IN.NOTA"


def test_independent_row_maps_to_sentinel(tmp_path: Path) -> None:
    """ECI IND row resolves to parties.IN.IND (no mint)."""
    eci_path = tmp_path / "2022_uttar_pradesh_10-Detailed Results.xlsx"
    _write_eci_xlsx(
        eci_path,
        rows=[
            (1, "Behat", "3 Hakumat Singh", "IND"),
        ],
    )
    cand_path = (
        tmp_path
        / "datasets"
        / "elections"
        / "assembly"
        / "state=uttar-pradesh"
        / "election=2022"
        / "candidacies.csv"
    )
    _write_candidacies_csv(
        cand_path,
        [
            {
                "state": "uttar-pradesh",
                "election_year": "2022",
                "constituency_no": "1",
                "constituency_name": "Behat",
                "candidate_name": "Hakumat Singh",
                "party_id": "parties.IN.UNK",
                "party_short_raw": "Independents",
            },
        ],
    )
    parties_path = tmp_path / "datasets" / "data" / "entities" / "parties.csv"
    _write_parties_csv(
        parties_path,
        [
            {"party_id": "parties.IN.UNK", "short": "UNK", "full": "Unknown", "is_sentinel": "true"},
            {"party_id": "parties.IN.NOTA", "short": "NOTA", "full": "NOTA", "is_sentinel": "true"},
            {"party_id": "parties.IN.IND", "short": "IND", "full": "Independent", "is_sentinel": "true"},
        ],
    )
    result = correlate(
        repo_root=tmp_path,
        parties_csv=parties_path,
        ephemeral_dir=tmp_path,
        verdict_root=tmp_path / "verdict",
        state_xlsx_map={
            "2022_uttar_pradesh_10-Detailed Results.xlsx": ("uttar-pradesh", 2022),
        },
    )
    with result["verdict_path"].open(encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    assert len(rows) == 1
    v = rows[0]
    assert v["action"] == "alias-add"
    assert v["proposed_party_id"] == "parties.IN.IND"


def test_unicode_punctuation_normalisation(tmp_path: Path) -> None:
    """Accented + en-dash variants still join across normalised key."""
    # ECI has plain ASCII; candidacies has accented + en-dash.
    eci_path = tmp_path / "2022_uttar_pradesh_10-Detailed Results.xlsx"
    _write_eci_xlsx(
        eci_path,
        rows=[
            # En-dash + accented characters in ECI side
            (1, "Behat\u2013Town", "1 Andr\u00e9 Garcia", "RSUP"),
        ],
    )
    cand_path = (
        tmp_path
        / "datasets"
        / "elections"
        / "assembly"
        / "state=uttar-pradesh"
        / "election=2022"
        / "candidacies.csv"
    )
    _write_candidacies_csv(
        cand_path,
        [
            # TCPD side: ASCII hyphen + plain "e"
            {
                "state": "uttar-pradesh",
                "election_year": "2022",
                "constituency_no": "1",
                "constituency_name": "Behat-Town",
                "candidate_name": "Andre Garcia",
                "party_id": "parties.IN.UNK",
                "party_short_raw": "Rashtra Uday Party",
            },
        ],
    )
    parties_path = tmp_path / "datasets" / "data" / "entities" / "parties.csv"
    _write_parties_csv(
        parties_path,
        [
            {"party_id": "parties.IN.UNK", "short": "UNK", "full": "Unknown", "is_sentinel": "true"},
            {"party_id": "parties.IN.NOTA", "short": "NOTA", "full": "NOTA", "is_sentinel": "true"},
            {"party_id": "parties.IN.IND", "short": "IND", "full": "Independent", "is_sentinel": "true"},
        ],
    )
    result = correlate(
        repo_root=tmp_path,
        parties_csv=parties_path,
        ephemeral_dir=tmp_path,
        verdict_root=tmp_path / "verdict",
        state_xlsx_map={
            "2022_uttar_pradesh_10-Detailed Results.xlsx": ("uttar-pradesh", 2022),
        },
    )
    with result["verdict_path"].open(encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    assert len(rows) == 1, f"expected 1 verdict row (unicode join must succeed), got {rows!r}"


def test_eci_internal_collision_skips(tmp_path: Path) -> None:
    """One publisher label -> 2 different ECI abbrs -> internal collision."""
    eci_path = tmp_path / "2022_uttar_pradesh_10-Detailed Results.xlsx"
    _write_eci_xlsx(
        eci_path,
        rows=[
            (1, "Behat", "1 Cand A", "PARTY1"),
            (2, "Other", "1 Cand B", "PARTY2"),
        ],
    )
    cand_path = (
        tmp_path
        / "datasets"
        / "elections"
        / "assembly"
        / "state=uttar-pradesh"
        / "election=2022"
        / "candidacies.csv"
    )
    _write_candidacies_csv(
        cand_path,
        [
            # Both UNK rows carry the SAME party_short_raw but join to
            # DIFFERENT ECI abbrs -> internal collision.
            {
                "state": "uttar-pradesh",
                "election_year": "2022",
                "constituency_no": "1",
                "constituency_name": "Behat",
                "candidate_name": "Cand A",
                "party_id": "parties.IN.UNK",
                "party_short_raw": "Ambiguous Label",
            },
            {
                "state": "uttar-pradesh",
                "election_year": "2022",
                "constituency_no": "2",
                "constituency_name": "Other",
                "candidate_name": "Cand B",
                "party_id": "parties.IN.UNK",
                "party_short_raw": "Ambiguous Label",
            },
        ],
    )
    parties_path = tmp_path / "datasets" / "data" / "entities" / "parties.csv"
    _write_parties_csv(
        parties_path,
        [
            {"party_id": "parties.IN.UNK", "short": "UNK", "full": "Unknown", "is_sentinel": "true"},
            {"party_id": "parties.IN.NOTA", "short": "NOTA", "full": "NOTA", "is_sentinel": "true"},
            {"party_id": "parties.IN.IND", "short": "IND", "full": "Independent", "is_sentinel": "true"},
        ],
    )
    result = correlate(
        repo_root=tmp_path,
        parties_csv=parties_path,
        ephemeral_dir=tmp_path,
        verdict_root=tmp_path / "verdict",
        state_xlsx_map={
            "2022_uttar_pradesh_10-Detailed Results.xlsx": ("uttar-pradesh", 2022),
        },
    )
    assert result["verdict_rows"] == 0
    with result["skipped_path"].open(encoding="utf-8") as f:
        skipped = list(csv.DictReader(f))
    assert any(s["reason"] == "eci-internal-collision" for s in skipped)


def test_existing_collision_disambiguated_mint(tmp_path: Path) -> None:
    """Naive mint pid collides; disambiguates to <abbr>_AE<year>_<state>."""
    eci_path = tmp_path / "2022_uttar_pradesh_10-Detailed Results.xlsx"
    _write_eci_xlsx(
        eci_path,
        rows=[
            (1, "Behat", "1 New Person", "DUP"),
        ],
    )
    cand_path = (
        tmp_path
        / "datasets"
        / "elections"
        / "assembly"
        / "state=uttar-pradesh"
        / "election=2022"
        / "candidacies.csv"
    )
    _write_candidacies_csv(
        cand_path,
        [
            {
                "state": "uttar-pradesh",
                "election_year": "2022",
                "constituency_no": "1",
                "constituency_name": "Behat",
                "candidate_name": "New Person",
                "party_id": "parties.IN.UNK",
                "party_short_raw": "New Party",
            },
        ],
    )
    parties_path = tmp_path / "datasets" / "data" / "entities" / "parties.csv"
    _write_parties_csv(
        parties_path,
        [
            {"party_id": "parties.IN.UNK", "short": "UNK", "full": "Unknown", "is_sentinel": "true"},
            {"party_id": "parties.IN.NOTA", "short": "NOTA", "full": "NOTA", "is_sentinel": "true"},
            {"party_id": "parties.IN.IND", "short": "IND", "full": "Independent", "is_sentinel": "true"},
            # Pre-existing parties.IN.DUP whose ``short`` + ``aliases``
            # do NOT include "DUP". This forces resolve_eci_label() past
            # the short/alias lookup paths into the mint branch, where
            # the naive parties.IN.DUP base collides with this existing
            # pid -> disambiguation triggers.
            {
                "party_id": "parties.IN.DUP",
                "short": "DIFFERENT_SHORT",
                "full": "Different Party Already",
                "aliases": "DIFFERENT_ALIAS",
            },
        ],
    )
    result = correlate(
        repo_root=tmp_path,
        parties_csv=parties_path,
        ephemeral_dir=tmp_path,
        verdict_root=tmp_path / "verdict",
        state_xlsx_map={
            "2022_uttar_pradesh_10-Detailed Results.xlsx": ("uttar-pradesh", 2022),
        },
    )
    with result["verdict_path"].open(encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    assert len(rows) == 1
    v = rows[0]
    assert v["action"] == "mint-new"
    # Disambiguated: parties.IN.DUP exists, so mint as
    # parties.IN.DUP_AE2022_UP (UP = STATE_SLUG_TO_ISO["uttar-pradesh"]).
    assert v["proposed_party_id"] == "parties.IN.DUP_AE2022_UP", (
        f"expected disambiguated pid, got {v['proposed_party_id']!r}"
    )


def test_unmapped_state_event_does_not_touch_corpus(tmp_path: Path) -> None:
    """ECI xlsx mapped to a state-year with no candidacies.csv emits 0 verdicts."""
    eci_path = tmp_path / "2022_uttar_pradesh_10-Detailed Results.xlsx"
    _write_eci_xlsx(
        eci_path,
        rows=[
            (1, "Behat", "1 Cand", "RSUP"),
        ],
    )
    # Note: NO candidacies.csv is written under
    # datasets/elections/assembly/state=uttar-pradesh/election=2022/.
    parties_path = tmp_path / "datasets" / "data" / "entities" / "parties.csv"
    _write_parties_csv(
        parties_path,
        [
            {"party_id": "parties.IN.UNK", "short": "UNK", "full": "Unknown", "is_sentinel": "true"},
            {"party_id": "parties.IN.NOTA", "short": "NOTA", "full": "NOTA", "is_sentinel": "true"},
            {"party_id": "parties.IN.IND", "short": "IND", "full": "Independent", "is_sentinel": "true"},
        ],
    )
    result = correlate(
        repo_root=tmp_path,
        parties_csv=parties_path,
        ephemeral_dir=tmp_path,
        verdict_root=tmp_path / "verdict",
        state_xlsx_map={
            "2022_uttar_pradesh_10-Detailed Results.xlsx": ("uttar-pradesh", 2022),
        },
    )
    assert result["verdict_rows"] == 0
    assert result["skipped_rows"] == 0


def test_year_derivation_from_xlsx_content(tmp_path: Path) -> None:
    """Tier 2 fallback: filename has no year, but xlsx title contains a year."""
    # Filename lacks any 20XX token; tier 1 (filename) misses.
    xlsx = tmp_path / "kerala_10-Detailed_Results_99999.xlsx"
    _write_eci_xlsx(
        xlsx,
        rows=[(1, "Manjeshwar", "1 A K M Ashraf", "IUML")],
        title="Detailed Results - General Election 2021 Kerala AE",
    )
    # Pass a different hardcoded value to prove tier-2 wins over tier-3.
    year, source = derive_year_for_file(xlsx, xlsx.name, hardcoded=1999)
    assert year == 2021, f"expected tier-2 year=2021, got {year}"
    assert source == "xlsx-title"


def test_year_derivation_filename_wins_over_xlsx_title(tmp_path: Path) -> None:
    """Tier 1 (filename year) takes priority over tier 2 (xlsx title)."""
    xlsx = tmp_path / "2022_uttar_pradesh_10-Detailed Results.xlsx"
    _write_eci_xlsx(
        xlsx,
        rows=[(1, "Behat", "1 Cand", "SP")],
        title="Detailed Results - 2017 Election (stale title)",
    )
    year, source = derive_year_for_file(xlsx, xlsx.name, hardcoded=2099)
    assert year == 2022
    assert source == "filename"


def test_year_derivation_hardcoded_fallback(tmp_path: Path) -> None:
    """Tier 3 (hardcoded) when neither filename nor xlsx title carry a year."""
    xlsx = tmp_path / "wb_10-Detailed_Results_99999.xlsx"
    _write_eci_xlsx(
        xlsx,
        rows=[(1, "Mekliganj", "1 Cand", "BJP")],
        title="Detailed Results",  # no year in title
    )
    year, source = derive_year_for_file(xlsx, xlsx.name, hardcoded=2021)
    assert year == 2021
    assert source == "hardcoded"


# ---------------------------------------------------------------------------
# Additional unit-level coverage
# ---------------------------------------------------------------------------


def test_normalise_strips_position_prefix_and_paren_suffix() -> None:
    assert normalise("10 MUNNA LAL") == "MUNNA LAL"
    assert normalise("1 Umar Ali Khan") == "UMAR ALI KHAN"
    assert normalise("Churah (SC)") == "CHURAH"
    assert normalise("MEKLIGANJ (ST)") == "MEKLIGANJ"


def test_normalise_unicode_decomp_and_punct() -> None:
    # NFKD-strips combining acute; en-dash collapses to space then drops.
    assert normalise("Andr\u00e9-Garcia") == "ANDRE GARCIA"
    assert normalise("Andr\u00e9\u2013Garcia") == "ANDRE GARCIA"


def test_resolve_eci_label_sentinels_and_existing() -> None:
    existing = {"parties.IN.SP"}
    short_to_pid = {"SP": "parties.IN.SP"}
    alias_to_pid = {"SAMAJWADI": "parties.IN.SP"}
    a, pid, _ = resolve_eci_label("NOTA", existing, short_to_pid, alias_to_pid, 2022, "UP")
    assert (a, pid) == ("alias-add", "parties.IN.NOTA")
    a, pid, _ = resolve_eci_label("IND", existing, short_to_pid, alias_to_pid, 2022, "UP")
    assert (a, pid) == ("alias-add", "parties.IN.IND")
    a, pid, _ = resolve_eci_label("SP", existing, short_to_pid, alias_to_pid, 2022, "UP")
    assert (a, pid) == ("alias-add", "parties.IN.SP")
    a, pid, dis = resolve_eci_label("NEWX", existing, short_to_pid, alias_to_pid, 2022, "UP")
    assert (a, pid, dis) == ("mint-new", "parties.IN.NEWX", False)


def test_state_slug_to_iso_covers_all_xlsx_map() -> None:
    """Every state slug in STATE_XLSX_MAP must have an ISO code."""
    from tools.correlate_unk_via_eci_state_xlsx.__main__ import STATE_XLSX_MAP

    for _fname, (slug, _yr) in STATE_XLSX_MAP.items():
        assert slug in STATE_SLUG_TO_ISO, f"missing ISO for slug={slug!r}"


def test_verdict_schema_matches_apply_consumer() -> None:
    """Verdict fieldnames must match what correlate_unk_apply reads."""
    from tools.correlate_unk_apply.__main__ import PARTIES_FIELDNAMES  # noqa: F401

    # The apply tool reads `action`, `proposed_party_id`,
    # `party_short_raw`, `tcpd_frequent_abbrev`, `tcpd_party_name`,
    # `tcpd_party_type`, `state`, `tcpd_start_year`, `tcpd_last_year`.
    expected = {
        "action",
        "proposed_party_id",
        "party_short_raw",
        "tcpd_frequent_abbrev",
        "tcpd_party_name",
        "tcpd_party_type",
        "state",
        "tcpd_start_year",
        "tcpd_last_year",
    }
    assert set(VERDICT_FIELDNAMES) == expected
