"""Pure-parser tests for backend/yen_gov/sources/rbi_xlsx/parsers.py.

Synthetic in-memory workbooks model the actual RBI Statement layout
documented in tools/rbi_inspect.py output for Statement 20:

    row 1: title text
    row 2: unit annotation
    row 3: empty
    row 4: header row — col 1 = "State/UT", cols 2..N = year periods
    row 5: column-index row — "1 2 3 …" (skipped)
    rows 6..: data rows — col 1 = "<ordinal>. <Name>", cols 2..N = values
    trailer rows: "All States and UTs #" aggregate, notes, source

No real RBI bytes touch the test suite.
"""
from __future__ import annotations

import io

import pytest
from openpyxl import Workbook

from yen_gov.sources.rbi_xlsx.parsers import (
    IndicatorSpec,
    RBIWorkbookShapeError,
    SHIPPED_SPECS,
    _coerce_value,
    _normalise_period,
    normalise_state_label,
    parse_workbook,
)


# ---------------------------------------------------------------------------
# normalise_state_label
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "raw,expected",
    [
        ("Andhra Pradesh", "S01"),
        ("1. Andhra Pradesh", "S01"),
        ("29. Jammu and Kashmir", "S09"),
        ("J&K", "S09"),
        ("  Tamil  Nadu  ", "S22"),
        ("Telangana", "S29"),
        ("Orissa", "S18"),  # legacy spelling
        ("NCT Delhi", "U05"),
        ("Puducherry", "U07"),
        ("All States and UTs", None),
        ("", None),
        (None, None),
    ],
)
def test_normalise_state_label(raw, expected):
    assert normalise_state_label(raw) == expected


# ---------------------------------------------------------------------------
# _normalise_period
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "raw,kind,expected",
    [
        ("2026", "fy_end_year", ("2026-03", None)),
        ("2025 (RE)", "fy_end_year", ("2025-03", "RE")),
        ("2026 (BE)", "fy_end_year", ("2026-03", "BE")),
        ("2022-23", "fy_span", ("2022-04", None)),
        ("2024-25 (BE)", "fy_span", ("2024-04", "BE")),
    ],
)
def test_normalise_period(raw, kind, expected):
    assert _normalise_period(raw, period_kind=kind) == expected


def test_normalise_period_unknown_kind_raises():
    with pytest.raises(RBIWorkbookShapeError):
        _normalise_period("2026", period_kind="never_heard")


def test_normalise_period_garbage_raises():
    with pytest.raises(RBIWorkbookShapeError):
        _normalise_period("not a year", period_kind="fy_end_year")


# ---------------------------------------------------------------------------
# _coerce_value
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "raw,expected",
    [
        (None, None),
        ("", None),
        ("—", None),
        ("–", None),
        ("-", None),
        ("N.A.", None),
        ("..", None),
        (3.14, 3.14),
        ("3.14", 3.14),
        ("1,234.5", 1234.5),
        ("(2.5)", -2.5),  # parens-as-negative
        ("not a number", None),
    ],
)
def test_coerce_value_default_sign(raw, expected):
    assert _coerce_value(raw, sign=1) == expected


def test_coerce_value_with_sign_flip():
    assert _coerce_value("3.14", sign=-1) == -3.14
    assert _coerce_value("(2.5)", sign=-1) == 2.5


# ---------------------------------------------------------------------------
# parse_workbook (synthetic Statement-20-style workbook)
# ---------------------------------------------------------------------------


def _make_synthetic_workbook(
    *,
    sheet_name: str = "ST_20",
    header_label: str = "State/UT",
    states: list[str] | None = None,
    years: list[str] | None = None,
    values: list[list[float | str | None]] | None = None,
    include_index_row: bool = True,
    include_aggregate_row: bool = True,
) -> bytes:
    if states is None:
        states = ["1. Andhra Pradesh", "2. Bihar"]
    if years is None:
        years = ["2024", "2025 (RE)", "2026 (BE)"]
    if values is None:
        values = [
            [33.6, 35.5, 35.4],
            [39.4, 38.0, 36.8],
        ]

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Title rows
    ws.append([])
    ws.append(["", "Statement 20: Total Outstanding Liabilities"])
    ws.append(["", "(Per cent)"])

    # Header row at row 4 (1-indexed)
    ws.append(["", header_label, *years])

    # Column-index row "1 2 3 …"
    if include_index_row:
        ws.append(["", "1", *[str(i + 2) for i in range(len(years))]])

    # Data rows
    for state, row_vals in zip(states, values):
        ws.append(["", state, *row_vals])

    if include_aggregate_row:
        # Aggregate row that must be skipped
        ws.append(["", "All States and UTs #", *[None] * len(years)])
        ws.append(["", "Notes: ..."])
        ws.append(["", "Source: RBI"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_SPEC = IndicatorSpec(
    indicator_id="fiscal/outstanding_debt_pct_gsdp",
    sheet_match="ST_20",
    header_label_match="state",
    period_kind="fy_end_year",
)


def test_parse_workbook_happy_path():
    wb_bytes = _make_synthetic_workbook()
    parsed = parse_workbook(wb_bytes, _SPEC)
    # 2 states × 3 years = 6 rows
    assert len(parsed.rows) == 6
    assert parsed.sheet_name == "ST_20"
    assert parsed.period_columns == 3
    assert parsed.unmatched_states == []

    # Verify per-state coverage
    by_state = {}
    for r in parsed.rows:
        by_state.setdefault(r.entity_id, []).append((r.time, r.value, r.facet))
    assert set(by_state.keys()) == {"S01", "S04"}
    ap = sorted(by_state["S01"])
    assert ap == [
        ("2024-03", 33.6, None),
        ("2025-03", 35.5, "RE"),
        ("2026-03", 35.4, "BE"),
    ]


def test_parse_workbook_skips_aggregate_row():
    wb_bytes = _make_synthetic_workbook(include_aggregate_row=True)
    parsed = parse_workbook(wb_bytes, _SPEC)
    # Aggregate row is excluded → only the 2 state rows materialise
    assert len(parsed.rows) == 6
    for r in parsed.rows:
        assert r.entity_id in {"S01", "S04"}


def test_parse_workbook_records_unmatched_states():
    wb_bytes = _make_synthetic_workbook(
        states=["1. Andhra Pradesh", "99. Atlantis"],
        values=[[10.0, 11.0, 12.0], [None, None, None]],
    )
    parsed = parse_workbook(wb_bytes, _SPEC)
    assert "99. Atlantis" in parsed.unmatched_states
    assert {r.entity_id for r in parsed.rows} == {"S01"}


def test_parse_workbook_handles_blanks_as_null():
    wb_bytes = _make_synthetic_workbook(
        states=["1. Telangana"],
        values=[["—", "—", 27.5]],
    )
    parsed = parse_workbook(wb_bytes, _SPEC)
    assert len(parsed.rows) == 3
    vals = sorted([(r.time, r.value) for r in parsed.rows])
    assert vals == [("2024-03", None), ("2025-03", None), ("2026-03", 27.5)]


def test_parse_workbook_raises_on_missing_sheet():
    wb_bytes = _make_synthetic_workbook(sheet_name="ST_99")
    with pytest.raises(RBIWorkbookShapeError, match="no sheet matched"):
        parse_workbook(wb_bytes, _SPEC)


def test_parse_workbook_raises_on_missing_header():
    wb_bytes = _make_synthetic_workbook(header_label="Region/Province")
    with pytest.raises(RBIWorkbookShapeError, match="no header row contained"):
        parse_workbook(wb_bytes, _SPEC)


def test_parse_workbook_raises_when_no_state_rows():
    # Both rows are unknown labels → no per-state output → raise
    wb_bytes = _make_synthetic_workbook(
        states=["1. Atlantis", "2. Wakanda"],
        values=[[1.0, 2.0, 3.0], [4.0, 5.0, 6.0]],
    )
    with pytest.raises(RBIWorkbookShapeError, match="no per-state rows"):
        parse_workbook(wb_bytes, _SPEC)


def test_shipped_specs_present_and_well_formed():
    assert len(SHIPPED_SPECS) >= 1
    ids = {s.indicator_id for s in SHIPPED_SPECS}
    assert "fiscal/outstanding_debt_pct_gsdp" in ids
    for s in SHIPPED_SPECS:
        assert s.period_kind in {"fy_end_year", "fy_span"}
        assert s.sign in {1, -1}


# ---------------------------------------------------------------------------
# Sub-column selector (Statement-17-style stacked Gross | Net headers)
# ---------------------------------------------------------------------------


def _make_subcolumn_workbook(
    *,
    sheet_name: str = "ST_17",
    states: list[str] | None = None,
    fiscal_years: list[str] | None = None,
    sub_labels: list[str] | None = None,
    values_per_state: list[list[float | None]] | None = None,
) -> bytes:
    """Build a synthetic Statement-17-shaped workbook.

    Layout:
      r00 blank
      r01 ['', '<title>']
      r02 ['', '(₹ Crore)']
      r03 ['', 'State/UT', '<FY1>', '', '<FY2>', '', ...]   (period header, sparse)
      r04 ['', '', 'Gross', 'Net*', 'Gross', 'Net*', ...]  (sub-header)
      r05 ['', '1', '2', '3', '4', '5', ...]              (column-index row)
      r06+ data rows: ['', '<state>', g1, n1, g2, n2, ...]
    """
    if states is None:
        states = ["1. Andhra Pradesh", "2. Bihar"]
    if fiscal_years is None:
        fiscal_years = ["2023-24 (Accounts)", "2024-25 (Revised Estimates)"]
    if sub_labels is None:
        sub_labels = ["Gross", "Net*"]
    n_subs = len(sub_labels)
    if values_per_state is None:
        values_per_state = [
            [85423.7, 83109.5, 90290.6, 88250.4],
            [150401.3, 147896.9, 195737.7, 193389.9],
        ]

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.append([])
    ws.append(["", "Statement 17: Devolution and Transfer of Resources from the Centre"])
    ws.append(["", "(\u20b9 Crore)"])

    # Period header row: each fiscal-year cell, then (n_subs - 1) blanks
    period_row: list[Any] = ["", "State/UT"]
    for fy in fiscal_years:
        period_row.append(fy)
        period_row.extend([""] * (n_subs - 1))
    ws.append(period_row)

    # Sub-header row
    sub_row: list[Any] = ["", ""]
    for _ in fiscal_years:
        sub_row.extend(sub_labels)
    ws.append(sub_row)

    # Column-index row
    n_data_cols = len(fiscal_years) * n_subs
    idx_row: list[Any] = ["", "1"] + [str(i + 2) for i in range(n_data_cols)]
    ws.append(idx_row)

    # Data rows
    for state, vals in zip(states, values_per_state):
        ws.append(["", state, *vals])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_subcolumn_selects_net_under_each_period():
    """Statement-17 shape: each fiscal year stacks Gross + Net*; the
    spec's value_column_label='Net' must pick the Net column, not the
    Gross column directly under the year header."""
    wb_bytes = _make_subcolumn_workbook()
    spec = IndicatorSpec(
        indicator_id="fiscal/net_transfers_from_centre",
        sheet_match="ST_17",
        header_label_match="state/ut",
        period_kind="fy_span",
        value_column_label="Net",
    )
    parsed = parse_workbook(wb_bytes, spec)

    # 2 states × 2 fiscal years = 4 rows
    assert len(parsed.rows) == 4
    assert parsed.unmatched_states == []

    by_state: dict[str, list[tuple[str, float | None, str | None]]] = {}
    for r in parsed.rows:
        by_state.setdefault(r.entity_id, []).append((r.time, r.value, r.facet))

    # AP Net values are columns 3 and 5 of each data row -> 83109.5 and 88250.4.
    ap = sorted(by_state["S01"])
    assert ap == [
        ("2023-04", 83109.5, None),
        ("2024-04", 88250.4, "RE"),
    ]
    # Bihar Net values are 147896.9 and 193389.9.
    br = sorted(by_state["S04"])
    assert br == [
        ("2023-04", 147896.9, None),
        ("2024-04", 193389.9, "RE"),
    ]


def test_subcolumn_label_not_found_raises():
    wb_bytes = _make_subcolumn_workbook()
    spec = IndicatorSpec(
        indicator_id="fiscal/test",
        sheet_match="ST_17",
        header_label_match="state/ut",
        period_kind="fy_span",
        value_column_label="Nonexistent",
    )
    with pytest.raises(RBIWorkbookShapeError, match="value_column_label 'Nonexistent' not found"):
        parse_workbook(wb_bytes, spec)


def test_subcolumn_does_not_cross_period_boundary():
    """If the desired label only exists under one period (e.g.
    ``Variation`` columns at the right of ST_17), the parser must
    raise rather than silently borrow a column from a neighbouring
    period."""
    # Build a workbook where the FIRST period has Gross+Net but the
    # SECOND only has Gross — second period must fail.
    wb = Workbook()
    ws = wb.active
    ws.title = "ST_17"
    ws.append([])
    ws.append(["", "Title"])
    ws.append(["", "(unit)"])
    # period header: FY1 spans 2 cols (Gross+Net), FY2 spans 1 col (Gross only)
    ws.append(["", "State/UT", "2023-24", "", "2024-25"])
    ws.append(["", "", "Gross", "Net*", "Gross"])
    ws.append(["", "1. Andhra Pradesh", 1.0, 2.0, 3.0])
    buf = io.BytesIO()
    wb.save(buf)

    spec = IndicatorSpec(
        indicator_id="fiscal/test",
        sheet_match="ST_17",
        header_label_match="state/ut",
        period_kind="fy_span",
        value_column_label="Net",
    )
    with pytest.raises(RBIWorkbookShapeError, match="sub-header 'Net' not found under period '2024-04'"):
        parse_workbook(buf.getvalue(), spec)


def test_value_column_label_none_preserves_old_behaviour():
    """Existing Statement-20 (no sub-header) artifact must still parse
    identically when value_column_label is left None."""
    wb_bytes = _make_synthetic_workbook()
    parsed = parse_workbook(wb_bytes, _SPEC)
    assert _SPEC.value_column_label is None
    assert len(parsed.rows) == 6
