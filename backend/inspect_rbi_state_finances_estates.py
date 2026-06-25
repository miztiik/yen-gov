"""Inspector for the RBI State Finances (e-STATES Database) workbook.

The Reserve Bank of India publishes its *State Finances: A Study of Budgets*
compilation as one large long-format XLSX - the "e-STATES Database" (~13 MB).
It is a single ``Data`` sheet of ~400k rows shaped:

    Appendix | State/UT | Budget Head | Fiscal Year | Account | Revised | Budget

covering every state/UT across ~36 fiscal years (1990-91 onward), where
``Account`` = audited Actuals, ``Revised`` = Revised Estimate, ``Budget`` =
Budget Estimate.

This is an OPERATOR / developer INSPECTION tool: it reads a locally-staged
workbook (no network, no ``backend`` imports - openpyxl + stdlib only) and
prints its shape so an implementing agent can build the ingest reader and the
``(appendix, budget_head) -> indicator_id`` crosswalk. The supersede ruling,
indicator definitions, and crosswalk seed live in
``docs/architecture/backend/sources-rbi-state-finances.md``; the indicator
vocabulary in ``docs/concepts/state-fiscal-health.md``.

Commands::

    sheets                 list sheet names + dimensions
    note                   print the workbook's Note sheet verbatim
    dump --sheet S         dump the first rows of sheet S
    analyze                inventory appendices, states, budget heads, years
    verify                 reconcile the headline aggregates (sanity check)

Examples::

    python backend/inspect_rbi_state_finances_estates.py sheets   --path <file.xlsx>
    python backend/inspect_rbi_state_finances_estates.py analyze  --path <file.xlsx>
    python backend/inspect_rbi_state_finances_estates.py dump --sheet Data --path <file.xlsx>
"""
from __future__ import annotations

import argparse
import collections
from pathlib import Path

from openpyxl import load_workbook


def _open(path: Path):
    if not path.exists():
        raise SystemExit(f"workbook not found: {path}")
    return load_workbook(path, read_only=True, data_only=True)


def _fmt(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    return text if len(text) <= 28 else text[:25] + "..."


def list_sheets(path: Path) -> None:
    wb = _open(path)
    print(f"SHEETS: {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        ws = wb[name]
        try:
            dim = ws.calculate_dimension()
        except Exception as exc:  # noqa: BLE001 - inspection tool, report and move on
            dim = f"<err {exc}>"
        print(f"  - {name!r}  dim={dim}  max_row={ws.max_row} max_col={ws.max_column}")
    wb.close()


def dump_sheet(path: Path, sheet: str, nrows: int, ncols: int) -> None:
    wb = _open(path)
    if sheet not in wb.sheetnames:
        raise SystemExit(f"sheet {sheet!r} not in {wb.sheetnames}")
    ws = wb[sheet]
    print(f"=== {sheet!r}  max_row={ws.max_row} max_col={ws.max_column} ===")
    for i, row in enumerate(
        ws.iter_rows(min_row=1, max_row=nrows, max_col=ncols, values_only=True)
    ):
        print(f"R{i + 1:>3}: " + " | ".join(_fmt(c) for c in row))
    wb.close()


def full_note(path: Path) -> None:
    wb = _open(path)
    if "Note" not in wb.sheetnames:
        raise SystemExit("no 'Note' sheet in this workbook")
    ws = wb["Note"]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        print(f"NOTE R{i + 1}: {row[0]}")
    wb.close()


def analyze(path: Path) -> None:
    wb = _open(path)
    ws = wb["Data"]
    appendix: collections.Counter = collections.Counter()
    states: collections.Counter = collections.Counter()
    heads_by_app: dict[str, set[str]] = collections.defaultdict(set)
    years: set[str] = set()
    n_acc = n_rev = n_bud = total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        app, state, head, fy, acc, rev, bud = row[:7]
        total += 1
        appendix[app] += 1
        states[state] += 1
        if app is not None and head is not None:
            heads_by_app[str(app)].add(str(head))
        if fy is not None:
            years.add(str(fy))
        if acc not in (None, ""):
            n_acc += 1
        if rev not in (None, ""):
            n_rev += 1
        if bud not in (None, ""):
            n_bud += 1
    wb.close()

    print(f"TOTAL DATA ROWS: {total}")
    if years:
        print(f"FISCAL YEARS: {len(years)}  range {min(years)} .. {max(years)}")
    print(f"DISTINCT STATE/UT: {len(states)}")
    print(f"NON-EMPTY  Account={n_acc}  Revised={n_rev}  Budget={n_bud}")
    print("\nAPPENDIX (rows each):")
    for key, count in appendix.most_common():
        heads = len(heads_by_app[str(key)])
        print(f"  {key!r:>14}: {count:>7} rows  |  {heads} distinct budget heads")
    print("\nSTATES/UT:")
    for key, count in sorted(states.items()):
        print(f"  {key!r:>32}: {count}")
    for app in sorted(heads_by_app):
        print(f"\n=== BUDGET HEADS for {app!r} ({len(heads_by_app[app])}) ===")
        for head in sorted(heads_by_app[app]):
            print(f"   - {head}")


# Headline aggregate budget-head strings (verbatim) used by the sanity check.
_AGG_HEADS = (
    "I.A: State's Own Tax Revenue (1 to 3)",
    "I.B: Share in Central Taxes (i to ix)",
    "II.C: State's Own Non-Tax Revenue (1 to 6)",
    "II.D: Grants from the Centre (1 to 7)",
    "Total: TOTAL REVENUE (I+II)",
    "Total: TOTAL EXPENDITURE (I+II+III)",
    "II.C.2: Interest Payments (i to iv)",
    "II.E: Pensions",
    "A: Surplus (+)/Deficit (-) on Revenue Account",
    "C: Overall Surplus (+)/Deficit (-) (A+B)",
    "I: Total Capital Outlay (1 + 2)",
)


def verify(path: Path, entity: str = "All States/UT", year: str = "2022-2023") -> None:
    """Reconcile the revenue-deficit identity and report bifurcation coverage.

    Confirms that ``TOTAL REVENUE - TOTAL EXPENDITURE`` equals App-4's
    published "Surplus/Deficit on Revenue Account" line, and reports the
    Account coverage span for a back-projected (bifurcated) state.
    """
    wb = _open(path)
    ws = wb["Data"]
    coverage: dict[str, list[str]] = collections.defaultdict(list)
    snap: dict[str, object] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        _app, state, head, fy, acc, _rev, _bud = row[:7]
        if head == "Total: TOTAL REVENUE (I+II)" and acc not in (None, ""):
            if state in ("Telangana", "Andhra Pradesh"):
                coverage[str(state)].append(str(fy))
        if state == entity and str(fy) == year and head in _AGG_HEADS:
            snap[str(head)] = acc
    wb.close()

    for state, fys in sorted(coverage.items()):
        print(f"{state} TOTAL REVENUE (Account) coverage: {min(fys)} .. {max(fys)} ({len(fys)} yrs)")
    print(f"\n{entity} {year} (Account, INR crore):")
    for head in _AGG_HEADS:
        print(f"  {snap.get(head, 'MISSING')!s:>18}  <-  {head}")
    tr = snap.get("Total: TOTAL REVENUE (I+II)")
    te = snap.get("Total: TOTAL EXPENDITURE (I+II+III)")
    ra = snap.get("A: Surplus (+)/Deficit (-) on Revenue Account")
    if isinstance(tr, (int, float)) and isinstance(te, (int, float)):
        print(f"\nCHECK revenue surplus/deficit: TOTAL REVENUE - TOTAL EXP = {tr - te:.2f}  vs App-4 'A' = {ra}")


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(
        description="Inspect the RBI State Finances (e-STATES Database) workbook.",
    )
    parser.add_argument(
        "command",
        choices=("sheets", "note", "dump", "analyze", "verify"),
        help="what to inspect",
    )
    parser.add_argument(
        "--path",
        required=True,
        type=Path,
        help="path to the locally-staged e-STATES .XLSX workbook",
    )
    parser.add_argument("--sheet", default="Data", help="sheet name for 'dump'")
    parser.add_argument("--rows", type=int, default=12, help="rows for 'dump'")
    parser.add_argument("--cols", type=int, default=16, help="columns for 'dump'")
    args = parser.parse_args(argv)

    if args.command == "sheets":
        list_sheets(args.path)
    elif args.command == "note":
        full_note(args.path)
    elif args.command == "dump":
        dump_sheet(args.path, args.sheet, args.rows, args.cols)
    elif args.command == "analyze":
        analyze(args.path)
    elif args.command == "verify":
        verify(args.path)


if __name__ == "__main__":
    main()
